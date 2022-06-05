import unittest
from typing import Union
import yaml
import jk_commentjson as jsonc
import matplotlib.pyplot as plt
import pandas as pd
import xlsxwriter
from openpyxl import formatting, styles, Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import webbrowser
import requests
import io
import os
import glob
import re
import getopt
import sys
import base64
import unidecode
from termcolor     import colored
from mako.template import Template
import mako.runtime
import warnings
import zipfile
import shutil
from importlib import reload
from importlib import reload
import matplotlib
from scipy.interpolate import make_interp_spline, BSpline
import numpy as np
import ftplib
import markdown
import logging
import datetime
from dateutil import parser

timestamp = datetime.datetime.now().strftime("%y%m%d-%H%M%S")
logFile   = "output"+os.sep+"_ConsommationFonciere_"+timestamp+".log"
logging.basicConfig(filename=logFile, filemode='w', format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.DEBUG)

reload(matplotlib)
matplotlib.use('Agg')

warnings.filterwarnings("ignore")
warnings.simplefilter(action='ignore', category=pd.errors.PerformanceWarning)


from warnings import simplefilter
simplefilter(action="ignore", category=pd.errors.PerformanceWarning)
pd.options.mode.chained_assignment = None

# Working Directories
output_dir = "output" + os.sep
data_dir   = "data"   + os.sep
input_dir  = "input"  + os.sep

configurationFile    = input_dir  + "Configuration.xlsx"
html_report_template = input_dir  + "report_template.html"
html_index_template  = input_dir  + "index_template.html"
plots_file           = input_dir  + "plots.json"
context_file         = output_dir + "context.yaml"
france_file          = output_dir + "france.json"
selection_file       = output_dir + "select.json"

global_context     = {}

# DataFile and Data Frames

##########################
### Sitadel Logements
##########################

sitadelSourcePage     = "https://www.data.gouv.fr/fr/datasets/base-des-permis-de-construire-et-autres-autorisations-durbanisme-sitadel"
global_context["URL_SOURCE_SITADEL"] = sitadelSourcePage

sitadelSource1316File = "https://www.data.gouv.fr/fr/datasets/r/67dd4ee1-0d73-4676-a90f-854fe9012f5d"
sitadelSource1721File = "https://www.data.gouv.fr/fr/datasets/r/1fa467ef-5e3a-456f-b961-be9032cfa3df"
sitadelSourceMetaFile = "https://www.data.gouv.fr/fr/datasets/r/9d7d6728-c3bc-44e4-8105-7335ad70d52e"
sitadel1316File       = data_dir + "PC_DP_creant_logements_2013_2016.csv"
sitadel1721File       = data_dir + "PC_DP_creant_logements_2017_2022.csv"
sitadelMetaFile       = data_dir + "dictionnaire_variables_logements_permis_construire.xls"


sitadel1316 = None
sitadel1721 = None
sitadelMeta = None
sitadel1321 = None

# Logements
# "REG";"DEP";"COMM";"Type_DAU";"Num_DAU";"Etat_DAU";"DATE_REELLE_AUTORISATION";"DATE_REELLE_DOC";"DATE_REELLE_DAACT";"DPC_AUT";"DPC_DOC";"DPC_DERN";"APE_DEM";"CJ_DEM";"DENOM_DEM";"SIREN_DEM";"SIRET_DEM";"CODPOST_DEM";"LOCALITE_DEM";"REC_ARCHI";"ADR_NUM_TER";"ADR_TYPEVOIE_TER";"ADR_LIBVOIE_TER";"ADR_LIEUDIT_TER";"ADR_LOCALITE_TER";"ADR_CODPOST_TER";"sec_cadastre1";"num_cadastre1";"sec_cadastre2";"num_cadastre2";"sec_cadastre3";"num_cadastre3";"SUPERFICIE_TERRAIN";"ZONE_OP";"NATURE_PROJET";"I_EXTENSION";"I_SURELEVATION";"I_NIVSUPP";"NB_NIV_MAX";"NB_CHAMBRES";"SURF_HAB_AVANT";"SURF_HAB_CREEE";"SURF_HAB_ISSUE_TRANSFO";"SURF_HAB_DEMOLIE";"SURF_HAB_TRANSFORMEE";"SURF_LOC_AVANT";"SURF_LOC_CREEE";"SURF_LOC_ISSUE_TRANSFO";"SURF_LOC_DEMOLIE";"SURF_LOC_TRANSFORMEE";"SURF_HEB_AVANT";"SURF_HEB_CREEE";"SURF_HEB_ISSUE_TRANSFO";"SURF_HEB_DEMOLIE";"SURF_HEB_TRANSFORMEE";"SURF_BUR_AVANT";"SURF_BUR_CREEE";"SURF_BUR_ISSUE_TRANSFO";"SURF_BUR_DEMOLIE";"SURF_BUR_TRANSFORMEE";"SURF_COM_AVANT";"SURF_COM_CREEE";"SURF_COM_ISSUE_TRANSFO";"SURF_COM_DEMOLIE";"SURF_COM_TRANSFORMEE";"SURF_ART_AVANT";"SURF_ART_CREEE";"SURF_ART_ISSUE_TRANSFO";"SURF_ART_DEMOLIE";"SURF_ART_TRANSFORMEE";"SURF_IND_AVANT";"SURF_IND_CREEE";"SURF_IND_ISSUE_TRANSFO";"SURF_IND_DEMOLIE";"SURF_IND_TRANSFORMEE";"SURF_AGR_AVANT";"SURF_AGR_CREEE";"SURF_AGR_ISSUE_TRANSFO";"SURF_AGR_DEMOLIE";"SURF_AGR_TRANSFORMEE";"SURF_ENT_AVANT";"SURF_ENT_CREEE";"SURF_ENT_ISSUE_TRANSFO";"SURF_ENT_DEMOLIE";"SURF_ENT_TRANSFORMEE";"SURF_PUB_AVANT";"SURF_PUB_CREEE";"SURF_PUB_ISSUE_TRANSFO";"SURF_PUB_DEMOLIE";"SURF_PUB_TRANSFORMEE";"TYPE_SERVICE_PUBLIC"


def load_sitadel(sitadel1316_file:  str = sitadel1316File,
                 sitadel1721_file:  str = sitadel1721File,
                 sitadel_meta_file: str = sitadelMetaFile):
    global sitadel1316, sitadel1721, sitadelMeta, sitadel1321
    if (sitadel1316 is None) or (sitadel1721 is None)  or (sitadelMeta is None):
        downloadFile(sitadelSource1316File, sitadel1316File, zip=True, zipped_file="PC_DP_creant_logements_2013_2016.csv")
        downloadFile(sitadelSource1721File, sitadel1721File, zip=True, zipped_file="PC_DP_creant_logements_2017_2022.csv")
        downloadFile(sitadelSourceMetaFile, sitadelMetaFile)
        print_blue("Lecture Sitadel Logements 2013-2016 : " + sitadel1316_file + " ...")
        sitadel1316 = pd.read_csv(sitadel1316_file, delimiter=';', index_col=4, encoding='latin-1', dtype={"DEP": str, "COMM": str, "Etat_DAU": str, "DPC_AUT": str, "NATURE_PROJET_DECLAREE" : str, "I_EXTENSION": str, "I_SURELEVATION": str, "I_NIVSUPP": str, "Type_DAU": str, "Num_DAU": str})
        sitadel1316["Parcelles"] = sitadel1316['sec_cadastre1'].map(str) + sitadel1316['num_cadastre1'].map(str)+" "+\
                                   sitadel1316['sec_cadastre2'].map(str) + sitadel1316['num_cadastre2'].map(str)+" "+\
                                   sitadel1316['sec_cadastre3'].map(str) + sitadel1316['num_cadastre3'].map(str)
        sitadel1316["NATURE_PROJET"]     = sitadel1316["NATURE_PROJET_DECLAREE"]
        sitadel1316["RESIDENCE_SERVICE"] = sitadel1316["RESIDENCE"]
        print_blue("Lecture Sitadel Logements 2017-2022 : " + sitadel1721_file + " ...")
        sitadel1721 = pd.read_csv(sitadel1721_file, delimiter=';', index_col=4, encoding='latin-1', dtype={"DEP": str, "COMM": str, "Etat_DAU": str, "DPC_AUT": str, "ADR_LOCALITE_TER" : str, "ADR_CODPOST_TER" : str, "NATURE_PROJET_DECLAREE" : str, "I_EXTENSION": str, "I_SURELEVATION": str, "I_NIVSUPP": str, "Type_DAU": str, "Num_DAU": str})
        sitadel1721["Parcelles"] = sitadel1721['sec_cadastre1'].map(str) + sitadel1721['num_cadastre1'].map(str)+" "+\
                                   sitadel1721['sec_cadastre2'].map(str) + sitadel1721['num_cadastre2'].map(str)+" "+\
                                   sitadel1721['sec_cadastre3'].map(str) + sitadel1721['num_cadastre3'].map(str)
        sitadel1721["NATURE_PROJET"]     = sitadel1721["NATURE_PROJET_DECLAREE"]
        sitadel1721["RESIDENCE_SERVICE"] = sitadel1721["RESIDENCE"]
        print_blue("Lecture Meta Logements Sitadel : " + sitadel_meta_file + " ...")
        sitadel1321 = pd.concat([sitadel1316, sitadel1721])
        xls = pd.ExcelFile(sitadel_meta_file)
        sitadelMeta = pd.read_excel(xls, 'Variables_Logements', index_col=0)
    return sitadel1316, sitadel1721, sitadelMeta, sitadel1321


##########################
### Sitadel Locaux
##########################

sitadelLocauxSource1316File = "https://www.data.gouv.fr/fr/datasets/r/3b987380-d1cf-4047-8dc5-1a19a3ecf812"
sitadelLocauxSource1721File = "https://www.data.gouv.fr/fr/datasets/r/98ff9fd3-a14e-474d-bb8f-12bde12d9f70"
sitadelLocauxSourceMetaFile = "https://www.data.gouv.fr/fr/datasets/r/b3ffee5b-fd75-4345-a086-02ded2018705"
sitadelLocaux1316File = data_dir + "PC_DP_creant_locaux_2013_2016.csv"
sitadelLocaux1721File = data_dir + "PC_DP_creant_locaux_2017_2022.csv"
sitadelLocauxMetaFile = data_dir + "dictionnaire_variables_locaux_permis_construire.xls"

sitadel_locaux_1316 = None
sitadel_locaux_1721 = None
sitadel_locaux_Meta = None
sitadel_locaux_1321 = None

# Locaux
# "REG";"DEP";"COMM";"Type_DAU";"Num_DAU";"Etat_DAU";"DATE_REELLE_AUTORISATION";"DATE_REELLE_DOC";"DATE_REELLE_DAACT";"DPC_AUT";"DPC_DOC";"DPC_DERN";"APE_DEM";"CJ_DEM";"DENOM_DEM";"SIREN_DEM";"SIRET_DEM";"CODPOST_DEM";"LOCALITE_DEM";"REC_ARCHI";"ADR_NUM_TER";"ADR_TYPEVOIE_TER";"ADR_LIBVOIE_TER";"ADR_LIEUDIT_TER";"ADR_LOCALITE_TER";"ADR_CODPOST_TER";"sec_cadastre1";"num_cadastre1";"sec_cadastre2";"num_cadastre2";"sec_cadastre3";"num_cadastre3";"SUPERFICIE_TERRAIN";"ZONE_OP";"NATURE_PROJET";"I_EXTENSION";"I_SURELEVATION";"I_NIVSUPP";"NB_NIV_MAX";"NB_CHAMBRES";"SURF_HAB_AVANT";"SURF_HAB_CREEE";"SURF_HAB_ISSUE_TRANSFO";"SURF_HAB_DEMOLIE";"SURF_HAB_TRANSFORMEE";"SURF_LOC_AVANT";"SURF_LOC_CREEE";"SURF_LOC_ISSUE_TRANSFO";"SURF_LOC_DEMOLIE";"SURF_LOC_TRANSFORMEE";"SURF_HEB_AVANT";"SURF_HEB_CREEE";"SURF_HEB_ISSUE_TRANSFO";"SURF_HEB_DEMOLIE";"SURF_HEB_TRANSFORMEE";"SURF_BUR_AVANT";"SURF_BUR_CREEE";"SURF_BUR_ISSUE_TRANSFO";"SURF_BUR_DEMOLIE";"SURF_BUR_TRANSFORMEE";"SURF_COM_AVANT";"SURF_COM_CREEE";"SURF_COM_ISSUE_TRANSFO";"SURF_COM_DEMOLIE";"SURF_COM_TRANSFORMEE";"SURF_ART_AVANT";"SURF_ART_CREEE";"SURF_ART_ISSUE_TRANSFO";"SURF_ART_DEMOLIE";"SURF_ART_TRANSFORMEE";"SURF_IND_AVANT";"SURF_IND_CREEE";"SURF_IND_ISSUE_TRANSFO";"SURF_IND_DEMOLIE";"SURF_IND_TRANSFORMEE";"SURF_AGR_AVANT";"SURF_AGR_CREEE";"SURF_AGR_ISSUE_TRANSFO";"SURF_AGR_DEMOLIE";"SURF_AGR_TRANSFORMEE";"SURF_ENT_AVANT";"SURF_ENT_CREEE";"SURF_ENT_ISSUE_TRANSFO";"SURF_ENT_DEMOLIE";"SURF_ENT_TRANSFORMEE";"SURF_PUB_AVANT";"SURF_PUB_CREEE";"SURF_PUB_ISSUE_TRANSFO";"SURF_PUB_DEMOLIE";"SURF_PUB_TRANSFORMEE";"TYPE_SERVICE_PUBLIC"


def load_sitadel_locaux(sitadelLocaux1316_file:  str = sitadelLocaux1316File,
                        sitadelLocaux1721_file:  str = sitadelLocaux1721File,
                        sitadelLocaux_meta_file: str = sitadelLocauxMetaFile):
    global sitadel_locaux_1316, sitadel_locaux_1721, sitadel_locaux_Meta, sitadel_locaux_1321
    if (sitadel_locaux_1316 is None) or (sitadel_locaux_1721 is None)  or (sitadel_locaux_Meta is None):
        downloadFile(sitadelLocauxSource1316File, sitadelLocaux1316File, zip=True, zipped_file="PC_DP_creant_locaux_2013_2016.csv")
        downloadFile(sitadelLocauxSource1721File, sitadelLocaux1721File, zip=True, zipped_file="PC_DP_creant_locaux_2017_2022.csv")
        downloadFile(sitadelLocauxSourceMetaFile, sitadelLocauxMetaFile)
        print_blue("Lecture Sitadel Locaux 2013-2016 : " + sitadelLocaux1316_file + " ...")
        sitadel_locaux_1316 = pd.read_csv(sitadelLocaux1316_file, delimiter=';', index_col=4, encoding='latin-1', dtype={"DEP": str, "COMM": str, "DPC_AUT": str, "Etat_DAU": str,"ZONE_OP": str, "NATURE_PROJET": str, "I_EXTENSION": str, "I_SURELEVATION": str, "I_NIVSUPP": str, "SUPERFICIE_TERRAIN": float, "SURF_HAB_AVANT": float})
        sitadel_locaux_1316["Parcelles"] = sitadel_locaux_1316['sec_cadastre1'].map(str) + sitadel_locaux_1316['num_cadastre1'].map(str)+" "+\
                                           sitadel_locaux_1316['sec_cadastre2'].map(str) + sitadel_locaux_1316['num_cadastre2'].map(str)+" "+\
                                           sitadel_locaux_1316['sec_cadastre3'].map(str) + sitadel_locaux_1316['num_cadastre3'].map(str)
        sitadel_locaux_1316["NATURE_PROJET"]     = sitadel_locaux_1316["NATURE_PROJET_DECLAREE"]
        print_blue("Lecture Sitadel Locaux 2017-2022 : " + sitadelLocaux1721_file + " ...")
        sitadel_locaux_1721 = pd.read_csv(sitadelLocaux1721_file, delimiter=';', index_col=4, encoding='latin-1', dtype={"DEP": str, "COMM": str, "Etat_DAU": str, "DPC_AUT": str, "ADR_LOCALITE_TER" : str, "ADR_CODPOST_TER" : str, "NATURE_PROJET" : str, "ZONE_OP": str, "I_EXTENSION": str, "I_SURELEVATION": str, "I_NIVSUPP": str, "SUPERFICIE_TERRAIN": float, "SURF_HAB_AVANT": float})
        sitadel_locaux_1721["Parcelles"] = sitadel_locaux_1721['sec_cadastre1'].map(str) + sitadel_locaux_1721['num_cadastre1'].map(str)+" "+\
                                           sitadel_locaux_1721['sec_cadastre2'].map(str) + sitadel_locaux_1721['num_cadastre2'].map(str)+" "+\
                                           sitadel_locaux_1721['sec_cadastre3'].map(str) + sitadel_locaux_1721['num_cadastre3'].map(str)
        sitadel_locaux_1721["NATURE_PROJET"]     = sitadel_locaux_1721["NATURE_PROJET_DECLAREE"]
        sitadel_locaux_1321 = pd.concat([sitadel_locaux_1316, sitadel_locaux_1721])
        print_blue("Lecture Meta Locaux Sitadel : " + sitadelLocaux_meta_file + " ...")
        xls = pd.ExcelFile(sitadelLocaux_meta_file)
        sitadel_locaux_Meta = pd.read_excel(xls, 'Variables_Locaux', index_col=0)
    return sitadel_locaux_1316, sitadel_locaux_1721, sitadel_locaux_Meta, sitadel_locaux_1321

############################
### Logements Paca 2010-2019
############################

sitadel1019SourcePage = "https://www.insee.fr/fr/statistiques/5395856?sommaire=5395912"
sitadel1019SourceFile = data_dir + "logements_commences_PACA_2010-2019.xls"

sitadel1019 = None


def load_logements_paca(sitadel1019S:  str = sitadel1019SourceFile):
    global sitadel1019
    if (sitadel1019 is None) :
        print_blue("Lecture Logements Paca 2010-2019 : " + sitadel1019SourceFile + " ...")
        xls = pd.ExcelFile(sitadel1019SourceFile)
        sitadel1019 = pd.read_excel(xls, 'logements_commences_PACA_1019')
    return sitadel1019

##########################
### Evolution 2008-2021
##########################

evolutionSourcePage = "https://www.insee.fr/fr/statistiques/1893198"
evolutionSourceFile = "https://www.insee.fr/fr/statistiques/fichier/1893198/evolution-population-reg-2008-2021.xlsx"
evolutionFile = data_dir + "evolution-population-dep-2008-2021.xlsx"
evolution0813 = None
evolution1318 = None
evolution1821 = None
global_context["URL_SOURCE_EVOLUTION"] = evolutionSourcePage

# Variation relative annuelle 2018-2021 (en %)
# Département, Estimations de population au 1er janvier 2021, Totale Due au solde naturel, Due au solde apparent des entrées et des sorties
# Ain	662,244	    0.7%	0.3%	0.4%


def load_evolution(evolution_file: str = evolutionFile):
    global evolution0813, evolution1318, evolution1821
    if (evolution0813 is None) or (evolution1318 is None)  or (evolution1821 is None):
        print_blue("Lecture Evolution Dept : " + evolution_file + " ...")
        xls = pd.ExcelFile(evolution_file)
        evolution0813 = pd.read_excel(xls, '2008-2013', index_col=0, dtype={"Unnamed: 0": str})
        evolution1318 = pd.read_excel(xls, '2013-2018', index_col=0, dtype={"Unnamed: 0": str})
        xls = pd.ExcelFile(evolution_file)
        evolution1821 = pd.read_excel(xls, '2018-2021', index_col=0, dtype={"Unnamed: 0": str})
    return evolution0813, evolution1318, evolution1821

##########################
### Projections 2013-2050
##########################

projectionsSourcePage = "https://www.insee.fr/fr/statistiques/2859843"
projectionsSourceFile = "https://www.insee.fr/fr/statistiques/fichier/2859843/projections_scenario_central.xls"
projectionsFile = data_dir + "projections_scenario_central.xls"
projectionsDPT = None
projectionsREG = None
global_context["URL_SOURCE_PROJECTIONS_2050"] = projectionsSourcePage

### Population_DEP
# Département       Libellé du département	Population en 2013	Population en 2014	Population en 2015	Population en 2016	Population en 2017	Population en 2018	Population en 2019	Population en 2020	Population en 2021	Population en 2022	Population en 2023	Population en 2024	Population en 2025	Population en 2026	Population en 2027	Population en 2028	Population en 2029	Population en 2030	Population en 2031	Population en 2032	Population en 2033	Population en 2034	Population en 2035	Population en 2036	Population en 2037	Population en 2038	Population en 2039	Population en 2040	Population en 2041	Population en 2042	Population en 2043	Population en 2044	Population en 2045	Population en 2046	Population en 2047	Population en 2048	Population en 2049	Population en 2050
# code_Departements	libelle_Departements	pop_2013	pop_2014	pop_2015	pop_2016	pop_2017	pop_2018	pop_2019	pop_2020	pop_2021	pop_2022	pop_2023	pop_2024	pop_2025	pop_2026	pop_2027	pop_2028	pop_2029	pop_2030	pop_2031	pop_2032	pop_2033	pop_2034	pop_2035	pop_2036	pop_2037	pop_2038	pop_2039	pop_2040	pop_2041	pop_2042	pop_2043	pop_2044	pop_2045	pop_2046	pop_2047	pop_2048	pop_2049	pop_2050

### Population_REG
# Région	Libellé de la région	Population en 2013	Population en 2014	Population en 2015	Population en 2016	Population en 2017	Population en 2018	Population en 2019	Population en 2020	Population en 2021	Population en 2022	Population en 2023	Population en 2024	Population en 2025	Population en 2026	Population en 2027	Population en 2028	Population en 2029	Population en 2030	Population en 2031	Population en 2032	Population en 2033	Population en 2034	Population en 2035	Population en 2036	Population en 2037	Population en 2038	Population en 2039	Population en 2040	Population en 2041	Population en 2042	Population en 2043	Population en 2044	Population en 2045	Population en 2046	Population en 2047	Population en 2048	Population en 2049	Population en 2050
# code_Regions	libelle_Regions	pop_2013	pop_2014	pop_2015	pop_2016	pop_2017	pop_2018	pop_2019	pop_2020	pop_2021	pop_2022	pop_2023	pop_2024	pop_2025	pop_2026	pop_2027	pop_2028	pop_2029	pop_2030	pop_2031	pop_2032	pop_2033	pop_2034	pop_2035	pop_2036	pop_2037	pop_2038	pop_2039	pop_2040	pop_2041	pop_2042	pop_2043	pop_2044	pop_2045	pop_2046	pop_2047	pop_2048	pop_2049	pop_2050


def load_projections(projections_file: str = projectionsFile):
    global projectionsREG, projectionsDPT
    if (projectionsREG is None) or (projectionsREG is None):
        print_blue("Lecture Projections Dept/Reg : " + projections_file + " ...")
        xls = pd.ExcelFile(projections_file)
        projectionsDPT = pd.read_excel(xls, 'Population_DEP', index_col=0, dtype={"Unnamed: 0": str, "Unnamed: 1": str})
        projectionsREG = pd.read_excel(xls, 'Population_REG', index_col=0, dtype={"Unnamed: 0": str, "Unnamed: 1": str})
    return projectionsDPT, projectionsREG


##############################
### Projections PACA 2030-2050
##############################

projectionsPacaSourcePage = "https://www.insee.fr/fr/statistiques/3202958?sommaire=3203271"
projectionsPacaSourceFile = "https://www.insee.fr/fr/statistiques/fichier/3202958/1_Population_evolutions.xls"
projectionsPacaFile = data_dir + "1_Population_evolutions.xls"
projectionsPaca = None
global_context["URL_SOURCE_PROJECTIONS_PACA"] = projectionsPacaSourcePage

# Type de Zone  EPCI	Nom de zone		Population haute 2013	Central* 2013	Population basse 2013	Sans migrations 2013	Population haute 2030	Central* 2030	Population basse 2030	Sans migrations 2030 Population haute 2050	Central* 2050	Population basse 2050	Sans migrations 2050


def load_projections_paca(projections_paca_file: str = projectionsPacaFile):
    global projectionsPaca
    if (projectionsPaca is None) :
        print_blue("Lecture Projections Paca : " + projections_paca_file + " ...")
        xls = pd.ExcelFile(projections_paca_file)
        projectionsPaca = pd.read_excel(xls, 'Projection', dtype={'["Unnamed: 0"]': str, '["Unnamed: 1"]': str, '["Unnamed: 2"]': str})
    return projectionsPaca


################
### Departements
################

departementsSourcePage = "https://www.data.gouv.fr/en/datasets/departements-de-france/"
departementsSourceFile = "https://www.data.gouv.fr/en/datasets/r/70cef74f-70b1-495a-8500-c089229c0254"
departementsFile = data_dir + "departements-france.csv"
departements     = None
global_context["URL_SOURCE_DEPARTEMENTS"] = departementsSourcePage

# code_departement	nom_departement	code_region	nom_region


def load_departements(dpt_file: str = departementsFile):
    global departements
    if (departements is None):
        print_blue("Lecture Departements : " + dpt_file + " ...")
        departements = pd.read_csv(dpt_file, delimiter=',', index_col=0, dtype={'code_departement': str, 'nom_departement': str, 'code_region': str, 'nom_region': str})
    return departements

####################
### Intercommunalite
####################

interCoSourcePage = "https://www.insee.fr/fr/information/2510634"
interCoSourceFile = "https://www.insee.fr/fr/statistiques/fichier/2510634/Intercommunalite_Metropole_au_01-01-2021.zip"
intercoFile = data_dir + "Intercommunalite-Metropole_au_01-01-2021.xlsx"
intercoDossier = None
intercoEPCI    = None
global_context["URL_SOURCE_INTERCOMMUNALITES"] = interCoSourcePage

# CODGEO	LIBGEO	EPCI	LIBEPCI	DEP	REG


def load_interco(interco_file: str = intercoFile):
    # EPCI		    EPCI - Métropole		        Code géographique de l'établissement public à fiscalité propre ou métropole
    # LIBEPCI		Libellé de l'EPCI / Métropole	Libellé de l'EPCI ou métropole
    # NATURE_EPCI	Nature d'EPCI			        Nature d'établissement public
    # NB_COM		Nombre communes			        Nombre de communes contenues dans l'objet géographique
    # CODGEO		Code géographique		        Code géographique
    # LIBGEO		Libellé géographique	        Libellé géographique
    # DEP		    Département			            Code géographique du département
    # REG		    Région				            Code géographique de la région
    global intercoDossier, intercoEPCI
    if (intercoDossier is None) or (intercoEPCI is None):
        print_blue("Lecture Donnees Intercommunalite : " + interco_file + " ...")
        xls = pd.ExcelFile(interco_file)
        intercoDossier = pd.read_excel(xls, 'Composition_communale', index_col=0, dtype={"EPCI": str, "LIBEPCI": str,  "NATURE_EPCI": str, "NB_COM": int, "CODGEO": str, "LIBGEO": str, "DEP": str, "REG": str})
        intercoEPCI    = pd.read_excel(xls, 'EPCI', index_col=0)
    return intercoDossier, intercoEPCI

####################
### SCoT
####################

SCoT_File = data_dir + "Table_EPCI_SCoT.xls"
SCOT_DATA = None
COMMUNES_HORS_PACA = None
EPCI_SCOT          = None
GROUPEMENTS        = None

def load_scot_data(SCoT_File: str = SCoT_File):
    global SCOT_DATA, COMMUNES_HORS_PACA, EPCI_SCOT, GROUPEMENTS
    if (SCOT_DATA is None):
        SCOT_DATA = {}
        print_blue("Lecture Donnees SCoT : " + SCoT_File + " ...")
        xls = pd.ExcelFile(SCoT_File)
        COMMUNES_HORS_PACA = pd.read_excel(xls, 'Commmunes hors PACA', dtype={"EPCI": str, "LIBEPCI": str,  "Commune": str,  "CODGEO": str })
        EPCI_SCOT          = pd.read_excel(xls, 'EPCI - SCoT', dtype={"SCoT": str, "LIBEPCI": str})
        GROUPEMENTS        = pd.read_excel(xls, 'Groupements', dtype={"Type": str, "DEPT": str, "EPCI": str, "REG": str, "NOM": str, "Territoires": str})
        SCOT_DATA["SCOT_EPCI"] = {}
        for index, row in EPCI_SCOT.iterrows():
            if (index == 0) : continue
            EPCI = row['Table de passage EPCI --> SCoT']
            SCOT = row['Unnamed: 2']
            if (SCOT == '////') : SCOT = "SCoT " + EPCI
            if (SCOT not in SCOT_DATA["SCOT_EPCI"]) : SCOT_DATA["SCOT_EPCI"][SCOT] = []
            SCOT_DATA["SCOT_EPCI"][SCOT].append(EPCI)
        SCOT_DATA["EPCI_AVEC_COMM_HORS_PACA"] = {}
        for index, row in COMMUNES_HORS_PACA.iterrows():
            if (index == 0) : continue
            COMM = row['Liste des communes hors PACA appartenant à des EPCI de PACA']
            EPCI = row['Unnamed: 3']
            if (EPCI not in SCOT_DATA["EPCI_AVEC_COMM_HORS_PACA"]): SCOT_DATA["EPCI_AVEC_COMM_HORS_PACA"][EPCI] = []
            SCOT_DATA["EPCI_AVEC_COMM_HORS_PACA"][EPCI].append(COMM)
        SCOT_DATA["GROUPEMENTS"] = {}
        for index, row in GROUPEMENTS.iterrows():
            if (str(row['Type']).startswith("#")) : continue
            NOM  = row['Nom']
            if (NOM not in SCOT_DATA["GROUPEMENTS"]): SCOT_DATA["GROUPEMENTS"][NOM] = {}
            SCOT_DATA["GROUPEMENTS"][NOM]["TYPE"] = row['Type']
            SCOT_DATA["GROUPEMENTS"][NOM]["NOM"]  = row['Nom']
            SCOT_DATA["GROUPEMENTS"][NOM]["LIST"] = row['Territoires']
    return SCOT_DATA

#################
### Codes Postaux
#################

codesPostauxSourcePage = "https://datanova.laposte.fr/explore/dataset/laposte_hexasmal/information/?disjunctive.code_commune_insee&disjunctive.nom_de_la_commune&disjunctive.code_postal&disjunctive.ligne_5"
codesPostauxSourceFile = "https://datanova.laposte.fr/explore/dataset/laposte_hexasmal/download/?format=csv&timezone=Europe/Berlin&lang=fr&use_labels_for_header=true&csv_separator=%3B"
codesPostauxFile = data_dir + "laposte_hexasmal.csv"
codesPostaux = None
global_context["URL_SOURCE_CODES_POSTAUX"] = codesPostauxSourcePage

# Code_commune_INSEE;Nom_commune;Code_postal;Ligne_5;LibellÃ©_d_acheminement;coordonnees_gps (lat,long)
# 02547;LA NEUVILLE HOUSSET;02250;;LA NEUVILLE HOUSSET;49.7881379377,3.731716273


def load_codes():
    global codesPostaux
    if (codesPostaux is None) :
        downloadFile(codesPostauxSourceFile, codesPostauxFile)
        print_blue("Lecture Codes Postaux Donnees Communes : " + codesPostauxFile + " ...")
        codesPostaux = pd.read_csv(codesPostauxFile, delimiter=';', index_col=0, dtype=str)
    return codesPostaux

#################
### SRU
#################

sru2017SourcePage = "http://www.paca.developpement-durable.gouv.fr/periode-triennale-2017-2019-a10879.html"
sru2017SourceFile = "http://www.paca.developpement-durable.gouv.fr/IMG/pdf/2017-2019_communes_sru_en_paca.pdf"
sru2020SourcePage = "http://www.paca.developpement-durable.gouv.fr/periode-triennale-2020-2022-a13129.html"
sru2020SourceFile = "http://www.paca.developpement-durable.gouv.fr/IMG/pdf/inventaire_010120.pdf"
sruFile = data_dir + "communes_sru_en_paca.xlsx"
sru2017 = None
sru2020 = None
global_context["URL_SOURCE_SRU"] = sru2020SourcePage

# REG;"DEP";"COMM";"Type_DAU";"Num_DAU";"Etat_DAU";"DATE_REELLE_AUTORISATION";"DATE_REELLE_DOC";"DATE_REELLE_DAACT";"DPC_AUT";"DPC_DOC";"DPC_DERN";"CAT_DEM";"APE_DEM";"CJ_DEM";"DENOM_DEM";"SIREN_DEM";"SIRET_DEM";"CODPOST_DEM";"LOCALITE_DEM";"REC_ARCHI";"ADR_NUM_TER";"ADR_TYPEVOIE_TER";"ADR_LIBVOIE_TER";"ADR_LIEUDIT_TER";"ADR_LOCALITE_TER";"ADR_CODPOST_TER";"sec_cadastre1";"num_cadastre1";"sec_cadastre2";"num_cadastre2";"sec_cadastre3";"num_cadastre3";"SUPERFICIE_TERRAIN";"ZONE_OP";"NATURE_PROJET";"I_EXTENSION";"I_SURELEVATION";"I_NIVSUPP";"NB_NIV_MAX";"UTILISATION";"RES_PRINCIP_OU_SECOND";"TYP_ANNEXE";"RESIDENCE_SERVICE";"NB_LGT_TOT_CREES";"NB_LGT_IND_CREES";"NB_LGT_COL_CREES";"NB_LGT_DEMOLIS";"NB_LGT_1P";"NB_LGT_2P";"NB_LGT_3P";"NB_LGT_4P";"NB_LGT_5P";"NB_LGT_6P_PLUS";"NB_LGT_PRET_LOC_SOCIAL";"NB_LGT_ACC_SOC_HORS_PTZ";"NB_LGT_PTZ";"SURF_HAB_AVANT";"SURF_HAB_CREEE";"SURF_HAB_ISSUE_TRANSFO";"SURF_HAB_DEMOLIE";"SURF_HAB_TRANSFORMEE";"SURF_LOC_AVANT";"SURF_LOC_CREEE";"SURF_LOC_ISSUE_TRANSFO";"SURF_LOC_DEMOLIE";"SURF_LOC_TRANSFORMEE"


def load_sru(sru_file: str = sruFile):
    global sru2017, sru2020
    if (sru2017 is None) or (sru2020 is None):
        print_blue("Lecture Carences SRU PACA : " + sru_file + " ...")
        xls = pd.ExcelFile(sru_file)
        sru2017 = pd.read_excel(xls, '2017-2019_communes_sru_en_paca', index_col=2, dtype={"REG": str, "DEP": str})
        sru2020 = pd.read_excel(xls, '2020-2022_communes_sru_en_paca', index_col=2, dtype={"REG": str, "DEP": str})
    return sru2017, sru2020

#################
### Communes
#################

metaDossierSourcePage = "https://www.insee.fr/fr/statistiques/5359146"
metaDossierSourceFile = "https://www.insee.fr/fr/statistiques/fichier/5359146/dossier_complet.zip"
metaDossierFile = data_dir + "meta_dossier_complet.csv"
dossierCompletFile = data_dir + "dossier_complet.csv"
metaDossier    = None
dossierComplet = None
global_context["URL_SOURCE_COMMUNES"] = metaDossierSourcePage

# CODGEO;P18_POP;P18_POP0014;P18_POP1529;P18_POP3044;P18_POP4559;P18_POP6074;P18_POP7589;P18_POP90P;P18_POPH;P18_H0014;P18_H1529;P18_H3044;P18_H4559;P18_H6074;P18_H7589;P18_H90P;P18_H0019;P18_H2064;P18_H65P;P18_POPF;P18_F0014;P18_F1529;P18_F3044;P18_F4559;P18_F6074;P18_F7589;P18_F90P;P18_F0019;P18_F2064;P18_F65P;P18_POP01P;P18_POP01P_IRAN1;P18_POP01P_IRAN2;P18_POP01P_IRAN3;P18_POP01P_IRAN4;P18_POP01P_IRAN5;P18_POP01P_IRAN6;P18_POP01P_IRAN7;P18_POP0114_IRAN2P;P18_POP0114_IRAN2;P18_POP0114_IRAN3P;P18_POP1524_IRAN2P;P18_POP1524_IRAN2;P18_POP1524_IRAN3P;P18_POP2554_IRAN2P;P18_POP2554_IRAN2;P18_POP2554_IRAN3P;P18_POP55P_IRAN2P;P18_POP55P_IRAN2;P18_POP55P_IRAN3P;C18_POP15P;C18_POP15P_CS1;C18_POP15P_CS2;C18_POP15P_CS3;C18_POP15P_CS4;C18_POP15P_CS5;C18_POP15P_CS6;C18_POP15P_CS7;C18_POP15P_CS8;C18_H15P;C18_H15P_CS1;C18_H15P_CS2;C18_H15P_CS3;C18_H15P_CS4;C18_H15P_CS5;C18_H15P_CS6;C18_H15P_CS7;C18_H15P_CS8;C18_F15P;C18_F15P_CS1;C18_F15P_CS2;C18_F15P_CS3;C18_F15P_CS4;C18_F15P_CS5;C18_F15P_CS6;C18_F15P_CS7;C18_F15P_CS8;C18_POP1524;C18_POP1524_CS1;C18_POP1524_CS2;C18_POP1524_CS3;C18_POP1524_CS4;C18_POP1524_CS5;C18_POP1524_CS6;C18_POP1524_CS7;C18_POP1524_CS8;C18_POP2554;C18_POP2554_CS1;C18_POP2554_CS2;C18_POP2554_CS3;C18_POP2554_CS4;C18_POP2554_CS5;C18_POP2554_CS6;C18_POP2554_CS7;C18_POP2554_CS8;C18_POP55P;C18_POP55P_CS1;C18_POP55P_CS2;C18_POP55P_CS3;C18_POP55P_CS4;C18_POP55P_CS5;C18_POP55P_CS6;C18_POP55P_CS7;C18_POP55P_CS8;P13_POP;P13_POP0014;P13_POP1529;P13_POP3044;P13_POP4559;P13_POP6074;P13_POP7589;P13_POP90P;P13_POPH;P13_H0014;P13_H1529;P13_H3044;P13_H4559;P13_H6074;P13_H7589;P13_H90P;P13_H0019;P13_H2064;P13_H65P;P13_POPF;P13_F0014;P13_F1529;P13_F3044;P13_F4559;P13_F6074;P13_F7589;P13_F90P;P13_F0019;P13_F2064;P13_F65P;P13_POP01P;P13_POP01P_IRAN1;P13_POP01P_IRAN2;P13_POP01P_IRAN3;P13_POP01P_IRAN4;P13_POP01P_IRAN5;P13_POP01P_IRAN6;P13_POP01P_IRAN7;P13_POP0114_IRAN2P;P13_POP0114_IRAN2;P13_POP0114_IRAN3P;P13_POP1524_IRAN2P;P13_POP1524_IRAN2;P13_POP1524_IRAN3P;P13_POP2554_IRAN2P;P13_POP2554_IRAN2;P13_POP2554_IRAN3P;P13_POP55P_IRAN2P;P13_POP55P_IRAN2;P13_POP55P_IRAN3P;C13_POP15P;C13_POP15P_CS1;C13_POP15P_CS2;C13_POP15P_CS3;C13_POP15P_CS4;C13_POP15P_CS5;C13_POP15P_CS6;C13_POP15P_CS7;C13_POP15P_CS8;C13_H15P;C13_H15P_CS1;C13_H15P_CS2;C13_H15P_CS3;C13_H15P_CS4;C13_H15P_CS5;C13_H15P_CS6;C13_H15P_CS7;C13_H15P_CS8;C13_F15P;C13_F15P_CS1;C13_F15P_CS2;C13_F15P_CS3;C13_F15P_CS4;C13_F15P_CS5;C13_F15P_CS6;C13_F15P_CS7;C13_F15P_CS8;C13_POP1524;C13_POP1524_CS1;C13_POP1524_CS2;C13_POP1524_CS3;C13_POP1524_CS4;C13_POP1524_CS5;C13_POP1524_CS6;C13_POP1524_CS7;C13_POP1524_CS8;C13_POP2554;C13_POP2554_CS1;C13_POP2554_CS2;C13_POP2554_CS3;C13_POP2554_CS4;C13_POP2554_CS5;C13_POP2554_CS6;C13_POP2554_CS7;C13_POP2554_CS8;C13_POP55P;C13_POP55P_CS1;C13_POP55P_CS2;C13_POP55P_CS3;C13_POP55P_CS4;C13_POP55P_CS5;C13_POP55P_CS6;C13_POP55P_CS7;C13_POP55P_CS8;P08_POP;P08_POP0014;P08_POP1529;P08_POP3044;P08_POP4559;P08_POP6074;P08_POP75P;P08_POPH;P08_H0014;P08_H1529;P08_H3044;P08_H4559;P08_H6074;P08_H7589;P08_H90P;P08_H0019;P08_H2064;P08_H65P;P08_POPF;P08_F0014;P08_F1529;P08_F3044;P08_F4559;P08_F6074;P08_F7589;P08_F90P;P08_F0019;P08_F2064;P08_F65P;P08_POP05P;P08_POP05P_IRAN1;P08_POP05P_IRAN2;P08_POP05P_IRAN3;P08_POP05P_IRAN4;P08_POP05P_IRAN5;P08_POP05P_IRAN6;P08_POP05P_IRAN7;P08_POP0514;P08_POP0514_IRAN2;P08_POP0514_IRAN3P;P08_POP1524;P08_POP1524_IRAN2;P08_POP1524_IRAN3P;P08_POP2554;P08_POP2554_IRAN2;P08_POP2554_IRAN3P;P08_POP55P;P08_POP55P_IRAN2;P08_POP55P_IRAN3P;C08_POP15P;C08_POP15P_CS1;C08_POP15P_CS2;C08_POP15P_CS3;C08_POP15P_CS4;C08_POP15P_CS5;C08_POP15P_CS6;C08_POP15P_CS7;C08_POP15P_CS8;C08_H15P;C08_H15P_CS1;C08_H15P_CS2;C08_H15P_CS3;C08_H15P_CS4;C08_H15P_CS5;C08_H15P_CS6;C08_H15P_CS7;C08_H15P_CS8;C08_F15P;C08_F15P_CS1;C08_F15P_CS2;C08_F15P_CS3;C08_F15P_CS4;C08_F15P_CS5;C08_F15P_CS6;C08_F15P_CS7;C08_F15P_CS8;C08_POP1524;C08_POP1524_CS1;C08_POP1524_CS2;C08_POP1524_CS3;C08_POP1524_CS4;C08_POP1524_CS5;C08_POP1524_CS6;C08_POP1524_CS7;C08_POP1524_CS8;C08_POP2554;C08_POP2554_CS1;C08_POP2554_CS2;C08_POP2554_CS3;C08_POP2554_CS4;C08_POP2554_CS5;C08_POP2554_CS6;C08_POP2554_CS7;C08_POP2554_CS8;C08_POP55P;C08_POP55P_CS1;C08_POP55P_CS2;C08_POP55P_CS3;C08_POP55P_CS4;C08_POP55P_CS5;C08_POP55P_CS6;C08_POP55P_CS7;C08_POP55P_CS8;C18_MEN;C18_MENPSEUL;C18_MENHSEUL;C18_MENFSEUL;C18_MENSFAM;C18_MENFAM;C18_MENCOUPSENF;C18_MENCOUPAENF;C18_MENFAMMONO;C18_PMEN;C18_PMEN_MENPSEUL;C18_PMEN_MENHSEUL;C18_PMEN_MENFSEUL;C18_PMEN_MENSFAM;C18_PMEN_MENFAM;C18_PMEN_MENCOUPSENF;C18_PMEN_MENCOUPAENF;C18_PMEN_MENFAMMONO;P18_POP15P;P18_POP1519;P18_POP2024;P18_POP2539;P18_POP4054;P18_POP5564;P18_POP6579;P18_POP80P;P18_POPMEN1519;P18_POPMEN2024;P18_POPMEN2539;P18_POPMEN4054;P18_POPMEN5564;P18_POPMEN6579;P18_POPMEN80P;P18_POP1519_PSEUL;P18_POP2024_PSEUL;P18_POP2539_PSEUL;P18_POP4054_PSEUL;P18_POP5564_PSEUL;P18_POP6579_PSEUL;P18_POP80P_PSEUL;P18_POP1519_COUPLE;P18_POP2024_COUPLE;P18_POP2539_COUPLE;P18_POP4054_COUPLE;P18_POP5564_COUPLE;P18_POP6579_COUPLE;P18_POP80P_COUPLE;P18_POP15P_MARIEE;P18_POP15P_PACSEE;P18_POP15P_CONCUB_UNION_LIBRE;P18_POP15P_VEUFS;P18_POP15P_DIVORCEE;P18_POP15P_CELIBATAIRE;C18_MEN_CS1;C18_MEN_CS2;C18_MEN_CS3;C18_MEN_CS4;C18_MEN_CS5;C18_MEN_CS6;C18_MEN_CS7;C18_MEN_CS8;C18_PMEN_CS1;C18_PMEN_CS2;C18_PMEN_CS3;C18_PMEN_CS4;C18_PMEN_CS5;C18_PMEN_CS6;C18_PMEN_CS7;C18_PMEN_CS8;C18_FAM;C18_COUPAENF;C18_FAMMONO;C18_HMONO;C18_FMONO;C18_COUPSENF;C18_NE24F0;C18_NE24F1;C18_NE24F2;C18_NE24F3;C18_NE24F4P;C13_MEN;C13_MENPSEUL;C13_MENHSEUL;C13_MENFSEUL;C13_MENSFAM;C13_MENFAM;C13_MENCOUPSENF;C13_MENCOUPAENF;C13_MENFAMMONO;C13_PMEN;C13_PMEN_MENPSEUL;C13_PMEN_MENHSEUL;C13_PMEN_MENFSEUL;C13_PMEN_MENSFAM;C13_PMEN_MENFAM;C13_PMEN_MENCOUPSENF;C13_PMEN_MENCOUPAENF;C13_PMEN_MENFAMMONO;P13_POP15P;P13_POP1519;P13_POP2024;P13_POP2539;P13_POP4054;P13_POP5564;P13_POP6579;P13_POP80P;P13_POPMEN1519;P13_POPMEN2024;P13_POPMEN2539;P13_POPMEN4054;P13_POPMEN5564;P13_POPMEN6579;P13_POPMEN80P;P13_POP1519_PSEUL;P13_POP2024_PSEUL;P13_POP2539_PSEUL;P13_POP4054_PSEUL;P13_POP5564_PSEUL;P13_POP6579_PSEUL;P13_POP80P_PSEUL;P13_POP1519_COUPLE;P13_POP2024_COUPLE;P13_POP2539_COUPLE;P13_POP4054_COUPLE;P13_POP5564_COUPLE;P13_POP6579_COUPLE;P13_POP80P_COUPLE;P13_POP15P_MARIEE;P13_POP15P_NONMARIEE;C13_MEN_CS1;C13_MEN_CS2;C13_MEN_CS3;C13_MEN_CS4;C13_MEN_CS5;C13_MEN_CS6;C13_MEN_CS7;C13_MEN_CS8;C13_PMEN_CS1;C13_PMEN_CS2;C13_PMEN_CS3;C13_PMEN_CS4;C13_PMEN_CS5;C13_PMEN_CS6;C13_PMEN_CS7;C13_PMEN_CS8;C13_FAM;C13_COUPAENF;C13_FAMMONO;C13_HMONO;C13_FMONO;C13_COUPSENF;C13_NE24F0;C13_NE24F1;C13_NE24F2;C13_NE24F3;C13_NE24F4P;C08_MEN;C08_MENPSEUL;C08_MENHSEUL;C08_MENFSEUL;C08_MENSFAM;C08_MENFAM;C08_MENCOUPSENF;C08_MENCOUPAENF;C08_MENFAMMONO;C08_PMEN;C08_PMEN_MENPSEUL;C08_PMEN_MENHSEUL;C08_PMEN_MENFSEUL;C08_PMEN_MENSFAM;C08_PMEN_MENFAM;C08_PMEN_MENCOUPSENF;C08_PMEN_MENCOUPAENF;C08_PMEN_MENFAMMONO;P08_POP15P;P08_POP1519;P08_POP2024;P08_POP2539;P08_POP4054;P08_POP5564;P08_POP6579;P08_POP80P;P08_POPMEN1519;P08_POPMEN2024;P08_POPMEN2539;P08_POPMEN4054;P08_POPMEN5564;P08_POPMEN6579;P08_POPMEN80P;P08_POP1519_PSEUL;P08_POP2024_PSEUL;P08_POP2539_PSEUL;P08_POP4054_PSEUL;P08_POP5564_PSEUL;P08_POP6579_PSEUL;P08_POP80P_PSEUL;P08_POP1519_COUPLE;P08_POP2024_COUPLE;P08_POP2539_COUPLE;P08_POP4054_COUPLE;P08_POP5564_COUPLE;P08_POP6579_COUPLE;P08_POP80P_COUPLE;P08_POP15P_MARIE;P08_POP15P_CELIB;P08_POP15P_VEUF;P08_POP15P_DIVOR;C08_MEN_CS1;C08_MEN_CS2;C08_MEN_CS3;C08_MEN_CS4;C08_MEN_CS5;C08_MEN_CS6;C08_MEN_CS7;C08_MEN_CS8;C08_PMEN_CS1;C08_PMEN_CS2;C08_PMEN_CS3;C08_PMEN_CS4;C08_PMEN_CS5;C08_PMEN_CS6;C08_PMEN_CS7;C08_PMEN_CS8;C08_FAM;C08_COUPAENF;C08_FAMMONO;C08_HMONO;C08_FMONO;C08_COUPSENF;C08_NE24F0;C08_NE24F1;C08_NE24F2;C08_NE24F3;C08_NE24F4P;P18_LOG;P18_RP;P18_RSECOCC;P18_LOGVAC;P18_MAISON;P18_APPART;P18_RP_1P;P18_RP_2P;P18_RP_3P;P18_RP_4P;P18_RP_5PP;P18_NBPI_RP;P18_RPMAISON;P18_NBPI_RPMAISON;P18_RPAPPART;P18_NBPI_RPAPPART;C18_RP_HSTU1P;C18_RP_HSTU1P_SUROCC;P18_RP_ACHTOT;P18_RP_ACH19;P18_RP_ACH45;P18_RP_ACH70;P18_RP_ACH90;P18_RP_ACH05;P18_RP_ACH15;P18_RPMAISON_ACH19;P18_RPMAISON_ACH45;P18_RPMAISON_ACH70;P18_RPMAISON_ACH90;P18_RPMAISON_ACH05;P18_RPMAISON_ACH15;P18_RPAPPART_ACH19;P18_RPAPPART_ACH45;P18_RPAPPART_ACH70;P18_RPAPPART_ACH90;P18_RPAPPART_ACH05;P18_RPAPPART_ACH15;P18_MEN;P18_MEN_ANEM0002;P18_MEN_ANEM0204;P18_MEN_ANEM0509;P18_MEN_ANEM10P;P18_MEN_ANEM1019;P18_MEN_ANEM2029;P18_MEN_ANEM30P;P18_PMEN;P18_PMEN_ANEM0002;P18_PMEN_ANEM0204;P18_PMEN_ANEM0509;P18_PMEN_ANEM10P;P18_NBPI_RP_ANEM0002;P18_NBPI_RP_ANEM0204;P18_NBPI_RP_ANEM0509;P18_NBPI_RP_ANEM10P;P18_RP_PROP;P18_RP_LOC;P18_RP_LOCHLMV;P18_RP_GRAT;P18_NPER_RP;P18_NPER_RP_PROP;P18_NPER_RP_LOC;P18_NPER_RP_LOCHLMV;P18_NPER_RP_GRAT;P18_ANEM_RP;P18_ANEM_RP_PROP;P18_ANEM_RP_LOC;P18_ANEM_RP_LOCHLMV;P18_ANEM_RP_GRAT;P18_RP_SDB;P18_RP_CCCOLL;P18_RP_CCIND;P18_RP_CINDELEC;P18_RP_ELEC;P18_RP_EAUCH;P18_RP_BDWC;P18_RP_CHOS;P18_RP_CLIM;P18_RP_TTEGOU;P18_RP_GARL;P18_RP_VOIT1P;P18_RP_VOIT1;P18_RP_VOIT2P;P18_RP_HABFOR;P18_RP_CASE;P18_RP_MIBOIS;P18_RP_MIDUR;P13_LOG;P13_RP;P13_RSECOCC;P13_LOGVAC;P13_MAISON;P13_APPART;P13_RP_1P;P13_RP_2P;P13_RP_3P;P13_RP_4P;P13_RP_5PP;P13_NBPI_RP;P13_RPMAISON;P13_NBPI_RPMAISON;P13_RPAPPART;P13_NBPI_RPAPPART;P13_RP_ACHTOT;P13_RP_ACH19;P13_RP_ACH45;P13_RP_ACH70;P13_RP_ACH90;P13_RP_ACH05;P13_RP_ACH10;P13_RPMAISON_ACH19;P13_RPMAISON_ACH45;P13_RPMAISON_ACH70;P13_RPMAISON_ACH90;P13_RPMAISON_ACH05;P13_RPMAISON_ACH10;P13_RPAPPART_ACH19;P13_RPAPPART_ACH45;P13_RPAPPART_ACH70;P13_RPAPPART_ACH90;P13_RPAPPART_ACH05;P13_RPAPPART_ACH10;P13_MEN;P13_MEN_ANEM0002;P13_MEN_ANEM0204;P13_MEN_ANEM0509;P13_MEN_ANEM10P;P13_MEN_ANEM1019;P13_MEN_ANEM2029;P13_MEN_ANEM30P;P13_PMEN;P13_PMEN_ANEM0002;P13_PMEN_ANEM0204;P13_PMEN_ANEM0509;P13_PMEN_ANEM10P;P13_NBPI_RP_ANEM0002;P13_NBPI_RP_ANEM0204;P13_NBPI_RP_ANEM0509;P13_NBPI_RP_ANEM10P;P13_RP_PROP;P13_RP_LOC;P13_RP_LOCHLMV;P13_RP_GRAT;P13_NPER_RP;P13_NPER_RP_PROP;P13_NPER_RP_LOC;P13_NPER_RP_LOCHLMV;P13_NPER_RP_GRAT;P13_ANEM_RP;P13_ANEM_RP_PROP;P13_ANEM_RP_LOC;P13_ANEM_RP_LOCHLMV;P13_ANEM_RP_GRAT;P13_RP_SDB;P13_RP_CCCOLL;P13_RP_CCIND;P13_RP_CINDELEC;P13_RP_ELEC;P13_RP_EAUCH;P13_RP_BDWC;P13_RP_CHOS;P13_RP_CLIM;P13_RP_TTEGOU;P13_RP_GARL;P13_RP_VOIT1P;P13_RP_VOIT1;P13_RP_VOIT2P;P13_RP_HABFOR;P13_RP_CASE;P13_RP_MIBOIS;P13_RP_MIDUR;P08_LOG;P08_RP;P08_RSECOCC;P08_LOGVAC;P08_MAISON;P08_APPART;P08_RP_1P;P08_RP_2P;P08_RP_3P;P08_RP_4P;P08_RP_5PP;P08_NBPI_RP;P08_RPMAISON;P08_NBPI_RPMAISON;P08_RPAPPART;P08_NBPI_RPAPPART;P08_RP_ACHTT;P08_RP_ACHT1;P08_RP_ACHT2;P08_RP_ACHT3;P08_RP_ACHT4;P08_RPMAISON_ACHT1;P08_RPMAISON_ACHT2;P08_RPMAISON_ACHT3;P08_RPMAISON_ACHT4;P08_RPAPPART_ACHT1;P08_RPAPPART_ACHT2;P08_RPAPPART_ACHT3;P08_RPAPPART_ACHT4;P08_MEN;P08_MEN_ANEM0002;P08_MEN_ANEM0204;P08_MEN_ANEM0509;P08_MEN_ANEM10P;P08_MEN_ANEM1019;P08_MEN_ANEM2029;P08_MEN_ANEM30P;P08_PMEN;P08_PMEN_ANEM0002;P08_PMEN_ANEM0204;P08_PMEN_ANEM0509;P08_PMEN_ANEM10P;P08_NBPI_RP_ANEM0002;P08_NBPI_RP_ANEM0204;P08_NBPI_RP_ANEM0509;P08_NBPI_RP_ANEM10P;P08_RP_PROP;P08_RP_LOC;P08_RP_LOCHLMV;P08_RP_GRAT;P08_NPER_RP;P08_NPER_RP_PROP;P08_NPER_RP_LOC;P08_NPER_RP_LOCHLMV;P08_NPER_RP_GRAT;P08_ANEM_RP;P08_ANEM_RP_PROP;P08_ANEM_RP_LOC;P08_ANEM_RP_LOCHLMV;P08_ANEM_RP_GRAT;P08_RP_SDB;P08_RP_CCCOLL;P08_RP_CCIND;P08_RP_CINDELEC;P08_RP_ELEC;P08_RP_EAUCH;P08_RP_BDWC;P08_RP_CHOS;P08_RP_CLIM;P08_RP_TTEGOU;P08_RP_GARL;P08_RP_VOIT1P;P08_RP_VOIT1;P08_RP_VOIT2P;P08_RP_HABFOR;P08_RP_CASE;P08_RP_MIBOIS;P08_RP_MIDUR;P18_POP0205;P18_POP0610;P18_POP1114;P18_POP1517;P18_POP1824;P18_POP2529;P18_POP30P;P18_SCOL0205;P18_SCOL0610;P18_SCOL1114;P18_SCOL1517;P18_SCOL1824;P18_SCOL2529;P18_SCOL30P;P18_H0205;P18_H0610;P18_H1114;P18_H1517;P18_H1824;P18_H2529;P18_H30P;P18_HSCOL0205;P18_HSCOL0610;P18_HSCOL1114;P18_HSCOL1517;P18_HSCOL1824;P18_HSCOL2529;P18_HSCOL30P;P18_F0205;P18_F0610;P18_F1114;P18_F1517;P18_F1824;P18_F2529;P18_F30P;P18_FSCOL0205;P18_FSCOL0610;P18_FSCOL1114;P18_FSCOL1517;P18_FSCOL1824;P18_FSCOL2529;P18_FSCOL30P;P18_NSCOL15P;P18_NSCOL15P_DIPLMIN;P18_NSCOL15P_BEPC;P18_NSCOL15P_CAPBEP;P18_NSCOL15P_BAC;P18_NSCOL15P_SUP2;P18_NSCOL15P_SUP34;P18_NSCOL15P_SUP5;P18_HNSCOL15P;P18_HNSCOL15P_DIPLMIN;P18_HNSCOL15P_BEPC;P18_HNSCOL15P_CAPBEP;P18_HNSCOL15P_BAC;P18_HNSCOL15P_SUP2;P18_HNSCOL15P_SUP34;P18_HNSCOL15P_SUP5;P18_FNSCOL15P;P18_FNSCOL15P_DIPLMIN;P18_FNSCOL15P_BEPC;P18_FNSCOL15P_CAPBEP;P18_FNSCOL15P_BAC;P18_FNSCOL15P_SUP2;P18_FNSCOL15P_SUP34;P18_FNSCOL15P_SUP5;P13_POP0205;P13_POP0610;P13_POP1114;P13_POP1517;P13_POP1824;P13_POP2529;P13_POP30P;P13_SCOL0205;P13_SCOL0610;P13_SCOL1114;P13_SCOL1517;P13_SCOL1824;P13_SCOL2529;P13_SCOL30P;P13_H0205;P13_H0610;P13_H1114;P13_H1517;P13_H1824;P13_H2529;P13_H30P;P13_HSCOL0205;P13_HSCOL0610;P13_HSCOL1114;P13_HSCOL1517;P13_HSCOL1824;P13_HSCOL2529;P13_HSCOL30P;P13_F0205;P13_F0610;P13_F1114;P13_F1517;P13_F1824;P13_F2529;P13_F30P;P13_FSCOL0205;P13_FSCOL0610;P13_FSCOL1114;P13_FSCOL1517;P13_FSCOL1824;P13_FSCOL2529;P13_FSCOL30P;P13_NSCOL15P;P13_NSCOL15P_DIPLMIN;P13_NSCOL15P_CAPBEP;P13_NSCOL15P_BAC;P13_NSCOL15P_SUP;P13_HNSCOL15P;P13_HNSCOL15P_DIPLMIN;P13_HNSCOL15P_CAPBEP;P13_HNSCOL15P_BAC;P13_HNSCOL15P_SUP;P13_FNSCOL15P;P13_FNSCOL15P_DIPLMIN;P13_FNSCOL15P_CAPBEP;P13_FNSCOL15P_BAC;P13_FNSCOL15P_SUP;P08_POP0205;P08_POP0610;P08_POP1114;P08_POP1517;P08_POP1824;P08_POP2529;P08_POP30P;P08_SCOL0205;P08_SCOL0610;P08_SCOL1114;P08_SCOL1517;P08_SCOL1824;P08_SCOL2529;P08_SCOL30P;P08_H0205;P08_H0610;P08_H1114;P08_H1517;P08_H1824;P08_H2529;P08_H30P;P08_HSCOL0205;P08_HSCOL0610;P08_HSCOL1114;P08_HSCOL1517;P08_HSCOL1824;P08_HSCOL2529;P08_HSCOL30P;P08_F0205;P08_F0610;P08_F1114;P08_F1517;P08_F1824;P08_F2529;P08_F30P;P08_FSCOL0205;P08_FSCOL0610;P08_FSCOL1114;P08_FSCOL1517;P08_FSCOL1824;P08_FSCOL2529;P08_FSCOL30P;P08_NSCOL15P;P08_NSCOL15P_DIPL0;P08_NSCOL15P_CEP;P08_NSCOL15P_BEPC;P08_NSCOL15P_CAPBEP;P08_NSCOL15P_BAC;P08_NSCOL15P_BACP2;P08_NSCOL15P_SUP;P08_HNSCOL15P;P08_HNSCOL15P_DIPL0;P08_HNSCOL15P_CEP;P08_HNSCOL15P_BEPC;P08_HNSCOL15P_CAPBEP;P08_HNSCOL15P_BAC;P08_HNSCOL15P_BACP2;P08_HNSCOL15P_SUP;P08_FNSCOL15P;P08_FNSCOL15P_DIPL0;P08_FNSCOL15P_CEP;P08_FNSCOL15P_BEPC;P08_FNSCOL15P_CAPBEP;P08_FNSCOL15P_BAC;P08_FNSCOL15P_BACP2;P08_FNSCOL15P_SUP;P18_POP1564;P18_POP1524;P18_POP2554;P18_H1564;P18_H1524;P18_H2554;P18_H5564;P18_F1564;P18_F1524;P18_F2554;P18_F5564;P18_ACT1564;P18_ACT1524;P18_ACT2554;P18_ACT5564;P18_HACT1564;P18_HACT1524;P18_HACT2554;P18_HACT5564;P18_FACT1564;P18_FACT1524;P18_FACT2554;P18_FACT5564;P18_ACTOCC1564;P18_ACTOCC1524;P18_ACTOCC2554;P18_ACTOCC5564;P18_HACTOCC1564;P18_HACTOCC1524;P18_HACTOCC2554;P18_HACTOCC5564;P18_FACTOCC1564;P18_FACTOCC1524;P18_FACTOCC2554;P18_FACTOCC5564;P18_CHOM1564;P18_HCHOM1564;P18_HCHOM1524;P18_HCHOM2554;P18_HCHOM5564;P18_FCHOM1564;P18_FCHOM1524;P18_FCHOM2554;P18_FCHOM5564;P18_INACT1564;P18_ETUD1564;P18_RETR1564;P18_AINACT1564;C18_ACT1564;C18_ACT1564_CS1;C18_ACT1564_CS2;C18_ACT1564_CS3;C18_ACT1564_CS4;C18_ACT1564_CS5;C18_ACT1564_CS6;C18_ACTOCC1564;C18_ACTOCC1564_CS1;C18_ACTOCC1564_CS2;C18_ACTOCC1564_CS3;C18_ACTOCC1564_CS4;C18_ACTOCC1564_CS5;C18_ACTOCC1564_CS6;P18_EMPLT;P18_ACTOCC;P18_ACT15P;P18_EMPLT_SAL;P18_EMPLT_FSAL;P18_EMPLT_SALTP;P18_EMPLT_NSAL;P18_EMPLT_FNSAL;P18_EMPLT_NSALTP;C18_EMPLT;C18_EMPLT_CS1;C18_EMPLT_CS2;C18_EMPLT_CS3;C18_EMPLT_CS4;C18_EMPLT_CS5;C18_EMPLT_CS6;C18_EMPLT_AGRI;C18_EMPLT_INDUS;C18_EMPLT_CONST;C18_EMPLT_CTS;C18_EMPLT_APESAS;C18_EMPLT_F;C18_AGRILT_F;C18_INDUSLT_F;C18_CONSTLT_F;C18_CTSLT_F;C18_APESASLT_F;C18_EMPLT_SAL;C18_AGRILT_SAL;C18_INDUSLT_SAL;C18_CONSTLT_SAL;C18_CTSLT_SAL;C18_APESASLT_SAL;C18_AGRILT_FSAL;C18_INDUSLT_FSAL;C18_CONSTLT_FSAL;C18_CTSLT_FSAL;C18_APESASLT_FSAL;C18_AGRILT_NSAL;C18_INDUSLT_NSAL;C18_CONSTLT_NSAL;C18_CTSLT_NSAL;C18_APESASLT_NSAL;C18_AGRILT_FNSAL;C18_INDUSLT_FNSAL;C18_CONSTLT_FNSAL;C18_CTSLT_FNSAL;C18_APESASLT_FNSAL;P13_POP1564;P13_POP1524;P13_POP2554;P13_H1564;P13_H1524;P13_H2554;P13_H5564;P13_F1564;P13_F1524;P13_F2554;P13_F5564;P13_ACT1564;P13_ACT1524;P13_ACT2554;P13_ACT5564;P13_HACT1564;P13_HACT1524;P13_HACT2554;P13_HACT5564;P13_FACT1564;P13_FACT1524;P13_FACT2554;P13_FACT5564;P13_ACTOCC1564;P13_ACTOCC1524;P13_ACTOCC2554;P13_ACTOCC5564;P13_HACTOCC1564;P13_HACTOCC1524;P13_HACTOCC2554;P13_HACTOCC5564;P13_FACTOCC1564;P13_FACTOCC1524;P13_FACTOCC2554;P13_FACTOCC5564;P13_CHOM1564;P13_HCHOM1564;P13_HCHOM1524;P13_HCHOM2554;P13_HCHOM5564;P13_FCHOM1564;P13_FCHOM1524;P13_FCHOM2554;P13_FCHOM5564;P13_INACT1564;P13_ETUD1564;P13_RETR1564;P13_AINACT1564;C13_ACT1564;C13_ACT1564_CS1;C13_ACT1564_CS2;C13_ACT1564_CS3;C13_ACT1564_CS4;C13_ACT1564_CS5;C13_ACT1564_CS6;C13_ACTOCC1564;C13_ACTOCC1564_CS1;C13_ACTOCC1564_CS2;C13_ACTOCC1564_CS3;C13_ACTOCC1564_CS4;C13_ACTOCC1564_CS5;C13_ACTOCC1564_CS6;P13_EMPLT;P13_ACTOCC;P13_ACT15P;P13_EMPLT_SAL;P13_EMPLT_FSAL;P13_EMPLT_SALTP;P13_EMPLT_NSAL;P13_EMPLT_FNSAL;P13_EMPLT_NSALTP;C13_EMPLT;C13_EMPLT_CS1;C13_EMPLT_CS2;C13_EMPLT_CS3;C13_EMPLT_CS4;C13_EMPLT_CS5;C13_EMPLT_CS6;C13_EMPLT_AGRI;C13_EMPLT_INDUS;C13_EMPLT_CONST;C13_EMPLT_CTS;C13_EMPLT_APESAS;C13_EMPLT_F;C13_AGRILT_F;C13_INDUSLT_F;C13_CONSTLT_F;C13_CTSLT_F;C13_APESASLT_F;C13_EMPLT_SAL;C13_AGRILT_SAL;C13_INDUSLT_SAL;C13_CONSTLT_SAL;C13_CTSLT_SAL;C13_APESASLT_SAL;C13_AGRILT_FSAL;C13_INDUSLT_FSAL;C13_CONSTLT_FSAL;C13_CTSLT_FSAL;C13_APESASLT_FSAL;C13_AGRILT_NSAL;C13_INDUSLT_NSAL;C13_CONSTLT_NSAL;C13_CTSLT_NSAL;C13_APESASLT_NSAL;C13_AGRILT_FNSAL;C13_INDUSLT_FNSAL;C13_CONSTLT_FNSAL;C13_CTSLT_FNSAL;C13_APESASLT_FNSAL;P08_POP1564;P08_H1564;P08_H1524;P08_H2554;P08_H5564;P08_F1564;P08_F1524;P08_F2554;P08_F5564;P08_ACT1564;P08_ACT1524;P08_ACT2554;P08_ACT5564;P08_HACT1564;P08_HACT1524;P08_HACT2554;P08_HACT5564;P08_FACT1564;P08_FACT1524;P08_FACT2554;P08_FACT5564;P08_ACTOCC1564;P08_ACTOCC1524;P08_ACTOCC2554;P08_ACTOCC5564;P08_HACTOCC1564;P08_HACTOCC1524;P08_HACTOCC2554;P08_HACTOCC5564;P08_FACTOCC1564;P08_FACTOCC1524;P08_FACTOCC2554;P08_FACTOCC5564;P08_CHOM1564;P08_HCHOM1564;P08_HCHOM1524;P08_HCHOM2554;P08_HCHOM5564;P08_FCHOM1564;P08_FCHOM1524;P08_FCHOM2554;P08_FCHOM5564;P08_INACT1564;P08_ETUD1564;P08_RETR1564;P08_AINACT1564;C08_ACT1564;C08_ACT1564_CS1;C08_ACT1564_CS2;C08_ACT1564_CS3;C08_ACT1564_CS4;C08_ACT1564_CS5;C08_ACT1564_CS6;C08_ACTOCC1564;C08_ACTOCC1564_CS1;C08_ACTOCC1564_CS2;C08_ACTOCC1564_CS3;C08_ACTOCC1564_CS4;C08_ACTOCC1564_CS5;C08_ACTOCC1564_CS6;P08_EMPLT;P08_ACTOCC;P08_ACT15P;P08_EMPLT_SAL;P08_EMPLT_FSAL;P08_EMPLT_SALTP;P08_EMPLT_NSAL;P08_EMPLT_FNSAL;P08_EMPLT_NSALTP;C08_EMPLT;C08_EMPLT_CS1;C08_EMPLT_CS2;C08_EMPLT_CS3;C08_EMPLT_CS4;C08_EMPLT_CS5;C08_EMPLT_CS6;C08_EMPLT_AGRI;C08_EMPLT_INDUS;C08_EMPLT_CONST;C08_EMPLT_CTS;C08_EMPLT_APESAS;C08_EMPLT_F;C08_AGRILT_F;C08_INDUSLT_F;C08_CONSTLT_F;C08_CTSLT_F;C08_APESASLT_F;C08_EMPLT_SAL;C08_AGRILT_SAL;C08_INDUSLT_SAL;C08_CONSTLT_SAL;C08_CTSLT_SAL;C08_APESASLT_SAL;C08_AGRILT_FSAL;C08_INDUSLT_FSAL;C08_CONSTLT_FSAL;C08_CTSLT_FSAL;C08_APESASLT_FSAL;C08_AGRILT_NSAL;C08_INDUSLT_NSAL;C08_CONSTLT_NSAL;C08_CTSLT_NSAL;C08_APESASLT_NSAL;C08_AGRILT_FNSAL;C08_INDUSLT_FNSAL;C08_CONSTLT_FNSAL;C08_CTSLT_FNSAL;C08_APESASLT_FNSAL;P18_ACTOCC15P;P18_SAL15P;P18_NSAL15P;P18_ACTOCC15P_TP;P18_SAL15P_TP;P18_HSAL15P_TP;P18_FSAL15P_TP;P18_NSAL15P_TP;P18_HACTOCC15P;P18_HSAL15P;P18_HSAL15P_CDI;P18_HSAL15P_CDD;P18_HSAL15P_INTERIM;P18_HSAL15P_EMPAID;P18_HSAL15P_APPR;P18_HNSAL15P;P18_HNSAL15P_INDEP;P18_HNSAL15P_EMPLOY;P18_HNSAL15P_AIDFAM;P18_FACTOCC15P;P18_FSAL15P;P18_FSAL15P_CDI;P18_FSAL15P_CDD;P18_FSAL15P_INTERIM;P18_FSAL15P_EMPAID;P18_FSAL15P_APPR;P18_FNSAL15P;P18_FNSAL15P_INDEP;P18_FNSAL15P_EMPLOY;P18_FNSAL15P_AIDFAM;P18_HSAL1564;P18_HSAL1524;P18_HSAL2554;P18_HSAL5564;P18_HSAL1564_TP;P18_HSAL1524_TP;P18_HSAL2554_TP;P18_HSAL5564_TP;P18_FSAL1564;P18_FSAL1524;P18_FSAL2554;P18_FSAL5564;P18_FSAL1564_TP;P18_FSAL1524_TP;P18_FSAL2554_TP;P18_FSAL5564_TP;P18_ACTOCC15P_ILT1;P18_ACTOCC15P_ILT2P;P18_ACTOCC15P_ILT2;P18_ACTOCC15P_ILT3;P18_ACTOCC15P_ILT4;P18_ACTOCC15P_ILT5;P18_ACTOCC15P_PASTRANS;P18_ACTOCC15P_MARCHE;P18_ACTOCC15P_VELO;P18_ACTOCC15P_2ROUESMOT;P18_ACTOCC15P_VOITURE;P18_ACTOCC15P_COMMUN;P13_ACTOCC15P;P13_SAL15P;P13_NSAL15P;P13_ACTOCC15P_TP;P13_SAL15P_TP;P13_HSAL15P_TP;P13_FSAL15P_TP;P13_NSAL15P_TP;P13_HACTOCC15P;P13_HSAL15P;P13_HSAL15P_CDI;P13_HSAL15P_CDD;P13_HSAL15P_INTERIM;P13_HSAL15P_EMPAID;P13_HSAL15P_APPR;P13_HNSAL15P;P13_HNSAL15P_INDEP;P13_HNSAL15P_EMPLOY;P13_HNSAL15P_AIDFAM;P13_FACTOCC15P;P13_FSAL15P;P13_FSAL15P_CDI;P13_FSAL15P_CDD;P13_FSAL15P_INTERIM;P13_FSAL15P_EMPAID;P13_FSAL15P_APPR;P13_FNSAL15P;P13_FNSAL15P_INDEP;P13_FNSAL15P_EMPLOY;P13_FNSAL15P_AIDFAM;P13_HSAL1564;P13_HSAL1524;P13_HSAL2554;P13_HSAL5564;P13_HSAL1564_TP;P13_HSAL1524_TP;P13_HSAL2554_TP;P13_HSAL5564_TP;P13_FSAL1564;P13_FSAL1524;P13_FSAL2554;P13_FSAL5564;P13_FSAL1564_TP;P13_FSAL1524_TP;P13_FSAL2554_TP;P13_FSAL5564_TP;P13_ACTOCC15P_ILT1;P13_ACTOCC15P_ILT2P;P13_ACTOCC15P_ILT2;P13_ACTOCC15P_ILT3;P13_ACTOCC15P_ILT4;P13_ACTOCC15P_ILT5;P13_ACTOCC15P_PASTRANS;P13_ACTOCC15P_MARCHE;P13_ACTOCC15P_2ROUES;P13_ACTOCC15P_VOITURE;P13_ACTOCC15P_COMMUN;P08_ACTOCC15P;P08_SAL15P;P08_NSAL15P;P08_ACTOCC15P_TP;P08_SAL15P_TP;P08_HSAL15P_TP;P08_FSAL15P_TP;P08_NSAL15P_TP;P08_HACTOCC15P;P08_HSAL15P;P08_HSAL15P_CDI;P08_HSAL15P_CDD;P08_HSAL15P_INTERIM;P08_HSAL15P_EMPAID;P08_HSAL15P_APPR;P08_HNSAL15P;P08_HNSAL15P_INDEP;P08_HNSAL15P_EMPLOY;P08_HNSAL15P_AIDFAM;P08_FACTOCC15P;P08_FSAL15P;P08_FSAL15P_CDI;P08_FSAL15P_CDD;P08_FSAL15P_INTERIM;P08_FSAL15P_EMPAID;P08_FSAL15P_APPR;P08_FNSAL15P;P08_FNSAL15P_INDEP;P08_FNSAL15P_EMPLOY;P08_FNSAL15P_AIDFAM;P08_HSAL1564;P08_HSAL1524;P08_HSAL2554;P08_HSAL5564;P08_HSAL1564_TP;P08_HSAL1524_TP;P08_HSAL2554_TP;P08_HSAL5564_TP;P08_FSAL1564;P08_FSAL1524;P08_FSAL2554;P08_FSAL5564;P08_FSAL1564_TP;P08_FSAL1524_TP;P08_FSAL2554_TP;P08_FSAL5564_TP;P08_ACTOCC15P_ILT1;P08_ACTOCC15P_ILT2P;P08_ACTOCC15P_ILT2;P08_ACTOCC15P_ILT3;P08_ACTOCC15P_ILT4;P08_ACTOCC15P_ILT5;D99_POP;D90_POP;D82_POP;D75_POP;D68_POP;SUPERF;NAIS1318;NAIS0813;NAIS9908;NAIS9099;NAIS8290;NAIS7582;NAIS6875;DECE1318;DECE0813;DECE9908;DECE9099;DECE8290;DECE7582;DECE6875;D99_LOG;D90_LOG;D82_LOG;D75_LOG;D68_LOG;D99_RP;D90_RP;D82_RP;D75_RP;D68_RP;D99_RSECOCC;D90_RSECOCC;D82_RSECOCC;D75_RSECOCC;D68_RSECOCC;D99_LOGVAC;D90_LOGVAC;D82_LOGVAC;D75_LOGVAC;D68_LOGVAC;D99_PMEN;D90_NPER_RP;D82_NPER_RP;D75_NPER_RP;D68_NPER_RP;NAISD14;NAISD15;NAISD16;NAISD17;NAISD18;NAISD19;NAISD20;DECESD14;DECESD15;DECESD16;DECESD17;DECESD18;DECESD19;DECESD20;NBMENFISC18;NBPERSMENFISC18;MED18;PIMP18;TP6018;TP60AGE118;TP60AGE218;TP60AGE318;TP60AGE418;TP60AGE518;TP60AGE618;TP60TOL118;TP60TOL218;PACT18;PTSA18;PCHO18;PBEN18;PPEN18;PPAT18;PPSOC18;PPFAM18;PPMINI18;PPLOGT18;PIMPOT18;D118;D918;RD18;SNHM19;SNHMC19;SNHMP19;SNHME19;SNHMO19;SNHMF19;SNHMFC19;SNHMFP19;SNHMFE19;SNHMFO19;SNHMH19;SNHMHC19;SNHMHP19;SNHMHE19;SNHMHO19;SNHM1819;SNHM2619;SNHM5019;SNHMF1819;SNHMF2619;SNHMF5019;SNHMH1819;SNHMH2619;SNHMH5019;ETTOT18;ETAZ18;ETBE18;ETFZ18;ETGU18;ETGZ18;ETOQ18;ETTEF018;ETAZ018;ETBE018;ETFZ018;ETGU018;ETGZ018;ETOQ018;ETTEF118;ETAZ118;ETBE118;ETFZ118;ETGU118;ETGZ118;ETOQ118;ETTEF1018;ETAZ1018;ETBE1018;ETFZ1018;ETGU1018;ETGZ1018;ETOQ1018;ETTEF2018;ETAZ2018;ETBE2018;ETFZ2018;ETGU2018;ETGZ2018;ETOQ2018;ETTEF5018;ETAZ5018;ETBE5018;ETFZ5018;ETGU5018;ETGZ5018;ETOQ5018;ETPTOT18;ETPAZ18;ETPBE18;ETPFZ18;ETPGU18;ETPGZ18;ETPOQ18;ETPTEF118;ETPAZ118;ETPBE118;ETPFZ118;ETPGU118;ETPGZ118;ETPOQ118;ETPTEF1018;ETPAZ1018;ETPBE1018;ETPFZ1018;ETPGU1018;ETPGZ1018;ETPOQ1018;ETPTEF2018;ETPAZ2018;ETPBE2018;ETPFZ2018;ETPGU2018;ETPGZ2018;ETPOQ2018;ETPTEF5018;ETPAZ5018;ETPBE5018;ETPFZ5018;ETPGU5018;ETPGZ5018;ETPOQ5018;ETPTEFCP18;ETPAZCP18;ETPBECP18;ETPFZCP18;ETPGUCP18;ETPGZCP18;ETPOQCP18;ETPRES18;ETNPRES18;ETPRESPUB18;ETNPRESPUB18;ETPPRES18;ETPNPRES18;ETPPRESPUB18;ETPNPRESPUB18;ETASSMAT18;ETAUTRES18;ENNTOT20;ENNBE20;ENNFZ20;ENNGI20;ENNJZ20;ENNKZ20;ENNLZ20;ENNMN20;ENNOQ20;ENNRU20;ENCTOT20;ENCBE20;ENCFZ20;ENCGI20;ENCJZ20;ENCKZ20;ENCLZ20;ENCMN20;ENCOQ20;ENCRU20;ENCTOT19;ENCTOT18;ENCTOT17;ENCTOT16;ENCTOT15;ENCTOT14;ENCTOT13;ENCTOT12;ENCTOT11;ENCITOT20;ENCIBE20;ENCIFZ20;ENCIGI20;ENCIJZ20;ENCIKZ20;ENCILZ20;ENCIMN20;ENCIOQ20;ENCIRU20;ENCITOT19;ENCITOT18;ENCITOT17;ENCITOT16;ENCITOT15;ENCITOT14;ENCITOT13;ENCITOT12;ENCITOT11;ETNTOT20;ETNBE20;ETNFZ20;ETNGI20;ETNJZ20;ETNKZ20;ETNLZ20;ETNMN20;ETNOQ20;ETNRU20;ETCTOT20;ETCBE20;ETCFZ20;ETCGI20;ETCJZ20;ETCKZ20;ETCLZ20;ETCMN20;ETCOQ20;ETCRU20;ETCTOT19;ETCBE19;ETCFZ19;ETCGI19;ETCJZ19;ETCKZ19;ETCLZ19;ETCMN19;ETCOQ19;ETCRU19;ETCTOT18;ETCBE18;ETCFZ18;ETCGI18;ETCJZ18;ETCKZ18;ETCLZ18;ETCMN18;ETCOQ18;ETCRU18;ETCTOT17;ETCBE17;ETCFZ17;ETCGI17;ETCJZ17;ETCKZ17;ETCLZ17;ETCMN17;ETCOQ17;ETCRU17;ETCTOT16;ETCBE16;ETCFZ16;ETCGI16;ETCJZ16;ETCKZ16;ETCLZ16;ETCMN16;ETCOQ16;ETCRU16;ETCTOT15;ETCBE15;ETCFZ15;ETCGI15;ETCJZ15;ETCKZ15;ETCLZ15;ETCMN15;ETCOQ15;ETCRU15;ETCTOT14;ETCBE14;ETCFZ14;ETCGI14;ETCJZ14;ETCKZ14;ETCLZ14;ETCMN14;ETCOQ14;ETCRU14;ETCTOT13;ETCBE13;ETCFZ13;ETCGI13;ETCJZ13;ETCKZ13;ETCLZ13;ETCMN13;ETCOQ13;ETCRU13;ETCTOT12;ETCBE12;ETCFZ12;ETCGI12;ETCJZ12;ETCKZ12;ETCLZ12;ETCMN12;ETCOQ12;ETCRU12;ETCTOT11;ETCBE11;ETCFZ11;ETCGI11;ETCJZ11;ETCKZ11;ETCLZ11;ETCMN11;ETCOQ11;ETCRU11;HT21;HT021;HT121;HT221;HT321;HT421;HT521;HTCH21;HTCH021;HTCH121;HTCH221;HTCH321;HTCH421;HTCH521;CPG21;CPG021;CPG121;CPG221;CPG321;CPG421;CPG521;CPGE21;CPGE021;CPGE121;CPGE221;CPGE321;CPGE421;CPGE521;CPGEL21;CPGEL021;CPGEL121;CPGEL221;CPGEL321;CPGEL421;CPGEL521;CPGEO21;CPGEO021;CPGEO121;CPGEO221;CPGEO321;CPGEO421;CPGEO521;VV21;VVUH21;VVLIT21;RT21;RTUH21;RTLIT21;AJCS21;AJCSUH21;AJCSLIT21


def load_communes(meta_dossier_file: str = metaDossierFile, dossier_complet_file: str = dossierCompletFile):
    global metaDossier, dossierComplet
    if (metaDossier is None) or (dossierComplet is None):
        downloadFile(metaDossierSourceFile, dossierCompletFile, zip=True, zipped_file="dossier_complet.csv")
        print_blue("Lecture Meta Donnees Communes : " + metaDossierFile + " ...")
        metaDossier = pd.read_csv(meta_dossier_file, delimiter=';', index_col=0)
        print_blue("Lecture Donnees Communes : " + dossier_complet_file + " ...")
        dossierComplet = pd.read_csv(dossier_complet_file, delimiter=';', dtype={"CODGEO": "string"}, index_col=0)
    return metaDossier, dossierComplet

#####################
### Artificialisation
#####################

artificialisationSourcePage = "https://artificialisation.biodiversitetousvivants.fr/les-donnees-au-1er-janvier-2020"
artificialisationSourceFile = "https://cerema.app.box.com/v/pnb-action7-indicateurs-ff/file/862179205781"
artificialisationSourceMeta = "https://artificialisation.biodiversitetousvivants.fr/sites/artificialisation/files/fichiers/2021/08/description%20indicateurs%202009%202020.pdf"
dossierArtificialisationFile = data_dir + "obs_artif_conso_com_2009_2020_V2.csv"
dossierArtificialisation = None
global_context["URL_SOURCE_ARTIFICIALISATION"] = artificialisationSourcePage

# idcom,idcomtxt,idreg,idregtxt,iddep,iddeptxt,epci20,epci20txt,aav2020,libaav2020,cateaav2020,
# naf09art10,art09act10,art09hab10,art09mix10,art09inc10,
# naf10art11,art10act11,art10hab11,art10mix11,art10inc11,
# naf11art12,art11act12,art11hab12,art11mix12,art11inc12,
# naf12art13,art12act13,art12hab13,art12mix13,art12inc13,
# naf13art14,art13act14,art13hab14,art13mix14,art13inc14,
# naf14art15,art14act15,art14hab15,art14mix15,art14inc15,
# naf15art16,art15act16,art15hab16,art15mix16,art15inc16,
# naf16art17,art16act17,art16hab17,art16mix17,art16inc17,
# naf17art18,art17act18,art17hab18,art17mix18,art17inc18,
# naf18art19,art18act19,art18hab19,art18mix19,art18inc19,
# naf19art20,art19act20,art19hab20,art19mix20,art19inc20,
# nafart0920,artact0920,arthab0920,artmix0920,artinc0920, artcom0920,
# pop12,pop17,pop1217,men12,men17,men1217,emp17,emp12,emp1217,mepart1217,menhab1217,artpop1217,surfcom20


def load_artificialisation(dossier_artificialisation_file: str = dossierArtificialisationFile):
    global dossierArtificialisation
    downloadFile(artificialisationSourceFile, dossierArtificialisationFile)
    if (dossierArtificialisation is None):
        print_blue("Lecture Donnees Artificialisation : " + dossier_artificialisation_file + " ...")
        dossierArtificialisation = pd.read_csv(dossier_artificialisation_file, delimiter=',', index_col=0, dtype={"idcom": str, "iddep": str, "epci20": str, "aav2020" : str,
              "naf09art10" : float, "art09act10" : float, "art09hab10" : float, "art09mix10" : float, "art09inc10" : float,
              "naf10art11" : float, "art10act11" : float, "art10hab11" : float, "art10mix11" : float, "art10inc11" : float,
              "naf11art12" : float, "art11act12" : float, "art11hab12" : float, "art11mix12" : float, "art11inc12" : float,
              "naf12art13" : float, "art12act13" : float, "art12hab13" : float, "art12mix13" : float, "art12inc13" : float,
              "naf13art14" : float, "art13act14" : float, "art13hab14" : float, "art13mix14" : float, "art13inc14" : float,
              "naf14art15" : float, "art14act15" : float, "art14hab15" : float, "art14mix15" : float, "art14inc15" : float,
              "naf15art16" : float, "art15act16" : float, "art15hab16" : float, "art15mix16" : float, "art15inc16" : float,
              "naf16art17" : float, "art16act17" : float, "art16hab17" : float, "art16mix17" : float, "art16inc17" : float,
              "naf17art18" : float, "art17act18" : float, "art17hab18" : float, "art17mix18" : float, "art17inc18" : float,
              "naf18art19" : float, "art18act19" : float, "art18hab19" : float, "art18mix19" : float, "art18inc19" : float,
              "naf19art20" : float, "art19act20" : float, "art19hab20" : float, "art19mix20" : float, "art19inc20" : float,
              "nafart0920" : float, "artact0920" : float, "arthab0920" : float, "artmix0920" : float, "artinc0920" : float, "artcom0920" : float,
              })
    return dossierArtificialisation


# IDCOM	IDCOMTXT	IDREG	IDREGTXT	IDDEP	IDDEPTXT	EPCI20	EPCI20TXT	AAV2020	LIBAAV2020	CATEAAV202
# NAF09ART10	ART09ACT10	ART09HAB10	ART09MIX10	ART09INC10
# NAF10ART11	ART10ACT11	ART10HAB11	ART10MIX11	ART10INC11
# NAF11ART12	ART11ACT12	ART11HAB12	ART11MIX12	ART11INC12
# NAF12ART13	ART12ACT13	ART12HAB13	ART12MIX13	ART12INC13
# NAF13ART14	ART13ACT14	ART13HAB14	ART13MIX14	ART13INC14
# NAF14ART15	ART14ACT15	ART14HAB15	ART14MIX15	ART14INC15
# NAF15ART16	ART15ACT16	ART15HAB16	ART15MIX16	ART15INC16
# NAF16ART17	ART16ACT17	ART16HAB17	ART16MIX17	ART16INC17
# NAF17ART18	ART17ACT18	ART17HAB18	ART17MIX18	ART17INC18
# NAF18ART19	ART18ACT19	ART18HAB19	ART18MIX19	ART18INC19
# NAF19ART20	ART19ACT20	ART19HAB20	ART19MIX20	ART19INC20
# NAFART0920	ARTACT0920	ARTHAB0920	ARTMIX0920	ARTINC0920  ARTCOM0920
# POP12	POP17	POP1217	MEN12	MEN17	MEN1217	EMP17	EMP12	EMP1217	MEPART1217	MENHAB1217	ARTPOP1217	SURFCOM20
dossierArtificialisationPacaFile = data_dir + "Fichier_CEREMA_Art_2009-2020_PACA.xls"
dossierArtificialisationPaca = None


def load_artificialisationPaca(dossier_artificialisationPaca_file: str = dossierArtificialisationPacaFile):
    global dossierArtificialisationPaca
    if (dossierArtificialisationPaca is None):
        print_blue("Lecture Donnees Artificialisation Paca : " + dossier_artificialisationPaca_file + " ...")
        xls = pd.ExcelFile(dossier_artificialisationPaca_file)
        dossierArtificialisationPaca = pd.read_excel(xls, index_col=0, dtype={"IDCOM": str})
    return dossierArtificialisationPaca

#############
### Flux 2018
#############

flux2018SourcePage = "https://www.insee.fr/fr/statistiques/5393826"
flux2018SourceFile = "https://www.insee.fr/fr/statistiques/fichier/5393826/base-csv-flux-mobilite-residentielle-2018.zip"
flux2018SourceMeta = "https://www.insee.fr/fr/statistiques/5393826#dictionnaire"

flux2018SourceData = data_dir + "base-flux-mobilite-residentielle-2018.csv"
flux2018 = None


def load_flux_2018(p_flux2018SourceFile:  str = flux2018SourceData):
    global flux2018
    if (flux2018 is None) :
        print_blue("Lecture Flux Mobilite Residentielle 2018 : " + p_flux2018SourceFile + " ...")
        flux2018 = pd.read_csv(p_flux2018SourceFile, delimiter=';', dtype={'CODGEO': str, 'LIBGEO': str,
                                                                         'DCRAN': str, 'L_DCRAN': str,
                                                                         'NBFLUX_C18_POP01P' : float })
    return flux2018

#############
### Flux 2017
#############

flux2017SourcePage = "https://www.insee.fr/fr/statistiques/4509335"
flux2017SourceFile = "https://www.insee.fr/fr/statistiques/fichier/4509335/base-csv-flux-mobilite-residentielle-2017.zip"
flux2017SourceMeta = "https://www.insee.fr/fr/statistiques/4509335#dictionnaire"

flux2017SourceData = data_dir + "base-flux-mobilite-residentielle-2017.csv"
flux2017 = None


def load_flux_2017(p_flux2017SourceFile:  str = flux2017SourceData):
    global flux2017
    if (flux2017 is None) :
        print_blue("Lecture Flux Mobilite Residentielle 2017 : " + p_flux2017SourceFile + " ...")
        flux2017 = pd.read_csv(p_flux2017SourceFile, delimiter=';', dtype={'CODGEO': str, 'LIBGEO': str,
                                                                         'DCRAN': str, 'L_DCRAN': str,
                                                                         'NBFLUX_C17_POP01P' : float })
    return flux2017

#############
### Flux 2014-15-16
#############

flux2016SourceFile = data_dir + "base-texte-flux-mobilite-residentielle-2016.csv"
flux2016 = None


def load_flux_2016(flux2016SourceFile:  str = flux2016SourceFile):
    global flux2016
    if (flux2016 is None) :
        print_blue("Lecture Flux Mobilite Residentielle 2016 : " + flux2016SourceFile + " ...")
        flux2016 = pd.read_csv(flux2016SourceFile, delimiter=';', encoding = "ISO-8859-1",
                                                                  dtype={'CODGEO': str, 'LIBGEO': str,
                                                                         'DCRAN': str, 'L_DCRAN': str,
                                                                         'NBFLUX_C16_POP01P' : float })
    return flux2016

flux2015SourceFile = data_dir + "base-texte-flux-mobilite-residentielle-2015.csv"
flux2015 = None


def load_flux_2015(flux2015SourceFile:  str = flux2015SourceFile):
    global flux2015
    if (flux2015 is None) :
        print_blue("Lecture Flux Mobilite Residentielle 2015 : " + flux2015SourceFile + " ...")
        flux2015 = pd.read_csv(flux2015SourceFile, delimiter=';', encoding = "ISO-8859-1",
                                                                  dtype={'CODGEO': str, 'LIBGEO': str,
                                                                         'DCRAN': str, 'L_DCRAN': str,
                                                                         'NBFLUX_C15_POP01P' : float})
    return flux2015

flux2014SourceFile = data_dir + "base-texte-flux-mobilite-residentielle-2014.csv"
flux2014 = None


def load_flux_2014(flux2014SourceFile:  str = flux2014SourceFile):
    global flux2014
    if (flux2014 is None) :
        print_blue("Lecture Flux Mobilite Residentielle 2014 : " + flux2014SourceFile + " ...")
        flux2014 = pd.read_csv(flux2014SourceFile, delimiter=';', encoding = "ISO-8859-1",
                                                                  dtype={'CODGEO': str, 'LIBGEO': str,
                                                                         'DCRAN': str, 'L_DCRAN': str,
                                                                         'NBFLUX_C14_POP01P' : float })
    return flux2014


######################
### Flux Travail 2018
######################

fluxpro2018SourcePage = "https://www.insee.fr/fr/statistiques/5393835"
fluxpro2018SourceFile = "https://www.insee.fr/fr/statistiques/fichier/5393835/base-csv-flux-mobilite-domicile-lieu-travail-2018.zip"
fluxpro2018SourceMeta = "https://www.insee.fr/fr/statistiques/5393835#dictionnaire"

fluxpro2018SourceData = data_dir + "base-flux-mobilite-domicile-lieu-travail-2018.csv"
fluxpro2018 = None

# CODGEO;LIBGEO;DCLT;L_DCLT;NBFLUX_C18_ACTOCC15P
def load_fluxpro_2018(p_fluxpro2018SourceFile:  str = fluxpro2018SourceData):
    global fluxpro2018
    if (fluxpro2018 is None) :
        print_blue("Lecture Flux Mobilite Domicile Travail 2018 : " + p_fluxpro2018SourceFile + " ...")
        fluxpro2018 = pd.read_csv(p_fluxpro2018SourceFile, delimiter=';', dtype={'CODGEO': str, 'LIBGEO': str,
                                                                         'DCRAN': str, 'L_DCRAN': str,
                                                                         'NBFLUX_C18_POP01P' : float })
    return fluxpro2018

######################
### Pop Communale 2019
######################

pop2019SourcePage = "https://www.insee.fr/fr/statistiques/6011070?sommaire=6011075"
pop2019SourceFile = "https://www.insee.fr/fr/statistiques/fichier/6011070/ensemble.zip"

pop2019SourceData = data_dir + "donnees_communes_pop_2019.csv"
pop2019 = None

# REG;DEP;COM;PMUN;PCAP;PTOT
def load_pop2019(p_pop2019SourceFile:  str = pop2019SourceData):
    global pop2019
    if (pop2019 is None) :
        print_blue("Lecture Populations Légales 2019 : " + p_pop2019SourceFile + " ...")
        pop2019 = pd.read_csv(p_pop2019SourceFile, index_col=2, delimiter=';', dtype={'REG': str, 'DEP': str,
                                                                       'COM': str, 'PMUN': int,
                                                                       'PCAP' : int, 'PTOT' : int })
    return pop2019

##############
### Stock 2018
##############

stock2018SourcePage = "https://www.insee.fr/fr/statistiques/5542867?sommaire=5395764"
stock2018SourceFile = "https://www.insee.fr/fr/statistiques/fichier/5542867/RP2018_LOGEMTZE_csv.zip"
stock2018SourceMeta = "https://www.insee.fr/fr/statistiques/fichier/5542867/dictionnaire_logemt_rp2018.pdf"

stock2018SourceData = data_dir + "FD_LOGEMTZE_2018.csv"
stock2018 = None

def load_stock_2018(p_stock2018SourceFile:  str = stock2018SourceData):
    global stock2018
    if (stock2018 is None) :
        downloadFile(stock2018SourceFile, stock2018SourceData, zip=True, zipped_file="FD_LOGEMTZE_2018.csv")
        print_blue("Lecture Besoins de Logements en Stock 2018 : " + stock2018SourceFile + " ...")
        stock2018 = pd.read_csv(p_stock2018SourceFile, delimiter=';', dtype={'CODGEO': str, 'LIBGEO': str,
                                                                         'DCRAN': str, 'L_DCRAN': str,
                                                                         'NBFLUX_C18_POP01P' : float })
    return stock2018


###
### Collect Data Description
###

collectDataFile     = configurationFile
collectDataMetrics  = None
collectDiagnostics  = None
collectCalculations = None
SCOT_OUEST          = None

# Key	Description	Source	Type	Data	Total
# Key	Description	Test    Message


def load_collectData(collect_file: str = collectDataFile):
    global collectDataMetrics, collectDiagnostics, collectCalculations, SCOT_OUEST
    if (collectDataMetrics is None) :
        print_blue("Lecture Data Metrics : " + collect_file + " ...")
        xls = pd.ExcelFile(collect_file)
        collectDataMetrics  = pd.read_excel(xls, 'Collect',    index_col=0, dtype=str)
        collectDiagnostics  = pd.read_excel(xls, 'Diagnostic', index_col=0, dtype=str)
        collectCalculations = pd.read_excel(xls, 'Calculs',    index_col=0, dtype=str)
        SCOT_OUEST          = pd.read_excel(xls, 'SCOT_OUEST', index_col=0, dtype={'Budget 2030': float, 'Budget 2040': float})

        print_yellow("  - Generation : " + "datametrics.json" + " ...")
        dm = pd.read_excel(xls, 'Collect', index_col=0, dtype=str).fillna('')
        dm = dm[~dm.index.duplicated(keep='first')]
        for index, row in dm.iterrows():
            if str(index) == "nan" or str(index).startswith("#")  : dm.drop(index, inplace=True)
        with open(output_dir + "datametrics.json", 'w') as f:
            d = dict()
            d["X"] = dm.to_dict(orient='index')
            d["Y"] = dm.to_dict()
            f.write(to_json(d, indent=4))
            f.close()

        print_yellow("  - Generation : " + "calculations.json" + " ...")
        dm = pd.read_excel(xls, 'Calculs', index_col=0, dtype=str).fillna('')
        dm = dm[~dm.index.duplicated(keep='first')]
        dm = dm[~dm.index.duplicated(keep='first')]
        for index, row in dm.iterrows():
            if str(index) == "nan" or str(index).startswith("#")  : dm.drop(index, inplace=True)
        with open(output_dir + "calculations.json", 'w') as f:
            d = dict()
            d["X"] = dm.to_dict(orient='index')
            d["Y"] = dm.to_dict()
            f.write(to_json(d, indent=4))
            f.close()

        print_yellow("  - Generation : " + "diagnostics.json" + " ...")
        dm = pd.read_excel(xls, 'Diagnostic', index_col=0, dtype=str).fillna('')
        dm = dm[~dm.index.duplicated(keep='first')]
        for index, row in dm.iterrows():
            if str(index) == "nan" or str(index).startswith("#")  : dm.drop(index, inplace=True)
        with open(output_dir + "diagnostics.json", 'w') as f:
            d = dict()
            d["X"] = dm.to_dict(orient='index')
            d["Y"] = dm.to_dict()
            f.write(to_json(d, indent=4))
            f.close()

    return collectDataMetrics

###
### File Management
###


def delete_pattern(directory=output_dir, pattern="*.bck"):
    # Get a list of all the file paths that ends with .txt from in specified directory
    fileList = glob.glob(directory+pattern)
    # Iterate over the list of filepaths & remove each file.
    for filePath in fileList:
        try:
            os.remove(filePath)
        except:
            print("Error while deleting file : ", filePath)

###
### Display in Browser
###

DISPLAY_HTML = True


def display_in_browser(html_file):
    if DISPLAY_HTML is False : return
    webbrowser.open_new_tab(html_file)


###
### Utils
###


def downloadFile(url: str, filename: str, zip=False, zipped_file: str = None) -> str:
    """ Download Files and Unzip"""
    if (os.path.isfile(filename)): return filename
    print_red("Downloading : "+filename)
    local_filename = filename
    with requests.get(url, stream=True) as r:
        r.raise_for_status()
        with open(local_filename, 'wb') as f:
            i = 1
            for chunk in r.iter_content(chunk_size=8192):
                i = i + 1
                print(".", end="")
                if i == 150:
                    print("")
                    i = 1
                f.write(chunk)
        f.close()
        print("")
    if (zip):
        print_red("Unzipping : " + filename)
        with zipfile.ZipFile(local_filename, 'r') as zip_ref:
            zip_ref.extractall("zip")
        shutil.move('zip' + os.sep + zipped_file, filename)
    print_red("Extracted : " + local_filename)
    return local_filename


def nan0(value, default=0) :
    if (value != value) : return default
    return value


def error0(value, default=0) :
    if ("error" in str(value).lower()) :
        logging.debug("error0 on value : "+str(value))
        return default
    return value


def round0(value, rounding=0) -> Union[float, int] :
    if (value != value) : value = 0  # NaN
    if (rounding == 0): return int(round(value, 0))
    return round(value, rounding)


def roundNumber(value, rounding=0) -> Union[float, int] :
    return round0(value, rounding)


def round0str(value, rounding=0) -> str :
    val = round0(value, rounding)
    if (rounding == 0) : return str(val)
    if (rounding == 1) : return f"{val:.1f}"
    if (rounding == 2) : return f"{val:.2f}"
    if (rounding == 3) : return f"{val:.3f}"
    if (rounding == 4) : return f"{val:.4f}"
    if (rounding == 5) : return f"{val:.5f}"
    return f"{val:.6}"


def perCent(value, rounding=0) -> float :
    return round0(value * 100, rounding)


def perCentStr(value, rounding=0) -> str :
    return round0str(value * 100, rounding)+"%"


def clean_name(name: str,sep="_") -> str:
    return unidecode.unidecode(name).replace(" ", sep).replace("\\", sep).replace("'", sep)


def str_list(the_list: list,sep=", ") -> str:
    string = ""
    s = ""
    for elem in the_list :
        string = string + s + str(elem)
        s = sep
    return string


### Print
def print_green(text):
    print(colored(text, "green"))
    logging.debug(text)


def print_red(text):
    print(colored(text, "red"))
    logging.debug(text)


def print_verbose(text):
    global VERBOSE
    if (VERBOSE) :
        print(colored(text, "magenta"))
    logging.debug(text)


def print_error(text):
    print(colored(text, "cyan"))
    logging.error(text)


def print_yellow(text):
    print(colored(text, "yellow"))
    logging.debug(text)


def print_grey(text):
    print(colored(text, "grey"))
    logging.debug(text)


def print_blue(text):
    print(colored(text, "blue"))
    logging.debug(text)


def print_white(text):
    print(colored(text, "white"))
    logging.debug(text)


def print_data(p_data_dict: dict):
    for l_key in p_data_dict.keys():
        print_white(l_key + " : " + str(p_data_dict[l_key]))


def print_data_frame(p_data_frame: pd.DataFrame, code_insee: str):
    for column in p_data_frame:
        print_white(column + " - " + str(p_data_frame[column]["meta"]) + " : " + str(p_data_frame[column][code_insee]))


def print_commune(commune):
    if (isinstance(commune, int)):
        print_white(str(commune)+" : "+nom_commune(commune))
        return
    for comm in commune :
        print_white(str(comm)+" : "+nom_commune(comm))
    return


def print_epci(ecpi) -> None :
    if (isinstance(ecpi, int)):
        print_white(str(ecpi)+" : "+nom_epci(ecpi))
        return
    for ep in ecpi :
        print_white(str(ep)+" : "+nom_epci(ep))
    return


def print_dept(dept) -> None :
    if (isinstance(dept, int)):
        print_white(str(dept)+" : "+nom_dept(dept))
        return
    for dp in dept :
        print_white(str(dp)+" : "+nom_dept(dp))
    return


def to_json(obj,  indent=4):
    return(jsonc.dumps(obj,  indent=indent))


def to_yaml(obj,  indent=4):
    return(yaml.safe_dump(obj,  indent=indent, default_flow_style=False))


def save_file(data, file_name: str):
    try:
        with open(file_name, 'w') as outfile:
            outfile.write(str(data))
            outfile.close()
    except Exception as e:
        print_red("Error Writing File : " + file_name + "-" + str(e))
        raise


## Save figure for HTML Format
def fig_to_base64(figure):
    img = io.BytesIO()
    figure.savefig(img, format='png', bbox_inches='tight')
    img.seek(0)
    return base64.b64encode(img.getvalue())


## Calculate Population Evolution and Rate
def calc_after(annee_depart : int, pop_depart, annee_arrivee : int, taux_croissance, rounding=3) -> float :
    """ Retourne l'after )) : P1 = P0 x ( 1 + T/100) puissance N """
    annee = annee_arrivee - annee_depart
    pop_arrivee = pop_depart * (1 + taux_croissance / 100) ** annee
    return round0(pop_arrivee, rounding)


def calc_taux(annee_depart : int, pop_depart, annee_arrivee : int, pop_arrivee, rounding=3) -> float :
    """ Retourne le taux  : T/100 = [ P1/P0 ] puissance 1/N - 1 """
    annee = annee_arrivee - annee_depart
    taux_croissance = (pop_arrivee / pop_depart) ** (1 / annee)
    return round0(- (1 - taux_croissance) * 100, rounding)


def taux(part, total, rounding=3) -> float :
    """ Retourne le taux """
    return round0(part/total, rounding)


def f_percent(part, full, rounding=1, suffix="%", format=""):
    """ Retourne formatted part in percent of full     Part = 90, Full = 200 => 45 """
    # Adds suffix: suffix="%"   =>   45%
    # Adds format: format="+"   =>  +45%
    # Adds format: format="+()" =>  (+45%) 90
    # Adds format: format="()"  =>   (45%) 90
    percent = 0 if (float(full) == 0)  else (float(part) / float(full)) * 100
    s_percent = round0str(percent, rounding)
    if (format == ""):           return s_percent+suffix
    if (("+" in format) and (percent > 0)): s_percent = "+"+s_percent
    if ("(" in format): return "("+s_percent+suffix+") "+str(part)
    return "("+s_percent+suffix+") " + str(part)


def f_diff(after, before, format="+"):
    """  Adds format: format="+"   =>  +45% """
    # After = 200, Before = 100 => 100
    diff = after - before
    if (("+" in format) and (float(diff) > 0)): return ("+"+str(diff))
    return ""+str(diff)


def f_val(val, format="+", suffix=""):
    if (("+" in format) and (float(val) > 0)): return ("+"+str(val)+suffix)
    return ""+str(val)+suffix


def f_taux(value, rounding=2, suffix="%", format="+"):
    if ((float(value) > 0) and ("+" in format)): return ("+"+round0str(value, rounding)+suffix)
    return ""+round0str(value, rounding)+suffix


def f_round(value, rounding=0) -> str :
    return round0str(value, rounding)


## Data Access
def is_scot_ouest(code_insee) -> bool:
    if (code_insee in communes_zone("SCoT_Ouest")): return True
    if (str(code_insee) == "200039915"): return True
    return False


def is_commune_hors_paca(code_insee) -> bool:
    global COMMUNES_HORS_PACA
    load_scot_data()
    for index, row in COMMUNES_HORS_PACA.iterrows():
        if (index == 0): continue
        COMM = row['Liste des communes hors PACA appartenant à des EPCI de PACA']
        if (COMM == code_insee) : return True
    return False


def get_code_insee_commune(code_postal) -> Union[int, str]:
    """ Retourne le code insee et le nom de la commune """
    load_codes()
    the_list = codesPostaux.index[codesPostaux['Code_postal'] == code_postal].tolist()
    if (not the_list):
        return "Pas de Code INSEE pour code Postal "+str(code_postal)
    code_insee = the_list[0]
    commune = codesPostaux["Nom_commune"][code_insee]
    if not isinstance(commune, str) : commune = commune[0]
    return code_insee, commune


## Data Access
def is_code_insee_commune(code_insee) -> bool:
    """ Retourne True si le code insee est celui d'une commune """
    load_codes()
    return (code_insee in codesPostaux.index)


def get_code_insee_commune_nom(nom_commune) -> int :
    """ Retourne le code insee et le nom de la commune """
    load_codes()
    the_list = codesPostaux.index[codesPostaux['Nom_commune'] == nom_commune.upper()].tolist()
    if (not the_list):
        return 0
    code_insee = the_list[0]
    return code_insee


def get_code_postal_commune(code_insee) -> Union[int, str]:
    """ Retourne le code postal et le nom de la commune """
    load_codes()
    try:
        code_postal = codesPostaux["Code_postal"][code_insee]
    except Exception as e:
        code_postal = "Pas de Code Postal pour Code INSEE "+str(code_insee)
        commune     = "Pas de Commune pour Code INSEE "+str(code_insee)
        return code_postal, commune
    if (not isinstance(code_postal, str)):
        code_postal = code_postal[0]
        commune     = codesPostaux["Nom_commune"][code_insee][0]
        return code_postal, commune
    commune     = codesPostaux["Nom_commune"][code_insee]
    return code_postal, commune


def get_code_postal(code_insee) -> Union[int, str]:
    code_postal, comm = get_code_postal_commune(code_insee)
    return code_postal


def get_code_insee(code_postal : Union[int, str]) -> int:
    """ Le Code INSEE du Code Postal """
    code_insee = get_code_insee_commune(code_postal)
    if str(code_insee).startswith("Pas de"): return 0
    return code_insee


def get_gps_insee(code_insee : Union[int, str]) -> Union[str, str]:
    """ Le GPS Coord INSEE du Code Postal """
    load_codes()
    if (code_insee not in codesPostaux["Code_postal"]) :  return "", ""
    if (isinstance(codesPostaux["coordonnees_gps"][code_insee], str)):
        gps = codesPostaux["coordonnees_gps"][code_insee]
    else:
        gps = codesPostaux["coordonnees_gps"][code_insee][0]
    lat  = gps.split(",")[0]
    long = gps.split(",")[1]
    return lat, long


def get_gps_lat_insee(code_insee : Union[int, str]) -> str:
    lat, long = get_gps_insee(code_insee)
    return lat


def get_gps_long_insee(code_insee : Union[int, str]) -> str:
    lat, long = get_gps_insee(code_insee)
    return long


def nom_epci(code_epci, clean=False) -> str:
    """ Le nom de l'EPCI """
    if (code_epci == 0) : return "Pas de Nom"
    load_interco()
    epci_list    = intercoDossier[intercoDossier["Unnamed: 2"] == code_epci].head(1)["Unnamed: 3"]
    if epci_list.empty : return None
    if (not clean): return epci_list[0]
    return clean_name(epci_list[0])


def code_epci(nom_epci : str) -> str:
    """ Le code de l'EPCI """
    load_interco()
    if (nom_epci == None) or (nom_epci.strip() == "") : return 0
    epci_list    = intercoDossier[intercoDossier["Unnamed: 3"] == nom_epci].head(1)["Unnamed: 2"]
    if epci_list.empty : return 0
    return epci_list[0]


def code_commune(nom_commune : str) -> str:
    """ Le code de la Commune """
    load_interco()
    if (nom_commune == None) or (nom_commune.strip() == "") : return 0
    comm_list    = intercoDossier[intercoDossier["Unnamed: 1"] == nom_commune].head(1)["Unnamed: 2"]
    if comm_list.empty : return 0
    return comm_list[0]


def nom_dept(code_dept, clean=False) -> str:
    """ Le nom du Departement """
    if (code_dept == 0) : return "Pas de Nom"
    load_departements()
    nom = departements["nom_departement"][code_dept]
    if (not clean): return nom
    return clean_name(nom)


def code_dept(nom_dept) -> str:
    """ Le code du Departement """
    if (nom_dept == None) : return 0
    load_departements()
    if (len(departements[departements["nom_departement"]==nom_dept])==0): return 0
    return departements[departements["nom_departement"]==nom_dept].index[0]


def nom_commune(code_insee=None, code_postal=None, clean=False) -> str:
    """ Le nom de la commune """
    load_codes()
    commune = "Pas de nom"
    if   (code_insee)  : code_postal, commune = get_code_postal_commune(code_insee)
    elif (code_postal) : code_insee,  commune = get_code_insee_commune(code_postal)
    nom = str(commune).title()
    if (not clean): return nom
    return clean_name(nom)


def nom_region(code_region, clean=False) -> str:
    """ Le nom de Region """
    if (code_region == 0) : return "Pas de Nom"
    load_departements()
    nom     = departements[departements["code_region"] == code_region]['nom_region'].to_list()[0]
    if (not clean): return nom
    return clean_name(nom)


def communes_epci(code_epci) -> list[int]:
    """ Les Codes INSEE des Communes d'une EPCI """
    load_interco()
    index = intercoDossier.index
    condition = intercoDossier["Unnamed: 2"] == code_epci
    epci_indices = index[condition]
    epci_indices_list = epci_indices.tolist()
    return sorted(epci_indices_list)


def list_nom_communes(codes : list) -> list[str]:
    """ Les noms des Communes de la liste des codes INSEE """
    list_nom = list()
    for code in codes :
        list_nom.append(nom_commune(code_insee=code))
    return list_nom


def list_nom_epci(codes : list) -> list[str]:
    """ Les noms des EPCI de la liste des codes INSEE """
    list_nom = list()
    for code in codes :
        list_nom.append(nom_epci(code, clean=False))
    return list_nom


def list_nom_zone(codes : list) -> list[str]:
    """ Les noms des Zones de la liste des codes Zone """
    list_nom = list()
    for code in codes :
        list_nom.append(nom_zone(code, clean=False))
    return list_nom


def communes_dept(code_dept) -> list[int]:
    """ Les Codes INSEE des Communes d'un Departement """
    load_interco()
    index = intercoDossier.index
    condition = intercoDossier["Unnamed: 4"] == code_dept
    epci_indices = index[condition]
    epci_indices_list = epci_indices.tolist()
    return sorted(epci_indices_list)


def communes_region(code_reg) -> list[int]:
    """ Les Codes INSEE des Communes d'une Region """
    load_interco()
    index = intercoDossier.index
    condition    = intercoDossier["Unnamed: 5"] == code_reg
    epci_indices = index[condition]
    epci_indices_list = epci_indices.tolist()
    return sorted(epci_indices_list)


def list_zones() -> list[str]:
    """ List des Zones """
    return scot_consolidation()["GROUPES_COMMUNES"].keys()


def list_zones_dept(code_dept="06") -> list[str]:
    """ List des Zones dans DEPT """
    global SCOT_DATA
    code_dept = str(code_dept)
    list_zone = list()
    for zone in list_zones() :
        if (SCOT_DATA["GROUPES_COMMUNES"][zone]["DEPT"] == code_dept) :
            list_zone.append(zone)
    return list_zone


def communes_zone(nom_zone) -> list[int]:
    """ Les Codes INSEE des Communes d'une Zone """
    SCOT_DATA = scot_consolidation()
    if (nom_zone not in list_zones()) : return []
    return SCOT_DATA["GROUPES_COMMUNES"][nom_zone]["COMMUNES"]


def dept_zone(nom_zone) :
    """ Le Code DEPT d'une Zone """
    SCOT_DATA = scot_consolidation()
    if (nom_zone in SCOT_DATA["GROUPES_COMMUNES"]) :
        return SCOT_DATA["GROUPES_COMMUNES"][nom_zone]["DEPT"]
    else:
        return "99"


def nom_zone(key, clean=False) -> str :
    """ Le nom d'une Zone """
    SCOT_DATA = scot_consolidation()
    if (key in SCOT_DATA["GROUPES_COMMUNES"]) :
        nom = SCOT_DATA["GROUPES_COMMUNES"][key]["NAME"]
        if (not clean): return nom
        return clean_name(nom)
    else:
        return "NZ"


def communes_territoire(territoire:str) -> [list[str], str]:
    """ Les Codes INSEE des Communes d'un Territoire, le type de territoire """
    # Territoire =  nom scot, code dept, nom dept
    #               code epci, nom epci, nom commune, code commune

    load_interco()

    territoire = territoire.strip()
    code_territoire = []
    type_territoire = ""
    if nom_epci(territoire):
        # Territoire = Nom epci
        code_territoire = communes_epci(territoire)
        type_territoire = "EPCI"
    elif code_epci(territoire):
        # Territoire = Code epci
        code_territoire = communes_epci(code_epci(territoire))
        type_territoire = "EPCI"
    elif get_code_insee_commune_nom(territoire):
        # Territoire = Nom Commune
        code_territoire = get_code_insee_commune_nom(territoire)
        type_territoire = "COMMUNE"
    elif is_code_insee_commune(territoire):
       # Territoire = Code INSEE Commune
       code_territoire = territoire
       type_territoire = "COMMUNE"
    elif get_code_insee(territoire):
        # Territoire = Code Postal Commune
        code_territoire, nom = get_code_insee_commune(territoire)
        type_territoire = "COMMUNE"
    elif len(communes_dept(territoire))!=0:
        # Territoire = Code Dept
        code_territoire = communes_dept(territoire)
        type_territoire = "DEPT"
    elif code_dept(territoire):
        # Territoire = Nom Dept
        code_territoire = communes_dept(code_dept(territoire))
        type_territoire = "DEPT"
    elif len(communes_region(territoire))!=0:
        # Territoire = Code Region
        code_territoire = communes_region(territoire)
        type_territoire = "REG"
    elif (territoire.lower()=="paca"):
        # Territoire = Code Region
        code_territoire = communes_region("93")
        type_territoire = "REG"

    if isinstance(code_territoire, list):
        return code_territoire, type_territoire
    else:
        return [code_territoire], type_territoire


def epci_dept(code_dept) -> list[int]:
    """ Les Codes INSEE des EPCI d'un Departement """
    load_interco()
    epci_list = intercoDossier[intercoDossier["Unnamed: 4"] == code_dept]["Unnamed: 2"]
    return sorted(list(set(epci_list)))


def dept_epci(code_epci) -> int :
    """ Le Code de Departement de l'EPCI """
    load_interco()
    epci_list = intercoDossier[intercoDossier["Unnamed: 2"] == code_epci]["Unnamed: 4"]
    if epci_list.empty : return 0
    return epci_list[0]


def region_epci(code_epci) -> int :
    """ Le Code de Region de l'EPCI """
    load_interco()
    epci_list = intercoDossier[intercoDossier["Unnamed: 2"] == code_epci]["Unnamed: 5"]
    if epci_list.empty : return 0
    return epci_list[0]


def region_dept(code_dept) -> int :
    """ Le Code de Region du Departement """
    load_interco()
    reg_list = intercoDossier[intercoDossier["Unnamed: 4"] == code_dept]["Unnamed: 5"]
    if reg_list.empty : return 0
    return reg_list[0]


def list_dept(code_region=None) -> list[int]:
    """ La liste des Departements """
    load_departements()
    liste_dept = sorted(list(map(str, list(set(departements.index.values.tolist())))))
    if (code_region):
        liste_dept_region = []
        for dept in liste_dept :
            if (str(region_dept(dept)) == str(code_region)):
                liste_dept_region.append(dept)
        return liste_dept_region
    else:
        return liste_dept


def epci_region(code_reg) -> list[int]:
    """ Les EPCI d'une Region """
    load_interco()
    epci_list    = intercoDossier[intercoDossier["Unnamed: 5"] == code_reg]["Unnamed: 2"]
    return sorted(list(set(epci_list)))


def epci_commune(code_commune) -> Union[int, None]:
    """ L'EPCI d'une Commune """
    load_interco()
    epci_list    = intercoDossier[intercoDossier.index == str(code_commune)]["Unnamed: 2"]
    if epci_list.size == 0 : return None
    return epci_list[0]


def list_region() -> list[str]:
    """ La liste des Regions """
    load_departements()
    return list(map(str, sorted(list(set(departements["code_region"])))))


def get_sru2017(key, code_insee, rounding=6):
    """ Retourne la valeur comsolidee de la cle SRU 17-19 pour la commune  """
    load_sru(sruFile)
    try:
        return round0(sru2017[key][int(code_insee)], rounding)
    except Exception as e:
        return 0


def get_sru2020(key, code_insee, rounding=6):
    """ Retourne la valeur comsolidee de la cle SRU 20-22 pour la commune  """
    load_sru(sruFile)
    try:
        return round0(sru2020[key][int(code_insee)], rounding)
    except Exception as e:
        return 0


def get_art(key, code_insee, rounding=6):
    """ Retourne la valeur comsolidee de la cle ART pour la commune  """
    load_artificialisation()
    try:
        return round0(dossierArtificialisation[key][code_insee], rounding)
    except Exception as e:
        return 0


def scot_consolidation() -> dict:
    """ La liste des Zones de Communes """
    global SCOT_DATA
    load_interco()
    load_scot_data()
    if ("GROUPES_COMMUNES" in SCOT_DATA): return SCOT_DATA
    SCOT_DATA["GROUPES_COMMUNES"] = {}
    for scot in SCOT_DATA["SCOT_EPCI"].keys():
        zone = clean_name(scot)
        SCOT_DATA["GROUPES_COMMUNES"][zone] = {}
        SCOT_DATA["GROUPES_COMMUNES"][zone]["TYPE"]  = "SCOT"
        SCOT_DATA["GROUPES_COMMUNES"][zone]["CLEAN"] = zone
        SCOT_DATA["GROUPES_COMMUNES"][zone]["NAME"]  = scot
        SCOT_DATA["GROUPES_COMMUNES"][zone]["COMMUNES"] = []
        SCOT_DATA["GROUPES_COMMUNES"][zone]["EPCI"] = []
        for epci in SCOT_DATA["SCOT_EPCI"][scot]:
            c_epci   = code_epci(epci)
            c_dept   = dept_epci(c_epci)
            c_reg    = region_epci(c_epci)
            communes = communes_epci(c_epci)
            SCOT_DATA["GROUPES_COMMUNES"][zone]["COMMUNES"].extend(communes)
            SCOT_DATA["GROUPES_COMMUNES"][zone]["EPCI"].append(c_epci)
            SCOT_DATA["GROUPES_COMMUNES"][zone]["DEPT"] = c_dept
            SCOT_DATA["GROUPES_COMMUNES"][zone]["REG"]  = c_reg
    for groupe in SCOT_DATA["GROUPEMENTS"].keys():
        zone = clean_name(groupe)
        territoires = SCOT_DATA["GROUPEMENTS"][groupe]["LIST"]
        SCOT_DATA["GROUPES_COMMUNES"][zone] = {}
        SCOT_DATA["GROUPES_COMMUNES"][zone]["TYPE"] = SCOT_DATA["GROUPEMENTS"][groupe]["TYPE"]
        SCOT_DATA["GROUPES_COMMUNES"][zone]["NAME"] = groupe
        SCOT_DATA["GROUPES_COMMUNES"][zone]["COMMUNES"] = []
        SCOT_DATA["GROUPES_COMMUNES"][zone]["EPCI"] = []
        for territoire in str(territoires).split(',') :
            # code commune, nom commune, code epci, nom epci, nom scot, code dept, nom dept
            communes, type = communes_territoire(territoire)
            if (communes is None) or (len(communes) == 0) :
                print_error("Territoire inconnu : " + territoire)
                continue
            c_epci   = epci_commune(communes[0])
            c_dept   = dept_epci(c_epci)
            c_reg    = region_epci(c_epci)
            SCOT_DATA["GROUPES_COMMUNES"][zone]["COMMUNES"] = sorted(list(set(SCOT_DATA["GROUPES_COMMUNES"][zone]["COMMUNES"]) | set(communes)))
            SCOT_DATA["GROUPES_COMMUNES"][zone]["EPCI"].append(c_epci)
            SCOT_DATA["GROUPES_COMMUNES"][zone]["DEPT"] = c_dept
            SCOT_DATA["GROUPES_COMMUNES"][zone]["REG"]  = c_reg
    save_file(to_json(SCOT_DATA,  indent=4), output_dir + os.sep + "a-scot.json")
    return SCOT_DATA

###
### Data Collection and Consolidation
###

source_data    = "DATA"
source_codes   = "CODE"
source_interco = "INTERCO"
source_insee   = "INSEE"
source_calc    = "CALC"
source_sru     = "SRU"
source_art     = "ART"
source_proj    = "PROJ"
source_evol    = "EVOL"
source_sit     = "SIT"

mode_sum       = "SUM"
mode_concat    = "CONCAT"
mode_ignore    = "IGNORE"
mode_equal     = "EQUAL"
mode_count     = "COUNT"
mode_custom    = "CUSTOM"
mode_na        = "N/A"
mode_average   = "AVG"
mode_max       = "MAX"
mode_min       = "MIN"

type_int       = "INT"
type_str       = "STR"
type_taux      = "TAUX"
type_percent   = "PERCENT"
type_float     = "FLOAT"

entite_commune = "COMMUNE"
entite_epci    = "EPCI"
entite_dept    = "DEPT"
entite_region  = "REGION"
entite_scot    = "SCOT"
entite_zone    = "ZONE"

DataStoreCache = {}


class DataStore():

    def __init__(self, store_name : str, store_type : str, store_code : str):
        self.store_name  = store_name   # Name of Commune, EPCI, DEPT, REGION ...
        self.store_type  = store_type   # Type of Entity for Store :  "COMMUNE", "EPCI", "DEPT", "REGION", "ZONE"
        self.store_code  = store_code   # Code INSEE for "COMMUNE", "EPCI", for pour "DEPT", "REGION", "ZONE"
        self.store_index = store_code   # Working Index : Code INSEE for "COMMUNE", "total" pour tables multi-communes
        if (self.store_type != entite_commune) : self.store_index = "total"
        self.data_frame  = pd.DataFrame()
        self.key_datas   = []  # Metric Keys
        self.meta_dict   = {}  # Semantic of this indicator / metric
        self.type_dict   = {}  # { "INT", "STR",    "TAUX",  "PERCENT", "FLOAT" }
        self.source_dict = {}  # { "SRU", "INSEE",  "ART",   "SIT",     "DATA", "PROJ", "EVOL", "CALC" }
        self.mode_dict   = {}  # { "SUM", "CONCAT", "IGNORE", "EQUAL", "COUNT", "CUSTOM", "N/A", "AVG", "MAX", "MIN", "SAME" }
        self.expr_dict   = {}  # Expression Used for Calculations
        self.error_dict  = {}  # Error while processing this indicator / metric
        self.metric_list = []  # List of Metrics for Rendering
        self.diagnostics = []  # List of Diagnostic for Rendering
        self.html        = None
        self.calc_done   = False
        self.diag_done   = False
        self.plot_done   = False

    def add_metric(self, key : str, meta : str, source: str, mode : str, type: str, index : str = None, data=None, expr="None"):
        """ Add a metric to the data store """
        self.key_datas.append(key)
        self.meta_dict[key]    = meta
        self.source_dict[key]  = source
        self.type_dict[key]    = type
        self.mode_dict[key]    = mode
        self.expr_dict[key]    = expr
        self.error_dict[key]   = "OK"
        self.metric_list.append({"key" : key , "meta" : meta , "source" : source , "type" : type, "mode" : mode, "expr" : expr})
        return self.add_data(key, index, data)

    def add_diagnostic(self, _key: str, _description: str, test: str, messageV: str, data : bool, messageF: str, cat: str, type: str):
        """ Add a diagnostic to the data store """
        diagnostic = {"key"  : _key , "description" : _description , "test" : test, "value" : data, "categorie" : cat,
                      "type" : type,  "messageSiFaux" : messageF, "messageSiVrai" : messageV}
        self.diagnostics.append(diagnostic)
        return diagnostic

    def add_data(self, key : str, index : str, data):
        """ Add a data to the data store - with type checking """
        if (data  is None): return
        if (key   is None): return
        if (index is None): index = self.store_index
        if (index is None): return
        try:
            if (self.type_dict[key]   == "STR")     : return self.add_value(key, index, str(data))
            elif (self.type_dict[key] == "INT")     : return self.add_value(key, index, int(round(data, 0)))
            elif (self.type_dict[key] == "FLOAT")   : return self.add_value(key, index, float(data))
            elif (self.type_dict[key] == "TAUX")    : return self.add_value(key, index, float(data))
            elif (self.type_dict[key] == "PERCENT") : return self.add_value(key, index, float(data))
            else:
                self.error_dict[key] = "Invalid Type  [" + str(index) + "][" + str(key) + "] - Not a " + str(self.type_dict[key]) + " :  " + str(data)
                print_red("else")
                print_red(self.error_dict[key])
                print_error("add_data Else : " + self.error_dict[key])
                self.add_value(key, index, str(data))
                # quit()
                return None
        except Exception as e:
            self.error_dict[key] = "Type Error [" + str(index) + "][" + str(key) + "] - Not a " + self.type_dict[key] + " :  " + str(data) + " : " + str(e)
            print_red("Exception")
            print_red(self.error_dict[key])
            print_error("add_data Exception : " + self.error_dict[key])
            self.add_value(key, index, str(data))
            # quit()
            return None

    def add_value(self, key : str, index : str, data):
        """ Add a raw value to the data store.  """
        self.data_frame.at[index, key] = data
        return data

    def __getitem__(self, key):
        return self.get(key=key, index=None)

    def __setitem__(self, key, value):
        return self.add_data(key, self.store_index, value)

    def get(self, key : str, index : str = None):
        """ Return a value from data store.  """
        if (key   is None): return None
        if (index is None): index = self.store_index
        if (index is None): return None
        return self.data_frame.at[index, key]

    def get_row_as_dict(self, index : str = None):
        """ Return the index raw as a dict with columm names as keys.  """
        if (index is None): index = self.store_index
        if (index is None): return None
        dc = {}
        for key in self.key_datas:
            dc[key] = self.get(key, index)
        return dc

    def get_fullname(self):
        """ Base file name for DataStore.  """
        if (self.store_type == "ZONE") :
            fullname = str(self.store_type) + "_" + str(self.store_name)
        else:
            fullname = str(self.store_type) + "_" + str(self.store_name) + "_" + str(self.store_code)
        return clean_name(fullname)

    def save_data(self):
        """ Save raw DataStore data in various formats (xlsx, json, csv)  """
        self.data_frame = self.data_frame.fillna('')
        if ("__builtins__" in self.data_frame.columns):
            self.data_frame.drop("__builtins__", axis=1,inplace=True)
        print_green("> Saving Data   : "+self.get_fullname())
        # DF TO JSON
        ## _s.json (Summary - used by GUI)
        print_yellow("  - Saving Data : "+self.get_fullname()+ "_s.json")
        self.data_frame.to_csv(output_dir + self.get_fullname() + "_s.json", sep=',')
        all = dict()
        summary_data = self.clean_data().fillna('')
        for name, values in summary_data.iteritems():
            all[name] = {}
            for name2, value2 in values.iteritems():
                all[name][name2] = value2
        with open(output_dir + self.get_fullname() + "_s.json", 'w') as f:
            data_s = summary_data.to_dict(orient='index')
            data_s["Data"] = all
            if ('__warningregistry__' in data_s["Data"]):
                data_s["Data"].pop('__warningregistry__', None)
            if ('__warningregistry__' in data_s["total"]):
                data_s["total"].pop('__warningregistry__', None)
            if ('__warningregistry__' in data_s["mode"]):
                data_s["mode"].pop('__warningregistry__', None)
            if ('__warningregistry__' in data_s["type"]):
                data_s["type"].pop('__warningregistry__', None)
            if ('__warningregistry__' in data_s["source"]):
                data_s["source"].pop('__warningregistry__', None)
            if ('__warningregistry__' in data_s["expr"]):
                data_s["expr"].pop('__warningregistry__', None)
            for k in data_s["Data"].keys():
                print_verbose(k)
                print_verbose(str(data_s["Data"][k]))
            data_s["Diagnostics"] = self.diagnostics
            global_context["JSON_DATA_SET_S"] = summary_data.to_dict(orient='index')
            try:
                f.write(to_json(data_s, indent=4))
            except Exception as e:
                print(str(e))
        # DF TO EXCEL
        if (not FAST):
            print_yellow("  - Saving Data : "+self.get_fullname()+ ".xlsx")
            writer = pd.ExcelWriter(output_dir + self.get_fullname() + ".xlsx")
            self.data_frame.to_excel(writer, "Data")
            pivot = self.data_frame.transpose()
            pivot.to_excel(writer, "Pivot")
            diag_df = self.data_frame.from_dict(self.diagnostics)
            diag_df.to_excel(writer, "Diagnotics")
            writer.save()
            writer.close()
        # DF TO CSV
        if (not FAST):
            print_yellow("  - Saving Data : "+self.get_fullname() + ".csv")
            self.data_frame.to_csv(output_dir + self.get_fullname() + ".csv", sep=',')
        # DF TO JSON
        ## _d.json (Diagnostics)
        if (not FAST and FAST):
            self.data_frame.to_csv(output_dir + self.get_fullname() + "_d.json", sep=',')
            with open(output_dir + self.get_fullname() + "_d.json", 'w') as f:
                global_context["JSON_DIAGNOSTICS"] = " { \"Diagnostic\" : "+to_json(jsonc.loads(jsonc.dumps(self.diagnostics)), indent=4) + "}"
                f.write(global_context["JSON_DIAGNOSTICS"])
        ## _m.json
        if (not FAST and FAST):
            print_yellow("  - Saving Data : "+self.get_fullname()+ "_m.json")
            all = {}
            for name, values in self.data_frame.iteritems():
                all[name] = {}
                for name2, value2 in values.iteritems():
                    all[name][name2] = value2
            with open(output_dir + self.get_fullname() + "_m.json", 'w') as f:
                global_context["JSON_DATA_SET_M"] = to_json(all, indent=4)
                f.write(global_context["JSON_DATA_SET_M"])
        ## _c.json
        if (not FAST and FAST):
            print_yellow("  - Saving Data : "+self.get_fullname()+ "_c.json")
            with open(output_dir + self.get_fullname() + "_c.json", 'w') as f:
                data_c = self.data_frame.fillna('').to_dict(orient='index')
                data_c["Data"] = all
                data_c["Diagnostics"] = self.diagnostics
                global_context["JSON_DATA_SET_C"] = self.data_frame.to_dict(orient='index')
                f.write(to_json(data_c, indent=4))


    def load_data(self):
        """ Load DataStore data from xlsx file  """
        # EXCEL to DF
        filename = output_dir + self.get_fullname() + ".xlsx"
        if not os.path.isfile(filename) : return None
        print_green("> Load Data : "+filename)
        xls = pd.ExcelFile(filename)
        self.data_frame  = pd.read_excel(xls, 'Data', index_col=0)
        self.key_datas   = list(self.data_frame.columns.values)
        self.meta_dict   = self.get_row_as_dict('meta')
        self.mode_dict   = self.get_row_as_dict('mode')
        self.expr_dict   = self.get_row_as_dict('expr')
        self.type_dict   = self.get_row_as_dict('type')
        self.source_dict = self.get_row_as_dict('source')
        return self.data_frame

    def number(self, key, round=0, suffix="") -> str:
        """ Return a INT key value as a rounded number formatted string, with a suffix (e.g. %)  """
        return round0str(self.data_frame[key][self.store_index], round)+suffix

    def str(self, key) -> str:
        """ Return a STR key value as a string """
        return str(self.data_frame[key][self.store_index])

    def taux(self, key, round=2, suffix="") -> str:
        """ Return a TAUX key value as a rounded number formatted string, with a suffix (e.g. %)  """
        val = self.data_frame[key][self.store_index]
        if (suffix == "%"): val = val * 100
        return round0str(val, round)+suffix

    def tauxp100(self, key, round=2) -> str:
        """ Return a TAUX key value as a percentage rounded number formatted string, with a suffix (e.g. %)  """
        return round0str(self.data_frame[key][self.store_index] * 100, round)+"%"

    def percent(self, key, round=2, suffix="%") -> str:
        """ Return a PERCENT key value as a rounded number formatted string, with a suffix (e.g. %)  """
        return round0str(self.data_frame[key][self.store_index], round) + suffix

    def getHTML(self) -> str:
        """ Return HTML of rendered Report """
        if (self.html) : return self.html
        self.render_report()
        return self.html

    def render_report(self, template=html_report_template, suffix=""):
        """ Render Report with specific HTML Mako Template """
        self.run_calculs()
        self.run_diagnostic()
        self.run_plots()
        # Building Mako Template Context
        html_file = output_dir + self.get_fullname() + suffix + ".html"
        print_green("> Render Report : " + html_file)
        context = {**self.get_row_as_dict(), **global_context}
        metric_list = []
        for metric in self.metric_list :
            metric['value'] = self[metric['key']]
            metric_list.append(metric)
        context["METRICS"] = metric_list
        diag_list = []
        for diagnostic in self.diagnostics :
            diag_list.append(diagnostic)
        context["DIAGNOSTICS"] = diag_list
        context["COMMUNE"] = str(context["NOM_COMMUNE"]).title()
        # Save Context File as reference TAGS for Make Template
        save_file(to_yaml(context), context_file)
        # Rendering Template
        mako.runtime.UNDEFINED = 'CONTEXT_MISSING_DATA'
        temp = Template(filename=template)
        self.html = temp.render(**context)
        # Saving to File
        f = open(html_file, 'w')
        f.write(self.html)
        f.close()
        return html_file

    # Add Meta Data
    def add_meta_data(self):
        """ Adds Meta data to DataStore """
        self.clean_meta_data(inplace=True)
        self.data_frame = self.data_frame.append(pd.Series(self.meta_dict,   name='meta'))
        self.data_frame = self.data_frame.sort_index(ascending=False)        # sorting by index
        self.data_frame = self.data_frame.append(pd.Series(self.mode_dict,   name='mode'))
        self.data_frame = self.data_frame.append(pd.Series(self.expr_dict,   name='expr'))
        self.data_frame = self.data_frame.append(pd.Series(self.type_dict,   name='type'))
        self.data_frame = self.data_frame.append(pd.Series(self.source_dict, name='source'))
        return self

    # Get Data Frame without Meta Data - Clean Meta Data
    def clean_meta_data(self, inplace=True) :
        """ Removes Meta data from DataStore """
        clean_data_frame = self.data_frame
        if 'meta'   in clean_data_frame.index : clean_data_frame.drop(labels="meta",   axis=0, inplace=inplace)
        if 'mode'   in clean_data_frame.index : clean_data_frame.drop(labels="mode",   axis=0, inplace=inplace)
        if 'type'   in clean_data_frame.index : clean_data_frame.drop(labels="type",   axis=0, inplace=inplace)
        if 'source' in clean_data_frame.index : clean_data_frame.drop(labels="source", axis=0, inplace=inplace)
        if 'expr'   in clean_data_frame.index : clean_data_frame.drop(labels="expr",   axis=0, inplace=inplace)
        return clean_data_frame

    # Get Data Frame without Individual Data - Keep Total & Meta Data
    def clean_data(self) :
        """ Removes Individual data from DataStore """
        clean_data_frame = self.data_frame
        for index, row in clean_data_frame.iterrows():
            if (index in ["meta", "total", "mode", "type", "source", "expr"]) : continue
            clean_data_frame = clean_data_frame.drop(index)
        return clean_data_frame

    def collect_data(self, code_postal=None, code_insee=None):
        """ Collect individual Commune data """

        commune = "Pas de Nom"
        if (code_postal):
            code_insee, commune = get_code_insee_commune(code_postal)

        if (code_insee):
            code_postal, commune = get_code_postal_commune(code_insee)

        if (str(code_insee) in DataStoreCache) :
            print_green("Cached Donnees : " + str(code_postal) + " : " + commune + " (Code INSEE : " + code_insee + ")")
            self.data_frame = merge_DataStoreCache(self, code_insee)
            return self

        save_index       = self.store_index
        self.store_index = code_insee

        print_green("> Collecte Donnees  : " + str(code_postal) + " : " + commune + " (Code INSEE : " + code_insee + ")")

        # Donnees Commune
        self.add_metric("CODE_INSEE",     "Code INSEE Commune",        source=source_codes,    mode=mode_count,  data=code_insee,          type="STR")
        self.add_metric("CODE_POSTAL",    "Code Postal Commune",       source=source_codes,    mode=mode_equal,  data=code_postal,         type="STR")
        self.add_metric("NOM_COMMUNE",    "Nom de Commune",            source=source_codes,    mode=mode_count,  data=commune,             type="STR")
        self.add_metric("LIBELLE",        "Libelle",                   source=source_codes,    mode=mode_count,  data=commune.title(),     type="STR")
        self.add_metric("BASE_NAME",      "Nom Unique",                source=source_codes,    mode=mode_custom, data=self.get_fullname(), type="STR")

        # Donnees EPCI
        self.add_metric("EPCI",      "Code EPCI - Métropole",          source=source_interco,  mode=mode_count, data=intercoDossier["Unnamed: 2"][code_insee],         type="STR")
        self.add_metric("LIBEPCI",   "Libellé de l'EPCI / Métropole",  source=source_interco,  mode=mode_count, data=intercoDossier["Unnamed: 3"][code_insee],         type="STR")
        self.add_metric("TYPE_EPCI", "Nature d'EPCI",                  source=source_interco,  mode=mode_count, data=intercoEPCI["Unnamed: 2"][self.get("EPCI")],      type="STR")
        self.add_metric("EPCI_COMMUNES", "Nombre communes EPCI",       source=source_interco,  mode=mode_count, data=intercoEPCI["Unnamed: 3"][self.get("EPCI")],      type="INT")
        self.add_metric("DEP",       "Departement",                    source=source_interco,  mode=mode_count, data=intercoDossier["Unnamed: 4"][code_insee],         type="STR")
        self.add_metric("DEP_NOM",   "Nom Departement",                source=source_interco,  mode=mode_count, data=departements["nom_departement"][self.get("DEP")], type="STR")
        self.add_metric("REG",       "Region",                         source=source_interco,  mode=mode_count, data=intercoDossier["Unnamed: 5"][code_insee],         type="STR")
        self.add_metric("REG_NOM",   "Nom Région",                     source=source_interco,  mode=mode_count, data=departements["nom_region"][self.get("DEP")],      type="STR")
        url_dossier = "https://www.insee.fr/fr/statistiques/2011101?geo=COM-" + self.str("CODE_INSEE")
        self.add_metric("DOSSIER_INSEE", "Dossier Complet INSEE",      source=source_insee,    mode=mode_custom, data=url_dossier,      type="STR")
        global_context["URL_SOURCE_DOSSIER"] = url_dossier

        # Donnees Display
        self.add_metric("DISPLAY_TYPE",   "Type de Territoire",             source=source_interco,  mode=mode_count, data="Pas de Type",     type="STR")
        self.add_metric("DISPLAY_COMM",   "Nom de Commune",                 source=source_interco,  mode=mode_count, data="Pas de Nom",      type="STR")
        self.add_metric("DISPLAY_EPCI",   "Nom de l'EPCI / Métropole",      source=source_interco,  mode=mode_count, data="Pas de Nom",      type="STR")
        self.add_metric("DISPLAY_ZONE",   "Nom de la Zone",                 source=source_interco,  mode=mode_count, data="Pas de Nom",      type="STR")
        self.add_metric("DISPLAY_DEPT",   "Nom du Departement",             source=source_interco,  mode=mode_count, data="Pas de Nom",      type="STR")
        self.add_metric("DISPLAY_REG",    "Nom de la Region",               source=source_interco,  mode=mode_count, data="Pas de Nom",      type="STR")
        self.add_metric("DISPLAY_NAME",   "Nom du Territoire",              source=source_interco,  mode=mode_count, data="Pas de Nom",      type="STR")

        # Donnees INSEE Commune
        load_communes()
        load_pop2019()

        # Donnees SRU
        load_sru()

        # Donnee Artificialisation
        load_artificialisation()
        load_artificialisationPaca()

        # Donnees Projections EPCI
        load_projections_paca()
        if (self.get("EPCI") == "ZZZZZZZZZ") :
            p_epci = "Sans Objet"
        else:
            p_epci = projectionsPaca.loc[projectionsPaca["Unnamed: 1"] == int(self.get("EPCI"))]
            p_epci = p_epci.head(1)

        # Donnees Projections SCOT
        load_projections_paca()
        p_scot = projectionsPaca.loc[projectionsPaca["Projection selon quatre scénarios à l'horizon 2030 et 2050  "] == "SCOT_50_000_ou_plus"]
        p_scot = p_scot.loc[p_scot["Unnamed: 1"].astype(str).str.contains(self.get("EPCI"))]
        p_scot = p_scot.head(1)

        # Donnees Projections Departement & Region 2013-2050
        load_projections()

        # Donnees Evolution Departement 2008-2021
        load_evolution()

        # Donnees Sitadel Logements
        load_sitadel()
        com_sitadel1316   = sitadel1316.loc[sitadel1316['COMM'] == str(code_insee)]
        com_sitadel1721   = sitadel1721.loc[sitadel1721['COMM'] == str(code_insee)]
        com_sitadel       = pd.concat([com_sitadel1316, com_sitadel1721])
        log_commences     = com_sitadel.loc[com_sitadel['Etat_DAU'] == "5"]
        log_termines      = com_sitadel.loc[com_sitadel['Etat_DAU'] == "6"]
        log_commences1316 = com_sitadel1316.loc[com_sitadel1316['Etat_DAU'] == "5"]
        log_termines1316  = com_sitadel1316.loc[com_sitadel1316['Etat_DAU'] == "6"]
        log_commences1721 = com_sitadel1721.loc[com_sitadel1721['Etat_DAU'] == "5"]
        log_termines1721  = com_sitadel1721.loc[com_sitadel1721['Etat_DAU'] == "6"]
        log_renouv        = com_sitadel.loc[(com_sitadel['NATURE_PROJET_DECLAREE'] == "2") & (com_sitadel['Etat_DAU'] != "4")]
        log_nouveau       = com_sitadel.loc[(com_sitadel['NATURE_PROJET_DECLAREE'] == "1") & (com_sitadel['Etat_DAU'] != "4")]
        log_principal     = com_sitadel.loc[(com_sitadel['RES_PRINCIP_OU_SECOND'] == 1) & (com_sitadel['Etat_DAU'] != "4")]
        log_secondaire    = com_sitadel.loc[(com_sitadel['RES_PRINCIP_OU_SECOND'] == 2) & (com_sitadel['Etat_DAU'] != "4")]
        log_particuliers  = com_sitadel.loc[com_sitadel['CAT_DEM'].isin([10, 11, 12])]
        log_organismes    = com_sitadel.loc[~com_sitadel['CAT_DEM'].isin([10, 11, 12])]
        res_sociales      = com_sitadel.loc[com_sitadel['RESIDENCE_SERVICE'].isin([1, 2, 4, 5, 6])]
        res_services      = com_sitadel.loc[~com_sitadel['RESIDENCE_SERVICE'].isin([1, 2, 4, 5, 6])]

        # Donnees Sitadel Locaux
        load_sitadel_locaux()
        com_sitadelLocaux1 = sitadel_locaux_1316.loc[sitadel_locaux_1316['COMM'] == str(code_insee)]
        com_sitadelLocaux2 = sitadel_locaux_1721.loc[sitadel_locaux_1721['COMM'] == str(code_insee)]
        com_sitadelLocaux  = pd.concat([com_sitadelLocaux1, com_sitadelLocaux2])
        loc_commences      = com_sitadelLocaux.loc[com_sitadelLocaux['Etat_DAU'] == "5"]
        loc_termines       = com_sitadelLocaux.loc[com_sitadelLocaux['Etat_DAU'] == "6"]
        loc_commences1316  = com_sitadelLocaux1.loc[com_sitadelLocaux1['Etat_DAU'] == "5"]
        loc_termines1316   = com_sitadelLocaux1.loc[com_sitadelLocaux1['Etat_DAU'] == "6"]
        loc_commences1721  = com_sitadelLocaux2.loc[com_sitadelLocaux2['Etat_DAU'] == "5"]
        loc_termines1721   = com_sitadelLocaux2.loc[com_sitadelLocaux2['Etat_DAU'] == "6"]
        loc_nouveau        = com_sitadelLocaux.loc[(com_sitadelLocaux['NATURE_PROJET_DECLAREE'] == "1") & (com_sitadelLocaux['Etat_DAU'] != "4")]
        loc_renouv         = com_sitadelLocaux.loc[(com_sitadelLocaux['NATURE_PROJET_DECLAREE'] == "2") & (com_sitadelLocaux['Etat_DAU'] != "4")]

        # Donnees Logements Paca 2010-2019
        load_logements_paca()

        # Donnees Flux
        load_flux_2014()
        load_flux_2015()
        load_flux_2016()
        load_flux_2017()
        load_flux_2018()
        load_fluxpro_2018()

        # Collected Data
        print_green("> Calculs Metriques : " + str(code_postal) + " : " + commune + " (Code INSEE : " + code_insee + ")")
        _line = 0
        for index, metric in collectDataMetrics.iterrows():
            _line = _line + 1
            _key         = index
            _description = metric["Description"]
            _source      = metric["Source"]
            _type        = metric["Type"]
            _data        = str(metric["Expr"])
            _expr        = str(metric["Expr"])
            _total       = metric["Total"]
            if (str(_key) == "nan") : continue          # Empty Line
            if (str(_key) == "")    : continue          # Empty Line
            if (str(_key).startswith("#")) : continue   # Ignore line / key starting with #
            # print_grey("Evaluating Metric " + str(_line) + ": " + str(_key) + " : " + str(_data))
            try:
                print_verbose("  - Evaluating Collect Line " + str(_line) + " :  [" + str(_key) + "] Expr : " + str(_expr))
                _data = re.sub("\${([A-Z0-9a-z-_]*)}", '\\1', _data)  # Replace ${VAR} by VAR
                value = eval(_data, self.get_row_as_dict(), {**globals(), **locals()})
                print_verbose("  - Evaluating Collect Line " + str(_line) + " :  [" + str(_key) + "] Value : " + str(value) + " - " + str(type(value)))
                self.add_metric(_key, _description, source=_source, mode=_total, data=value, type=_type, expr=_expr)
            except Exception as e :
                error = "collect_data Error evaluating metrique : Line " + str(_line) + " Key = " + str(_key) + " expr : " + str(_expr) + " - Error : " + str(e)
                print_error(error)
                if (_expr.startswith("error0")) :
                    value = 0
                self.add_metric(_key, _description, source=_source, mode=_total, data=value, type=_type, expr=_expr)

        update_DataStoreCache(self)
        self.store_index = save_index
        return self

    def total_data(self, meta=True):
        """ Consolidate total data for a set of Commune data """

        data_clean = self.clean_meta_data()
        if 'total'   in data_clean.index : data_clean.drop(labels="total",   axis=0, inplace=True)

        print_green("> Total Donnees   : " + str(self.store_code) + " : " + self.store_name)

        total_dict = {}
        _line = 0
        for key in self.key_datas:
            _line = _line + 1
            total_dict[key] = "ERROR Invalid Mode"
            mode = self.mode_dict[key].strip()
            try:
                print_verbose("  - Evaluating Total Line " + str(_line) + ": [" + str(key) + "] Mode : " + str(mode))
                if (key == "CODE_INSEE"): total_dict[key] = self.store_index
                elif (mode == "SUM"):     total_dict[key] = data_clean[key].sum()
                elif (mode == "CONCAT"):  total_dict[key] = re.sub(' +', ' ',data_clean[key].str.cat(sep=', '))
                elif (mode == "EQUAL"):   total_dict[key] = data_clean[key][0]
                elif (mode == "SAME"):    total_dict[key] = "SAME not evaluated"
                elif (mode == "COUNT"):   total_dict[key] = data_clean.shape[0]
                elif (mode == "IGNORE"):  total_dict[key] = "IGNORE"
                elif (mode == "N/A"):     total_dict[key] = "N/A"
                elif (mode == "AVG"):     total_dict[key] = data_clean[key].average()
                elif (mode == "MIN"):     total_dict[key] = data_clean[key].min()
                elif (mode == "MAX"):     total_dict[key] = data_clean[key].max()
                elif (mode == "CUSTOM"):  total_dict[key] = "CUSTOM not Implemented"
                else:                     total_dict[key] = "CUSTOM not evaluated"
                print_verbose("  - Evaluating Total Line " + str(_line) + ": [" + str(key) + "] Value : " + str(total_dict[key]))
            except Exception as e:
                print_error("Exception Evaluating Total Line : [" + str(mode) + " ] for key : [" + str(key) + "] : " + str(e))

        if (self.store_type == "COMMUNE"):
            total_dict["DISPLAY_TYPE"]  = "COMMUNE"
            total_dict["DISPLAY_NAME"]  = nom_commune(code_insee=self.store_code, clean=False)
            total_dict["DISPLAY_COMM"]  = nom_commune(code_insee=self.store_code, clean=False)
            total_dict["DISPLAY_EPCI"]  = nom_epci(epci_commune(self.store_code), clean=False)
            total_dict["DISPLAY_ZONE"]  = nom_epci(epci_commune(self.store_code), clean=False)
            total_dict["DISPLAY_DEPT"]  = nom_dept(dept_epci(epci_commune(self.store_code)), clean=False)
            total_dict["DISPLAY_REG"]   = nom_region(region_epci(epci_commune(self.store_code)), clean=False)

            total_dict["BASE_NAME"]     = self.get_fullname()
            total_dict["LIBELLE"]       = self.store_name
            total_dict["TYPE_EPCI"]     = "COMMUNE"
            total_dict["CODE_POSTAL"]   = get_code_postal(self.store_code)
            url_dossier = "https://www.insee.fr/fr/statistiques/2011101?geo=COM-" + str(self.store_code)
            total_dict["DOSSIER_INSEE"]          = url_dossier
            global_context["URL_SOURCE_DOSSIER"] = url_dossier
            total_dict["EPCI"]          = epci_commune(self.store_code)
            total_dict["LIBEPCI"]       = nom_epci(total_dict["EPCI"], clean=True)
            total_dict["TYPE_EPCI"]     = "COMMUNE"
            total_dict["EPCI_COMMUNES"] = 1
            total_dict["DEP"]           = dept_epci(total_dict["EPCI"])
            total_dict["DEP_NOM"]       = nom_dept(dept_epci(total_dict["EPCI"]), clean=True)
            total_dict["REG"]           = region_epci(total_dict["EPCI"])
            total_dict["REG_NOM"]       = nom_region(region_epci(total_dict["EPCI"]), clean=True)
            URL_VILLE_DATA = "https://ville-data.com/"+clean_name(nom_commune(code_insee=self.store_code) , sep="-").title().replace("-Le-" , "-le-").replace("-La-" , "-la-")+"-"+str(total_dict["CODE_POSTAL"])+".html"
            total_dict["URL_VILLE_DATA"]         = URL_VILLE_DATA
            URL_LINTERNAUTE = "https://www.linternaute.com/ville/"+clean_name(nom_commune(code_insee=self.store_code) , sep="-").lower()+"/ville-"+str(self.store_code)
            total_dict["URL_LINTERNAUTE"]        = URL_LINTERNAUTE
            total_dict["URL_GOOGLE"]             = "https://www.google.com/search?q="+total_dict["LIBELLE"]

        if (self.store_type == "EPCI"):
            total_dict["DISPLAY_TYPE"]  = "EPCI"
            total_dict["DISPLAY_NAME"]  = nom_epci(self.store_code, clean=False)
            total_dict["DISPLAY_COMM"]  = str_list(list_nom_communes(communes_epci(self.store_code)))
            total_dict["DISPLAY_EPCI"]  = nom_epci(self.store_code, clean=False)
            total_dict["DISPLAY_ZONE"]  = nom_epci(self.store_code, clean=False)
            total_dict["DISPLAY_DEPT"]  = nom_dept(dept_epci(self.store_code), clean=False)
            total_dict["DISPLAY_REG"]   = nom_region(region_epci(self.store_code), clean=False)

            total_dict["BASE_NAME"]     = self.get_fullname()
            total_dict["LIBELLE"]       = self.store_name
            total_dict["CODE_INSEE"]    = self.store_code
            total_dict["CODE_POSTAL"]   = "EPCI"
            # total_dict["NOM_COMMUNE"]   = "EPCI"
            total_dict["EPCI"]          = self.store_code
            total_dict["LIBEPCI"]       = self.store_name
            total_dict["TYPE_EPCI"]     = "EPCI"
            total_dict["EPCI_COMMUNES"] = len(communes_epci(self.store_code))
            total_dict["DEP"]           = dept_epci(self.store_code)
            total_dict["DEP_NOM"]       = nom_dept(dept_epci(self.store_code), clean=True)

            total_dict["REG"]           = region_epci(self.store_code)
            total_dict["REG_NOM"]       = nom_region(region_epci(self.store_code), clean=True)
            url_dossier = "https://www.insee.fr/fr/statistiques/2011101?geo=EPCI-" + str(self.store_code)
            total_dict["DOSSIER_INSEE"]          = url_dossier
            global_context["URL_SOURCE_DOSSIER"] = url_dossier
            total_dict["URL_VILLE_DATA"]         = "https://www.google.com/search?q="+total_dict["LIBELLE"]
            total_dict["URL_LINTERNAUTE"]        = "https://www.google.com/search?q="+total_dict["LIBELLE"]
            total_dict["URL_GOOGLE"]             = "https://www.google.com/search?q="+total_dict["LIBELLE"]

        if (self.store_type == "DEPT"):
            total_dict["DISPLAY_TYPE"]  = "DEPT"
            total_dict["DISPLAY_NAME"]  = nom_dept(self.store_code, clean=False)
            total_dict["DISPLAY_COMM"]  = str_list(list_nom_communes(communes_dept(self.store_code)))
            total_dict["DISPLAY_EPCI"]  = str_list(epci_dept(self.store_code))
            total_dict["DISPLAY_ZONE"]  = str_list(list_zones_dept(self.store_code))
            total_dict["DISPLAY_DEPT"]  = nom_dept(self.store_code, clean=False)
            total_dict["DISPLAY_REG"]   = nom_region(region_dept(self.store_code), clean=False)

            total_dict["BASE_NAME"]     = self.get_fullname()
            total_dict["LIBELLE"]       = self.store_name
            total_dict["CODE_INSEE"]    = self.store_code
            total_dict["CODE_POSTAL"]   = "DEPT"
            total_dict["NOM_COMMUNE"]   = nom_dept(self.store_code, clean=True)
            total_dict["EPCI"]          = "DEPT"
            total_dict["LIBEPCI"]       = "DEPT"
            total_dict["TYPE_EPCI"]     = "DEPT"
            total_dict["EPCI_COMMUNES"] = len(communes_dept(self.store_code))
            total_dict["DEP"]           = self.store_code
            total_dict["DEP_NOM"]       = nom_dept(self.store_code, clean=True)
            total_dict["REG"]           = region_dept(self.store_code)
            total_dict["REG_NOM"]       = nom_region(region_dept(self.store_code), clean=True)
            url_dossier = "https://www.insee.fr/fr/statistiques/2011101?geo=DEP-" + str(self.store_code)
            total_dict["DOSSIER_INSEE"]          = url_dossier
            global_context["URL_SOURCE_DOSSIER"] = url_dossier
            URL_VILLE_DATA = "https://ville-data.com/"+clean_name(str(total_dict["LIBELLE"]) , sep="-").lower()+".html"
            total_dict["URL_VILLE_DATA"]         = URL_VILLE_DATA
            URL_LINTERNAUTE = "https://www.linternaute.com/ville/"+clean_name(str(total_dict["LIBELLE"]) , sep="-").lower()+"/departement-"+str(self.store_code)
            total_dict["URL_LINTERNAUTE"]        = URL_LINTERNAUTE
            total_dict["URL_GOOGLE"]             = "https://www.google.com/search?q="+total_dict["LIBELLE"]

        if (self.store_type == "REGION"):
            total_dict["DISPLAY_TYPE"]  = "REGION"
            total_dict["DISPLAY_NAME"]  = nom_region(self.store_code, clean=False)
            total_dict["DISPLAY_COMM"]  = str_list(list_nom_communes(communes_region(self.store_code)))
            total_dict["DISPLAY_EPCI"]  = str_list(epci_region(self.store_code))
            total_dict["DISPLAY_ZONE"]  = str_list(list_zones(self.store_code))
            total_dict["DISPLAY_DEPT"]  = str_list(list_dept(self.store_code, clean=False))
            total_dict["DISPLAY_REG"]   = nom_region(self.store_code, clean=False)

            total_dict["BASE_NAME"]     = self.get_fullname()
            total_dict["LIBELLE"]       = self.store_name
            total_dict["CODE_INSEE"]    = self.store_code
            total_dict["CODE_POSTAL"]   = "REGION"
            total_dict["NOM_COMMUNE"]   = nom_region(region_dept(self.store_code), clean=True)
            total_dict["EPCI"]          = "REGION"
            total_dict["LIBEPCI"]       = "REGION"
            total_dict["TYPE_EPCI"]     = "REGION"
            total_dict["EPCI_COMMUNES"] = len(communes_region(self.store_code))
            total_dict["DEP"]           = "REGION"
            total_dict["DEP_NOM"]       = "REGION"
            total_dict["REG"]           = self.store_code
            total_dict["REG_NOM"]       = nom_region(region_dept(self.store_code), clean=True)
            url_dossier = "https://www.insee.fr/fr/statistiques/2011101?geo=REG-" + str(self.store_code)
            total_dict["DOSSIER_INSEE"]          = url_dossier
            global_context["URL_SOURCE_DOSSIER"] = url_dossier
            total_dict["URL_VILLE_DATA"]         = "https://ville-data.com/paca.html"
            total_dict["URL_LINTERNAUTE"]        = "https://fr.wikipedia.org/wiki/Provence-Alpes-C%C3%B4te_d%27Azur"
            total_dict["URL_GOOGLE"]             = "https://www.google.com/search?q="+total_dict["LIBELLE"]

        if (self.store_type == "ZONE"):
            total_dict["DISPLAY_TYPE"]  = "ZONE"
            total_dict["DISPLAY_NAME"]  = self.store_code
            total_dict["DISPLAY_COMM"]  = str_list(list_nom_communes(communes_zone(self.store_code)))
            total_dict["DISPLAY_EPCI"]  = self.store_code
            total_dict["DISPLAY_ZONE"]  = self.store_code
            total_dict["DISPLAY_DEPT"]  = nom_dept(dept_zone(self.store_code), clean=False)
            total_dict["DISPLAY_REG"]   = nom_region(region_dept(dept_zone(self.store_code)), clean=False)

            total_dict["BASE_NAME"]     = self.get_fullname()
            total_dict["LIBELLE"]       = self.store_name
            total_dict["CODE_INSEE"]    = self.store_code
            total_dict["CODE_POSTAL"]   = "ZONE"
            # total_dict["NOM_COMMUNE"]   = "ZONE"
            total_dict["EPCI"]          = self.store_name
            total_dict["LIBEPCI"]       = self.store_name
            total_dict["TYPE_EPCI"]     = "ZONE"
            total_dict["EPCI_COMMUNES"] = communes_zone(self.store_name)
            total_dict["DEP"]           = SCOT_DATA["GROUPES_COMMUNES"][self.store_name]["DEPT"]
            total_dict["DEP_NOM"]       = nom_dept(SCOT_DATA["GROUPES_COMMUNES"][self.store_name]["DEPT"])
            total_dict["REG"]           = SCOT_DATA["GROUPES_COMMUNES"][self.store_name]["REG"]
            total_dict["REG_NOM"]       = nom_region(SCOT_DATA["GROUPES_COMMUNES"][self.store_name]["REG"])
            url_dossier = "https://www.insee.fr/"
            total_dict["DOSSIER_INSEE"]          = url_dossier
            global_context["URL_SOURCE_DOSSIER"] = url_dossier
            total_dict["URL_VILLE_DATA"]         = "https://www.google.com/search?q="+total_dict["LIBELLE"]
            total_dict["URL_LINTERNAUTE"]        = "https://www.google.com/search?q="+total_dict["LIBELLE"]
            total_dict["URL_GOOGLE"]             = "https://www.google.com/search?q="+total_dict["LIBELLE"]

        for key in self.key_datas:
            mode = self.mode_dict[key]
            if (key == "CODE_INSEE"): continue
            elif (mode == "SUM"):     continue
            elif (mode == "MIN"):     continue
            elif (mode == "MAX"):     continue
            elif (mode == "CONCAT"):  continue
            elif (mode == "EQUAL"):   continue
            elif (mode == "COUNT"):   continue
            elif (mode == "IGNORE"):  continue
            elif (mode == "N/A"):     continue
            elif (mode == "AVG"):     continue
            elif (mode == "CUSTOM"):  continue
            else:
                if (mode == "SAME"):
                    mode = self.expr_dict[key]
                try:
                    print_verbose("  - Evaluating Total Line : [" + str(key) + "] Mode : " + str(mode))
                    mode  = re.sub("\${([A-Z0-9a-z-_]*)}", '\\1', mode)    # Replace ${VAR} by VAR
                    value = eval(mode, total_dict, globals())
                    print_verbose("  - Evaluating Total Line : [" + str(key) + "] Value : " + str(value) + " - " + str(type(value)))
                    total_dict[key] = value
                except Exception as e:
                    error = "total_data Error evaluating metrique total mode  : " + key + " + expr : " + mode + " - Error : " + str(e)
                    print_error(error)
                    if (mode.startswith("error0")):
                        error = 0
                    total_dict[key] = error

        # Store in Data Frame
        self.data_frame = self.data_frame.append(pd.Series(total_dict, name="total"))
        if (meta):
            # Add Meta Data
            self.data_frame = self.data_frame.append(pd.Series(self.meta_dict,   name='meta'))
            self.data_frame = self.data_frame.sort_index(ascending=False)        # sorting by index
            self.data_frame = self.data_frame.append(pd.Series(self.mode_dict,   name='mode'))
            self.data_frame = self.data_frame.append(pd.Series(self.type_dict,   name='type'))
            self.data_frame = self.data_frame.append(pd.Series(self.source_dict, name='source'))
            self.data_frame = self.data_frame.append(pd.Series(self.expr_dict,   name='expr'))

        update_DataStoreCache(self)
        return self

    def run_calculs(self, meta=True, force=False):
        """ Run Calculs from Configuration file, Calculs tab """
        if ((self.calc_done) and (force==False)) : return self

        print_green("> Calculs Donnees : " + str(self.store_code) + " : " + self.store_name)

        save_index       = self.store_index
        self.store_index = "total"

        # Calculated Data
        _line = 1
        for index, metric in collectCalculations.iterrows():
            _line = _line + 1
            _key         = index
            _description = metric["Description"]
            _source      = metric["Source"]
            _type        = metric["Type"]
            _flag        = metric["Flag"]
            _python      = str(metric["Python"])
            _javascript  = str(metric["JavaScript"])
            if (str(_key) == "nan") : continue          # Empty Line
            if (str(_key) == "")    : continue          # Empty Line
            if (str(_key).startswith("#")) : continue   # Ignore line / key starting with #
            try:
                print_verbose("  - Evaluating Calcul Line " + str(_line) + ": [" + str(_key) + "] Eval : " + str(_python))
                _python = re.sub("\${([A-Z0-9a-z-_]*)}", '\\1', _python)  # Replace ${VAR} by VAR
                value = eval(_python, self.get_row_as_dict(), {**globals(), **locals()})
                print_verbose("  - Evaluating Calcul Line " + str(_line) + ": [" + str(_key) + "] Value : " + str(value) + " - " + str(type(value)))
                if   ((pd.isna(value)) and (_type == "STR")):     value = ""
                elif ((pd.isna(value)) and (_type == "INT")):     value = 0
                elif ((pd.isna(value)) and (_type == "FLOAT")):   value = 0
                elif ((pd.isna(value)) and (_type == "TAUX")):    value = 0
                elif ((pd.isna(value)) and (_type == "PERCENT")): value = 0
                if   (_type == "STR"):     value = str(value)
                elif (_type == "INT"):     value = int(round(float(str(value)), 0))
                elif (_type == "FLOAT"):   value = float(value)
                elif (_type == "TAUX"):    value = float(value)
                elif (_type == "PERCENT"): value = float(value)
                # self.add_metric(_key, _description, source=_source, mode="$JS:"+_javascript, data=value, type=_type, expr=_python)
                self.add_metric(_key, _description, source=_source, mode=_python, data=value, type=_type, expr=_python)
            except Exception as e :
                print_error("Error evaluating Calcul : Line " + str(_line) + " Key = " + str(_key) + " + expr : " + str(_python) + " - Error : " + str(e))
                self.add_metric(_key, _description, source=_source, mode=_python, data=str(e), type=_type, expr=_python)
        self.store_index = save_index
        self.add_meta_data()
        self.calc_done = True
        return self

    def run_diagnostic(self, force=False):
        """ Run Diagnostics from Configuration file, Diagnostics tab """
        if ((self.diag_done) and (force==False)) : return self
        load_min_data()
        self.run_calculs()
        print_green("> Etablissement des diagnostics : " + str(self.store_code) + " : " + self.store_name)
        _line = 0
        for index, metric in collectDiagnostics.iterrows():
            _line = _line + 1
            _key         = index  # Key
            _description = metric["Description"]
            _test        = str(metric["Test"])
            _messageV    = metric["MessageSiVrai"]
            _messageF    = metric["MessageSiFaux"]
            _categorie   = metric["Categorie"]
            _type        = metric["Type"]
            if (str(_key) == "nan")        : continue   # Ignore empty lines
            if (str(_key).startswith("#")) : continue   # Ignore key starting with #
            _data = re.sub("\${([A-Z0-9a-z-_]*)}", '\\1', _test)    # Replace ${VAR} by VAR
            try:
                print_verbose("  - Evaluating Diagnostic Line " + str(_line) + " : [" + str(_key) + "] Eval : " + str(_test))
                value = bool(eval(_test, self.get_row_as_dict(), {**globals(), **locals()}))
                print_verbose("  - Evaluating Diagnostic Line " + str(_line) + " : [" + str(_key) + "] Value : " + str(value) + " - " + str(type(value)))
                if (str(_messageV) == "nan") : _messageV = ""          # Empty Line
                if (str(_messageV).startswith("\"")):
                    _messageV = re.sub("\${([A-Z0-9a-z-_]*)}", '\\1', _messageV)  # Replace ${VAR} by VAR
                    _messageV = str(eval(_messageV, self.get_row_as_dict(), {**globals(), **locals()}))
                if (str(_messageF) == "nan") : _messageF = ""          # Empty Line
                if (str(_messageF).startswith("\"")):
                    _messageF = re.sub("\${([A-Z0-9a-z-_]*)}", '\\1', _messageF)    # Replace ${VAR} by VAR
                    _messageF = str(eval(_messageF, self.get_row_as_dict(), {**globals(), **locals()}))
                self.add_diagnostic(_key, _description, test=_test, messageV=_messageV, data=value, messageF=_messageF, cat=_categorie, type=_type)
            except Exception as e :
                print_error("Error evaluating Diagnostic : " + _key + " + eval : " + _test + " - Error : " + str(e))
                self.add_diagnostic(_key, _description, test=_test, messageV=str(e), data=value, messageF=str(e), cat=_categorie, type=_type)
        self.diag_done = True
        return self

    def run_plots(self, force=False):
        """ Draw Graphics """
        if ((self.plot_done) and (force==False)) : return self
        print_green("> Generation des Graphiques : " + str(self.store_code) + " : " + self.store_name)
        self.run_calculs()
        plots(self)
        self.plot_done = True
        return self

    def report(self, force=True, data_only: bool = False, ftp_push: bool = False):
        """ Generate report for DataStore. Force = True will re-calculate DataStore data from source. """
        load_min_data()
        print_blue("Preparation Rapport "+self.store_type + " " + self.store_name + " (Code INSEE : "+self.store_code+")")
        loaded = None
        if (force is False) :
            loaded = self.load_data()
        if (force is True) or (loaded is None):
            if (self.store_type == entite_commune):
                self.collect_data(code_insee=self.store_code)
                self.total_data(meta=True)
                self.store_index = self.store_code
            if (self.store_type == entite_epci):
                for commune in communes_epci(self.store_code):
                    self.collect_data(code_insee=commune)
                self.total_data(meta=True)
                self.store_index = 'total'
            if (self.store_type == entite_dept):
                for commune in communes_dept(self.store_code):
                    self.collect_data(code_insee=commune)
                self.total_data(meta=True)
                self.store_index = 'total'
            if (self.store_type == entite_zone) or (self.store_type == entite_scot):
                for commune in communes_zone(self.store_code):
                    self.collect_data(code_insee=commune)
                self.total_data(meta=True)
                self.store_index = 'total'

        global FAST
        html_report_file = None
        if (not FAST and FAST): # Disabling HTML Reports
            self.run_calculs()
            self.run_diagnostic()
            self.run_plots()
            gen_tracker(self)
            html_report_file = gen_report(ds=self)

        self.save_data()

        if (ftp_push):
            ftp_push_ds(self)

        if (html_report_file):
            display_in_browser(html_report_file)

        return self


def scot_ouest(code_insee, start_date="2021-05-20", file="scot_ouest"):
    global SCOT_OUEST
    load_sitadel()
    load_scot_data()
    load_collectData()
    # Annee / Log Aut / Log Commences / Nbre Log / Surface Terrain
    com_2021 = sitadel1721[(sitadel1721['COMM'] == str(code_insee)) &
                           (sitadel1721["DATE_REELLE_AUTORISATION"] > start_date)]

    if (com_2021.size == 0) : return None
    # (sitadel1721["NATURE_PROJET"] == "1") &
    # (sitadel1721["Etat_DAU"] != "4") &
    # (sitadel1721["Type_DAU"] == "PC") &

    com_2021.fillna(0)

    parcelles   = ""
    for index, row in com_2021.iterrows():
        com_2021.loc[index, "Parcelles"] = str(row['sec_cadastre1'])+str(row['num_cadastre1'])+" "+str(row['sec_cadastre2'])+str(row['num_cadastre2'])+" "+str(row['sec_cadastre3'])+str(row['num_cadastre3'])
        parcelles = parcelles + " " + com_2021.loc[index, "Parcelles"]
        if   (row['Etat_DAU'] == "2") : com_2021.loc[index, "Etat"]   = "Autorisé"
        elif (row['Etat_DAU'] == "4") : com_2021.loc[index, "Etat"]   = "Annulé"
        elif (row['Etat_DAU'] == "5") : com_2021.loc[index, "Etat"]   = "Commencé"
        elif (row['Etat_DAU'] == "6") : com_2021.loc[index, "Etat"]   = "Terminé"
        else : com_2021.loc[index, "Etat"] = "NR"
        com_2021.loc[index, "Nature"]      = "UNA" if (row['NATURE_PROJET_DECLAREE'] == "1") else "RU"
        com_2021.loc[index, "Extension"]   = int(row['I_EXTENSION']) + int(row['I_SURELEVATION']) + int(row['I_NIVSUPP'])
        com_2021.loc[index, "Renouv"]      = 1 if ((float(row['SURF_HAB_DEMOLIE']) + float(row['SURF_LOC_DEMOLIE']) + float(row['SURF_HAB_AVANT']) + float(row['NB_LGT_DEMOLIS']) + float(row['SURF_LOC_AVANT']))>0) else 0

    for index, row in com_2021.iterrows():
        com_2021.loc[index, "Artif"]       = row['SUPERFICIE_TERRAIN'] if ((row['Extension'] == 0) and (row['Renouv'] == 0) and (row['Etat_DAU'] != 4)) else 0
        com_2021.loc[index, "Date"]        = row['DATE_REELLE_AUTORISATION']

    com_2021 = com_2021[["Date", "Nature", "Extension", "Renouv", "Etat", "Artif", "Parcelles",
                         "DEP", "COMM", "Type_DAU", "Etat_DAU", "NATURE_PROJET", "NATURE_PROJET_DECLAREE",
                         "NB_LGT_TOT_CREES",   "NB_LGT_PRET_LOC_SOCIAL", "NB_LGT_ACC_SOC_HORS_PTZ", "NB_LGT_PTZ",
                         "SUPERFICIE_TERRAIN", "DATE_REELLE_AUTORISATION",
                         "SURF_HAB_DEMOLIE",   "SURF_LOC_DEMOLIE", "SURF_HAB_AVANT", "SURF_LOC_AVANT", "I_EXTENSION", "I_SURELEVATION",
                         "SURF_HAB_CREEE",     "SURF_LOC_CREEE"]]
    com_2021.sort_values("DATE_REELLE_AUTORISATION")

    superficie = sum(com_2021["SUPERFICIE_TERRAIN"])
    logements  = sum(com_2021["NB_LGT_TOT_CREES"])
    logsoc     = sum(com_2021["NB_LGT_PRET_LOC_SOCIAL"])
    artif      = sum(com_2021["Artif"])
    artif_18   = sum(com_2021[(com_2021["DATE_REELLE_AUTORISATION"] > "2018-01-01")]["Artif"])
    artif_19   = sum(com_2021[(com_2021["DATE_REELLE_AUTORISATION"] > "2019-01-01")]["Artif"])
    artif_20   = sum(com_2021[(com_2021["DATE_REELLE_AUTORISATION"] > "2020-01-01")]["Artif"])
    artif_21   = sum(com_2021[(com_2021["DATE_REELLE_AUTORISATION"] > "2021-01-01")]["Artif"])
    artif_22   = sum(com_2021[(com_2021["DATE_REELLE_AUTORISATION"] > "2022-01-01")]["Artif"])
    artif_SO   = sum(com_2021[(com_2021["DATE_REELLE_AUTORISATION"] > "2021-05-20")]["Artif"])

    count       = len(com_2021.index)
    count0      = len(com_2021.loc[com_2021["SUPERFICIE_TERRAIN"] == 0].index)
    superficie0 = round0(superficie + ((superficie/(count-count0))*count0), 0) if (count != count0) else superficie

    NOM_COMMUNE = nom_commune(code_insee=code_insee).upper()
    BUDGET_2030 = SCOT_OUEST['Budget 2030'][NOM_COMMUNE] * 10000 if (NOM_COMMUNE in SCOT_OUEST.index) else 0
    BUDGET_2040 = SCOT_OUEST['Budget 2040'][NOM_COMMUNE] * 10000 if (NOM_COMMUNE in SCOT_OUEST.index) else 0
    time_percent = ((10 - (2030 - datetime.date.today().year)-1) / 10)
    pace_percent = round0(artif_SO/BUDGET_2030, 0) if (BUDGET_2030 > 0) else 0
    trajectoire  = "OK"
    if (pace_percent > time_percent) :
        trajectoire = "Consommation Trop Rapide"

    com_2021.loc["Total / Count", "SUPERFICIE_TERRAIN"]     = superficie0
    com_2021.loc["Total / Count", "Type_DAU"]               = count
    com_2021.loc["Total / Count", "NB_LGT_TOT_CREES"]       = logements
    com_2021.loc["Total / Count", "NB_LGT_PRET_LOC_SOCIAL"] = logsoc
    com_2021.loc["Total / Count", "Artif"]                  = artif
    com_2021.loc["Total / Budget 2030", "Artif"]            = BUDGET_2030
    com_2021.loc["Total / 2018-01-01", "Artif"]             = artif_18
    com_2021.loc["Total / 2019-01-01", "Artif"]             = artif_19
    com_2021.loc["Total / 2020-01-01", "Artif"]             = artif_20
    com_2021.loc["Total / 2021-01-01", "Artif"]             = artif_21
    com_2021.loc["Total / 2022-01-01", "Artif"]             = artif_22
    com_2021.loc["Total / 2021-05-20", "Artif"]             = artif_SO
    com_2021.loc["Total / 2020-01-01", "Etat"]              = str(round0(((artif_20/BUDGET_2030)*100), 0))+"%" if (BUDGET_2030 > 0) else 0
    com_2021.loc["Total / 2021-01-01", "Etat"]              = str(round0(((artif_21/BUDGET_2030)*100), 0))+"%" if (BUDGET_2030 > 0) else 0
    com_2021.loc["Total / 2021-05-20", "Etat"]              = str(round0(((artif_SO/BUDGET_2030)*100), 0))+"%" if (BUDGET_2030 > 0) else 0
    com_2021.loc["Total / Time",       "Etat"]              = str(round0(time_percent*100, 0))+"%"
    com_2021.loc["Total / Time",       "Artif"]             = trajectoire
    com_2021.loc["Total / Budget 2040", "Artif"]            = BUDGET_2040

    file_name = output_dir + os.sep + "scot_ouest_"+clean_name(NOM_COMMUNE)+".xlsx"
    writer = pd.ExcelWriter(file_name, engine='xlsxwriter')
    com_2021.to_excel(writer, sheet_name=NOM_COMMUNE)
    worksheet1 = writer.sheets[NOM_COMMUNE]
    worksheet1.set_tab_color('green')
    red_format = writer.book.add_format({'bg_color': 'red'})
    worksheet1.conditional_format('H1:H1000', {'type': 'cell', 'criteria': 'equal to', 'value': 0, 'format': red_format})
    # writer.save()
    writer.close()

    file_123  = file + ".xlsx"
    sheet_123 = clean_name(NOM_COMMUNE)
    sheet_sum = "Summary"

    if (os.path.exists(file_123)):
        work_book = load_workbook(file_123)
    else:
        work_book = Workbook()

    # Data
    if (sheet_123 in work_book.sheetnames):
        work_book.remove(work_book[sheet_123])
    work_sheet = work_book.create_sheet(0)
    work_sheet.title = sheet_123
    for row in dataframe_to_rows(com_2021, index=True, header=True):
        work_sheet.append(row)
    data_number_of_rows = len(com_2021.index)


    red_color      = 'ffc7ce'
    red_color_font = '9c0103'
    green_color    = '00ff00'

    red_font   = styles.Font(bold=True, color=red_color_font)
    red_fill   = styles.PatternFill(start_color=red_color,   end_color=red_color,   fill_type='solid')
    green_fill = styles.PatternFill(start_color=green_color, end_color=green_color, fill_type='solid')
    dxf_green_fill = styles.differential.DifferentialStyle(fill=green_fill)
    dxf_red_fill   = styles.differential.DifferentialStyle(font=red_font, border=None, fill=red_fill)

    work_sheet.conditional_formatting.add('C3:C' + str(data_number_of_rows-4), formatting.rule.Rule(type='containsText', text='RU', dxf=dxf_green_fill))
    work_sheet.conditional_formatting.add('D3:D' + str(data_number_of_rows-4), formatting.rule.CellIsRule(operator='greaterThan',  formula=['0'], fill=green_fill))
    work_sheet.conditional_formatting.add('E3:E' + str(data_number_of_rows-4), formatting.rule.CellIsRule(operator='greaterThan',  formula=['0'], fill=green_fill))
    work_sheet.conditional_formatting.add('F3:F' + str(data_number_of_rows-4), formatting.rule.Rule(type='containsText', text='Annulé', dxf=dxf_green_fill))
    work_sheet.conditional_formatting.add('G3:G' + str(data_number_of_rows-4), formatting.rule.CellIsRule(operator='greaterThan',  formula=['0'], fill=red_fill, font=red_font))

    red_colour = 'ffc7ce'
    red_colour_font = '9c0103'

    red_font = styles.Font(size=14, bold=True, color=red_colour_font)
    red_fill = styles.PatternFill(start_color=red_colour, end_color=red_colour, fill_type='solid')

    rule = formatting.rule.Rule(type='containsText', text="RU", stopIfTrue=False)
    rule.dxf = styles.differential.DifferentialStyle(font=red_font, border=None, fill=red_fill)
    work_sheet.conditional_formatting.add('C3:C' + str(data_number_of_rows-4), rule)


    # Summary
    try:
        df_sum = pd.read_excel(file_123, sheet_name=sheet_sum, index_col=0)
    except:
        df_sum = pd.DataFrame()
        df_sum.index.name = 'Commune'
        df_sum = df_sum.rename_axis('Commune')
    sum_number_of_rows = len(com_2021.index)

    # df_sum.at[sheet_123, "Commune"]            = sheet_123
    df_sum.at[sheet_123, "# Permis"]           = count
    df_sum.at[sheet_123, "Logements"]          = logements
    df_sum.at[sheet_123, "Sociaux"]            = logsoc
    df_sum.at[sheet_123, "Superficie"]         = superficie0
    df_sum.at[sheet_123, "Consommation SCoT"]  = artif
    df_sum.at[sheet_123, "Budget 2030"]        = BUDGET_2030
    df_sum.at[sheet_123, "Conso / 2020-01-01"]   = artif_20
    df_sum.at[sheet_123, "% / 2020-01-01"]       = str(round0(((artif_20/BUDGET_2030)*100), 0))+"%" if (BUDGET_2030 > 0) else 0
    df_sum.at[sheet_123, "Conso / 2021-01-01"]   = artif_21
    df_sum.at[sheet_123, "% / 2021-01-01"]       = str(round0(((artif_21/BUDGET_2030)*100), 0))+"%" if (BUDGET_2030 > 0) else 0
    df_sum.at[sheet_123, "Conso / 2021-05-20"]   = artif_SO
    df_sum.at[sheet_123, "% / 2021-05-20"]       = str(round0(((artif_SO/BUDGET_2030)*100), 0))+"%" if (BUDGET_2030 > 0) else 0
    df_sum.at[sheet_123, "Total / Time"]         = str(round0(((time_percent)*100), 0))+"%"
    df_sum.at[sheet_123, "Budget 2040"]          = BUDGET_2040
    df_sum.at[sheet_123, "Trajectoire"]          = trajectoire

    if (sheet_sum in work_book.sheetnames):
        work_book.remove(work_book[sheet_sum])
    work_sheet = work_book.create_sheet(0)
    work_sheet.title = sheet_sum

    for row in dataframe_to_rows(df_sum, header=True):
        work_sheet.append(row)

    if ("Sheet" in work_book.sheetnames):
        work_book.remove(work_book["Sheet"])

    work_book.save(filename=file_123)

    return file_123


def excel_flux():
    # Donnees Flux
    load_flux_2018()
    load_flux_2017()
    load_flux_2016()
    load_flux_2015()
    load_flux_2014()
    excel_flux_details(flux2014, "NBFLUX_C14_POP01P" , "93", 'Flux2014-93.xlsx')
    excel_flux_details(flux2015, "NBFLUX_C15_POP01P" , "93", 'Flux2015-93.xlsx')
    excel_flux_details(flux2016, "NBFLUX_C16_POP01P" , "93", 'Flux2016-93.xlsx')
    excel_flux_details(flux2017, "NBFLUX_C17_POP01P" , "93", 'Flux2017-93.xlsx')
    excel_flux_details(flux2018, "NBFLUX_C18_POP01P" , "93", 'Flux2018-93.xlsx')


def excel_flux_details(data_file, field, region, xls_file):
        # Donnees Flux
        print_green("Flux : "+xls_file)
        print_green("Flux - Communes")
        data1 = data_file.loc[(flux2018['CODGEO'].isin(communes_region(region)))  | (data_file['DCRAN'].isin(communes_region(region)))]
        print_green("Flux - EPCI")

        data2 = pd.DataFrame(columns=epci_region(region) , index=epci_region(region))
        data2 = data2.append(pd.DataFrame(index=['99', 'Nom']))
        data2["99"] = ""
        data2["Nom"] = ""
        for x_epci in epci_region(region) :
            for y_epci in epci_region(region):
                xy_data = round(data_file.loc[flux2018['CODGEO'].isin(communes_epci(x_epci))  & data_file['DCRAN'].isin(communes_epci(y_epci)) & (data_file['DCRAN'] != data_file['CODGEO'])][field].sum(),0)
                data2[x_epci][y_epci] = xy_data
                data2['Nom'][y_epci] = nom_epci(y_epci)
                xy_data = round(data_file.loc[(~data_file['CODGEO'].isin(communes_region(region))) & (data_file['DCRAN'].isin(communes_epci(y_epci))) & (data_file['DCRAN'] != data_file['CODGEO'])][field].sum(), 0)
                data2['99'][y_epci] = xy_data
                data2['Nom'][y_epci] = nom_epci(y_epci)
            xy_data = round(data_file.loc[data_file['CODGEO'].isin(communes_epci(x_epci)) & (~data_file['DCRAN'].isin(communes_region(region))) & (data_file['DCRAN'] != data_file['CODGEO'])][field].sum(), 0)
            data2[x_epci]['99'] = xy_data
            data2[x_epci]['Nom'] = nom_epci(x_epci)

        print_green("Flux - DEPT")
        data3 = pd.DataFrame(columns=list_dept(region) , index=list_dept(region))
        data3 = data3.append(pd.DataFrame(index=['99', 'Nom']))
        data3["99"] = ""
        data3["Nom"] = ""
        for x_epci in list_dept(region) :
            for y_epci in list_dept(region):
                xy_data = round(data_file.loc[(data_file['CODGEO'].isin(communes_dept(x_epci)))  & (data_file['DCRAN'].isin(communes_dept(y_epci))) & (data_file['DCRAN'] != data_file['CODGEO'])][field].sum(),0)
                data3[x_epci][y_epci] = xy_data
                data3['Nom'][y_epci] = nom_dept(y_epci)
                xy_data = round(data_file.loc[(~(data_file['CODGEO'].isin(communes_region(region)))) & (data_file['DCRAN'].isin(communes_dept(y_epci))) & (data_file['DCRAN'] != data_file['CODGEO'])][field].sum(), 0)
                data3['99'][y_epci]  = xy_data
                data3['Nom'][y_epci]  = nom_dept(y_epci)
            xy_data = round(data_file.loc[(data_file['CODGEO'].isin(communes_dept(x_epci))) & (~data_file['DCRAN'].isin(communes_region(region))) & (data_file['DCRAN'] != data_file['CODGEO'])][field].sum(), 0)
            data3[x_epci]['99']  = xy_data
            data3[x_epci]['Nom'] = nom_dept(x_epci)

        # Create excel writer
        excel_file = output_dir + xls_file
        print_green("Flux : "+excel_file)
        writer = pd.ExcelWriter(excel_file)
        # Write dataframe to excel sheet named 'marks'
        data3.to_excel(writer, 'dept')
        data2.to_excel(writer, 'epci')
        data1.to_excel(writer, 'communes')
        # Save the excel file
        writer.save()


def update_DataStoreCache(ds : DataStore, code_insee=None):
    """ Cache a DataStore for re-use without re-calculations """
    if code_insee and ((isinstance(code_insee, int)) or (isinstance(code_insee, str))):
        DataStoreCache[str(code_insee)] = ds
        print_green("> Added in Cache  : DataStore with code INSEE " + str(ds.store_code) + " : "  + ds.store_name)
    elif (isinstance(ds["CODE_INSEE"], int)) or (isinstance(ds["CODE_INSEE"], str)):
        DataStoreCache[str(ds["CODE_INSEE"])] = ds
        print_green("> Added in Cache  : DataStore with code INSEE " + str(ds["CODE_INSEE"]) + " : "  + ds.store_name)
    else:
        print_red("- Not Added in Cache : DataStore without code INSEE : " + ds.store_name)


def merge_DataStoreCache(ds : DataStore, code_insee=None) -> pd.DataFrame :
    """ Update a DataStore from cache for Data with code_insee """
    l_data_frame = ds.data_frame.append(DataStoreCache[code_insee].data_frame.loc[code_insee, :])
    # Add Meta Data
    if 'meta'   not in l_data_frame.index: l_data_frame = l_data_frame.append(DataStoreCache[code_insee].data_frame.loc["meta", :])
    if 'mode'   not in l_data_frame.index: l_data_frame = l_data_frame.append(DataStoreCache[code_insee].data_frame.loc["mode", :])
    if 'type'   not in l_data_frame.index: l_data_frame = l_data_frame.append(DataStoreCache[code_insee].data_frame.loc["type", :])
    if 'source' not in l_data_frame.index: l_data_frame = l_data_frame.append(DataStoreCache[code_insee].data_frame.loc["source", :])
    if 'expr'   not in l_data_frame.index: l_data_frame = l_data_frame.append(DataStoreCache[code_insee].data_frame.loc["expr", :])
    ds.source_dict = DataStoreCache[code_insee].source_dict
    ds.key_datas   = DataStoreCache[code_insee].key_datas
    ds.meta_dict   = DataStoreCache[code_insee].meta_dict
    ds.type_dict   = DataStoreCache[code_insee].type_dict
    ds.source_dict = DataStoreCache[code_insee].source_dict
    ds.mode_dict   = DataStoreCache[code_insee].mode_dict
    ds.expr_dict   = DataStoreCache[code_insee].expr_dict
    return l_data_frame


###
### HTML Reports
###

def render_index(template=html_index_template, region="93"):
    """" Index HTML file of Regions,Dept, EPCI, Communes """
    # TODO
    context = {**report_region_dict(region=region, filename=france_file), **global_context}
    # Rendering Template
    mako.runtime.UNDEFINED = 'MISSING_CONTEXT'
    mako.runtime.UNDEFINED = 'MISSING_CONTEXT'
    temp = Template(filename=template)
    index_html = temp.render(**context)
    # Saving to File
    p_html_index_filename = output_dir + "index.html"
    f = open(p_html_index_filename, 'w')
    f.write(index_html)
    f.close()
    return p_html_index_filename


def report_diagnostic(ds: DataStore):
    """ Diagnostic Report in HTML """
    data_html = "<h3>" + "Diagnostic des Donnees" + "</h3>"
    for diagnostic in ds.diagnostics:
        if (diagnostic["value"]==True):
            data_html = data_html + "<p style=color:green;>" + " OK  : "      + str(diagnostic["messageSiVrai"]) + "</p>"
        else:
            data_html = data_html + "<p style=color:red;>"   + " Problem  : " + str(diagnostic["messageSiFaux"]) + "</p>"
    return data_html


def report_source(ds: DataStore):
    """ Sources in HTML """
    # Summary Table
    data_html = "<h3>" + "Sources de Donnees" + "</h3>"
    data_html = data_html + "<p>" + "Artificialisation :    <a target=\"_blank\" href=\"" + artificialisationSourcePage + "\">Dossier Cerema  </a>" + "</p>"
    data_html = data_html + "<p>" + "Logements Sociaux :    <a target=\"_blank\" href=\"" + sru2020SourcePage           + "\">Dossier Dreal   </a>" + "</p>"
    data_html = data_html + "<p>" + "Logements Construits : <a target=\"_blank\" href=\"" + sitadelSourcePage           + "\">Dossier Sitadel </a>" + "</p>"
    data_html = data_html + "<p>" + "Donnees Communes :     <a target=\"_blank\" href=\"" + metaDossierSourcePage       + "\">Dossier Insee   </a>" + "</p>"
    data_html = data_html + "<p>" + "Projections 2050 :     <a target=\"_blank\" href=\"" + projectionsSourcePage       + "\">Dossier Insee   </a>" + "</p>"
    data_html = data_html + "<p>" + "Projections Paca :     <a target=\"_blank\" href=\"" + projectionsPacaSourcePage   + "\">Dossier Omphale </a>" + "</p>"
    data_html = data_html + "<p>" + "Dossier " + ds.str("NOM_COMMUNE").title() + "  :  <a target=\"_blank\" href=\"" + ds["DOSSIER_INSEE"] + "\">Dossier Insee   </a>" + "</p>"
    global_context["HTML_SOURCES"] = data_html
    return data_html


def report_sru(ds: DataStore):
    """ SRU Data in HTML """
    data_dict = ds.get_row_as_dict(ds.store_index)
    df = pd.DataFrame(columns=['Item', '2017', '2020'])
    row = {'Item': "Residences Principales",     '2017': ds.number("SRU_RP_2017"),    '2020': ds.number("SRU_RP_2020")}
    df = df.append(row, ignore_index=True)
    row = {'Item': "Logements Sociaux",          '2017': ds.number("SRU_LLS_2017"),   '2020': ds.number("SRU_LLS_2020")}
    df = df.append(row, ignore_index=True)
    row = {'Item': "Taux Logements Sociaux",     '2017': ds.tauxp100("SRU_TX_LLS_2017"), '2020': ds.tauxp100("SRU_TX_LLS_2020")}
    df = df.append(row, ignore_index=True)

    carence2017 = round0(data_dict["SRU_RP_2017"] * (0.25 - data_dict["SRU_TX_LLS_2017"]), 0)
    carence2020 = round0(data_dict["SRU_RP_2020"] * (0.25 - data_dict["SRU_TX_LLS_2020"]), 0)
    row = {'Item': "Carence",                    '2017': round0str(carence2017), '2020' : round0str(carence2020)}
    df = df.append(row, ignore_index=True)
    row = {'Item': "Objectifs Trienaux",         '2017': ds.number("SRU_OBJ_2017_2019"), '2020': ds.number("SRU_OBJ_2020_2022")}
    df = df.append(row, ignore_index=True)

    mod3565_2017 = round(data_dict["SRU_OBJ_2017_2019"]/35*100, 0)
    mod3565_2020 = round(data_dict["SRU_OBJ_2020_2022"]/35*100, 0)
    mod3565_carence = round(carence2020/35*100, 0)
    row = {'Item': "Log 35/65 Trienaux",                  '2017': round0str(mod3565_2017), '2020': round0str(mod3565_2020)}
    df = df.append(row, ignore_index=True)

    evol_res_pr  = round(data_dict["SRU_RP_2020"]-data_dict["SRU_RP_2017"], 0)
    evol_carence = round((data_dict["SRU_RP_2020"] - data_dict["SRU_RP_2017"]) * 0.25, 0)
    log_sociaux  = round(carence2017 + evol_carence - carence2020, 0)
    html = "<h2>" + "Logements Sociaux " + ds.store_name + "</h2>"
    html = html   + "<p>" + df.to_html() + "</p>"
    html = html   + "<p>" + "Evolution 2017-2020 des RP : "               + round0str(evol_res_pr, 0)     + "</p>"
    html = html   + "<p>" + "Evolution 2017-2020 Carence : "              + round0str(evol_carence, 0)    + "</p>"
    html = html   + "<p>" + "LS Realises en 2017-2020 : "                 + round0str(log_sociaux, 0)     + "</p>"
    html = html   + "<p>" + "Total de logements a construire en 36/65 : " + round0str(mod3565_carence, 0) + "</p>"
    global_context["HTML_TABLE_SRU"] = html
    return html


def report_artificialisation(ds: DataStore):
    """ Artificialisation Data in HTML """
    data_dict = ds.get_row_as_dict(ds.store_index)
    html = "<h2>" + "Artificialisation sur 10 ans (2009-2020)" + "</h2>"
    html = html   + "<p>" + "Surface du Territoire       = " + round0str(data_dict["SURFACE_COMMUNE"] / 10000, 1) + " ha"+"</p>"
    html = html   + "<p>" + "Artificialisation 2009-2020 = " + round0str(data_dict["ART_TOTAL"] / 10000, 1)  + " ha"+"</p>"
    html = html   + "<p>" + "Supplement 2009-2020        = " + round0str(data_dict["ART_POURCENT"] , 1) + "%"+"</p>"
    global_context["HTML_TABLE_ARTIFICIALISATION"] = html
    return html


def report_historique(ds: DataStore):
    """ Historiques Data in HTML """
    data_dict = ds.get_row_as_dict(ds.store_index)
    # Summary Table
    df = pd.DataFrame(columns=['Item', '2008', '2013', '2018'])
    row = {'Item': "Population",             '2008': ds.number("P08_POP"),    '2013': ds.number("P13_POP"),    '2018': ds.number("P18_POP")}
    df = df.append(row, ignore_index=True)
    row = {'Item': "- Evolution #",          '2008': "-",                     '2013': round0str(data_dict["P13_POP"]-data_dict["P08_POP"]), '2018': round0str(data_dict["P18_POP"]-data_dict["P13_POP"])}
    df = df.append(row, ignore_index=True)
    row = {'Item': "- Evolution %",          '2008': "-",                     '2013': ds.percent("TXPOP_0813"), '2018': ds.percent("TXPOP_1318")}
    df = df.append(row, ignore_index=True)
    row = {'Item': "Population Menages ",    '2008': ds.number("C08_PMEN"),   '2013': ds.number("C13_PMEN"),   '2018': ds.number("C18_PMEN")}
    df = df.append(row, ignore_index=True)
    row = {'Item': "Population Hors Menages ", '2008': round0str(data_dict["P08_POP"]-data_dict["C08_PMEN"]),   '2013': round0str(data_dict["P13_POP"]-data_dict["C13_PMEN"]),   '2018': round0str(data_dict["P18_POP"]-data_dict["C18_PMEN"])}
    df = df.append(row, ignore_index=True)
    row = {'Item': "Menages ",               '2008': ds.number("C08_MEN"),    '2013': ds.number("C13_MEN"),    '2018': ds.number("C18_MEN")}
    df = df.append(row, ignore_index=True)
    row = {'Item': "Taille des Menages",     '2008': ds.number("TM_2008", 2), '2013': ds.number("TM_2013", 2), '2018': ds.number("TM_2018", 2)}
    df = df.append(row, ignore_index=True)
    row = {'Item': "- Evolution TM %",       '2008': "-",                     '2013': ds.percent("TXTM_0813", 2), '2018': ds.percent("TXTM_1318", 2)}
    df = df.append(row, ignore_index=True)
    row = {'Item': "Logements",              '2008': ds.number("P08_LOG"),    '2013': ds.number("P13_LOG"),    '2018': ds.number("P18_LOG")}
    df = df.append(row, ignore_index=True)
    row = {'Item': "- Residences Principales", '2008': ds.number("P08_RP"),   '2013': ds.number("P13_RP"),     '2018': ds.number("P18_RP")}
    df = df.append(row, ignore_index=True)
    row = {'Item': "- Residences Secondaires", '2008': ds.number("P08_RSECOCC"), '2013': ds.number("P13_RSECOCC"), '2018': ds.number("P18_RSECOCC")}
    df = df.append(row, ignore_index=True)
    row = {'Item': "- Logements Vacants",    '2008': ds.number("P08_LOGVAC"), '2013': ds.number("P13_LOGVAC"),  '2018': ds.number("P18_LOGVAC")}
    df = df.append(row, ignore_index=True)
    row = {'Item': "- Maisons",              '2008': ds.number("P08_MAISON"), '2013': ds.number("P13_MAISON"),  '2018': ds.number("P18_MAISON")}
    df = df.append(row, ignore_index=True)
    row = {'Item': "- Appartements ",        '2008': ds.number("P08_APPART"), '2013': ds.number("P13_APPART"),  '2018': ds.number("P18_APPART")}
    df = df.append(row, ignore_index=True)
    row = {'Item': "Proprietaires en RP",    '2008': ds.number("P08_RP_PROP"), '2013': ds.number("P13_RP_PROP"), '2018': ds.number("P18_RP_PROP")}
    df = df.append(row, ignore_index=True)
    row = {'Item': "Locataires en RP",       '2008': ds.number("P08_RP_LOC"),  '2013': ds.number("P13_RP_LOC"),  '2018': ds.number("P18_RP_LOC")}
    df = df.append(row, ignore_index=True)

    html = "<h2>" + "Historique " + ds.store_name + "</h2>"
    html = html   + "<p>" + df.to_html() + "</p>"
    global_context["HTML_TABLE_HISTORIQUE"] = html
    return html


def report_projection(ds: DataStore):
    """ Projection Data in HTML """
    data_dict = ds.get_row_as_dict(ds.store_index)
    # Summary Table
    df = pd.DataFrame(columns=['Item', '2020', '2030', '2040', '2050'])
    row = {'Item': "Population",             '2020': ds.number("POP_2020"),   '2030': ds.number("POP_2030"),    '2040': ds.number("POP_2040"),     '2050': ds.number("POP_2050")}
    df = df.append(row, ignore_index=True)
    row = {'Item': "- Evolution %",          '2020': ds.percent("TXPOP_1318", 2), '2030': ds.percent("TX_POP_2030", 2), '2040': ds.percent("TX_POP_3040", 2),  '2050': ds.percent("TX_POP_4050", 2)}
    df = df.append(row, ignore_index=True)
    df = df.append(row, ignore_index=True)
    row = {'Item': "Taille des Menages",     '2020': ds.number("TM_2020", 2), '2030': ds.number("TM_2030", 2),  '2040': ds.number("TM_2040", 2),   '2050': ds.number("TM_2050", 2)}
    df = df.append(row, ignore_index=True)
    row = {'Item': "Besoins en Logements = ",  '2020': data_dict["LOG_2020"], '2030': data_dict["LOG_2030"],    '2040': data_dict["LOG_2040"],     '2050': data_dict["LOG_2050"]}
    df = df.append(row, ignore_index=True)
    row = {'Item': "Besoins en Logements # 10 ans",  '2020': "-",             '2030': round0str(data_dict["LOG_2030"]-data_dict["LOG_2020"]),      '2040': round0str(data_dict["LOG_2040"]-data_dict["LOG_2030"]),      '2050': round0str(data_dict["LOG_2050"]-data_dict["LOG_2040"])}
    df = df.append(row, ignore_index=True)
    row = {'Item': "Besoins en Logements # / an ", '2020': "-",               '2030': round0str((data_dict["LOG_2030"]-data_dict["LOG_2020"])/10), '2040': round0str((data_dict["LOG_2040"]-data_dict["LOG_2030"])/10), '2050': round0str((data_dict["LOG_2050"]-data_dict["LOG_2040"])/10)}
    df = df.append(row, ignore_index=True)

    html = "<h2>" + "Projections " + ds.store_name + "</h2>"
    html = html   + "<p>" + df.to_html() + "</p>"
    global_context["HTML_TABLE_PROJECTIONS"] = html
    return html


def report_summary_data(ds: DataStore):
    """ Summary Data in HTML """
    data_dict = {}
    df = pd.DataFrame(columns=['Key', 'Value', 'Meta'])
    for column in ds.data_frame:
        data_dict[column] = ds.data_frame[column][ds.store_index]
        row = {'Key': column, 'Value': str(ds.data_frame[column][ds.store_index]), 'Meta': ds.data_frame[column]["meta"]}
        df = df.append(row, ignore_index=True)
    data_html = "<h3>" + "Donnees Brutes Consolidees pour : " + ds.store_name  + "</h3>"
    data_html = data_html + df.to_html(index=True)
    global_context["HTML_TABLE_SUMMARY"] = data_html
    return data_html


def report_full_data(ds: DataStore):
    """ Full Data in HTML """
    data_html = "<h3>" + "Toutes les Donnees - Details : " + ds.store_name + "</h3>"
    data_html = data_html + ds.data_frame.to_html(index=True)
    global_context["HTML_FULL_DATA"] = data_html
    return data_html


###
### Plotting
###

index_figure = 0


def new_figure():
    global index_figure
    index_figure += 1
    plt.close('all')
    plt.figure(index_figure)
    return plt


def plots(ds: DataStore):
    """ Generates Plots from plots.json """
    with open(plots_file) as json_file:
        desc = jsonc.load(json_file)
    for plot in desc["Plots"] :
        if (plot["Type"] == "LINES")  : plots_lines(plot, ds)
        if (plot["Type"] == "PIE")    : plots_pie(plot, ds)
        if (plot["Type"] == "DONUT")  : plots_donut2(plot, ds)
        if (plot["Type"] == "DONUT2") : plots_donut1(plot, ds)


def plots_donut1(plot, ds: DataStore):
    """ Generic : Draw plot Donut1 with DataStore values """
    key = plot["Key"]
    ctx1 = ds.get_row_as_dict()
    ctx2 = {**globals(), **locals()}
    # Title & Legend
    titre = eval(plot["Title"], ctx1, ctx2)
    fig, ax = plt.subplots(figsize=(6, 3), subplot_kw=dict(aspect="equal"))

    labels = eval(plot["Labels"], ctx1, ctx2)
    data   = eval(plot["Values"], ctx1, ctx2)

    """
    def func(pct, allvals):
        absolute = int(round(pct/100.*np.sum(allvals)))
        return "{:.1f}%\n({:d})".format(pct, absolute)

    recipe = ["Residences Principales :\n" + str(data_dict["TX_RES_PR_18"]*100+"%"),
              "Residences Secondaires :\n" + str(data_dict["TX_RES_SEC_18"]*100+"%"),
              "Logements Vacants :\n"      + str(data_dict["TX_RES_VAC_18"]*100)+"%"]
    """

    wedges, texts = ax.pie(data, wedgeprops=dict(width=0.5), startangle=-40, normalize=True)
    bbox_props = dict(boxstyle="square,pad=0.3", fc="w", ec="k", lw=0.72)
    kw = dict(arrowprops=dict(arrowstyle="-") , bbox=bbox_props, zorder=0, va="center")
    for i, p in enumerate(wedges):
        ang = (p.theta2 - p.theta1)/2. + p.theta1
        y = np.sin(np.deg2rad(ang))
        x = np.cos(np.deg2rad(ang))
        horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
        connectionstyle = "angle,angleA=0,angleB={}".format(ang)
        kw["arrowprops"].update({"connectionstyle": connectionstyle})
        ax.annotate(labels[i], xy=(x, y), xytext=(1.35*np.sign(x), 1.4*y) , horizontalalignment=horizontalalignment, **kw)

    ax.set_title(titre, color=plot["ColorTitle"], weight='bold')

    ## Save Locally
    image_file_name = output_dir + ds.get_fullname() + "_" + key.replace(" ", "_") + ".png"
    fig.savefig(image_file_name)
    ds.add_metric(key="GRAPHIQUE_"+key.upper() , meta=titre,source="PLOT" , mode="EQUAL" , type="STR" , data=image_file_name , expr=image_file_name)
    return image_file_name


def plots_donut2(plot, ds: DataStore):
    """ Generic : Draw plot Donut2 with DataStore values """
    key = plot["Key"]
    ctx1 = ds.get_row_as_dict()
    ctx2 = {**globals(), **locals()}
    # Title & Legend
    titre = eval(plot["Title"], ctx1, ctx2)
    fig, ax = plt.subplots(figsize=(6, 3), subplot_kw=dict(aspect="equal"))

    values   = eval(plot["Values"], ctx1, ctx2)
    labels   = eval(plot["Labels"], ctx1, ctx2)
    colors   = eval(plot["Colors"], ctx1, ctx2)

    # Pie Chart
    _, _, autotexts = plt.pie(values, colors=colors, labels=labels,
                    autopct='%1.1f%%', pctdistance=0.75, normalize=True)
    for autotext in autotexts:
        autotext.set_color('white')

    # Draw circle
    centre_circle = plt.Circle((0, 0), 0.50, fc='white')
    fig = plt.gcf()

    # Adding Circle in Pie chart
    fig.gca().add_artist(centre_circle)

    # Add Legends
    # plt.legend(labels, loc="upper right")

    # Adding Title of chart
    plt.title(titre, color=plot["ColorTitle"], weight='bold')

    ## Save Locally
    image_file_name = output_dir + ds.get_fullname() + "_" + key.replace(" ", "_") + ".png"
    fig.savefig(image_file_name)
    ds.add_metric(key="GRAPHIQUE_"+key.upper() , meta=titre , source="PLOT" , mode="EQUAL" , type="STR" , data=image_file_name , expr=image_file_name)
    return image_file_name


def plots_pie(plot, ds: DataStore):
    """ Generic : Draw plot Pie with DataStore values """
    plt = new_figure()
    key = plot["Key"]
    ctx1 = ds.get_row_as_dict()
    ctx2 = {**globals(), **locals()}
    # Title & Legend
    titre  = eval(plot["Title"], ctx1, ctx2)
    legend = eval(plot["Legende"], ctx1, ctx2)
    plt.title(titre, color=plot["ColorTitle"], weight='bold')
    fig, ax = plt.subplots(figsize=(6, 3), subplot_kw=dict(aspect="equal"))

    labels  = eval(plot["Labels"], ctx1, ctx2)
    data    = eval(plot["Values"], ctx1, ctx2)
    colors  = eval(plot["Colors"], ctx1, ctx2)

    def func(pct, allvals):
        absolute = int(round(pct/100.*np.sum(allvals)))
        return "{:.1f}%\n({:d})".format(pct, absolute)
    # Pie Chart
    wedges, texts, autotexts = ax.pie(data, autopct=lambda pct: func(pct, data), textprops=dict(color="w"), colors=colors)
    ax.legend(wedges, labels, title=legend, loc="center left", bbox_to_anchor=(1, 0, 0.5, 1))
    plt.setp(autotexts, size=8, weight="bold")
    plt.title(titre, color=plot["ColorTitle"], weight='bold')

    ## Save Locally
    image_file_name = output_dir + ds.get_fullname() + "_" + key.replace(" ", "_") + ".png"
    plt.savefig(image_file_name)
    ds.add_metric(key="GRAPHIQUE_"+key.upper() , meta=titre , source="PLOT" , mode="EQUAL" , type="STR" , data=image_file_name , expr=image_file_name)
    return image_file_name


def plots_lines(plot, ds: DataStore):
    """ Generic : Draw plot Lines with DataStore values """
    plt = new_figure()
    key = plot["Key"]
    ctx1 = ds.get_row_as_dict()
    ctx2 = {**globals(), **locals()}
    # Naming the axis
    plt.xlabel(plot["xLabel"], color=plot["ColorLabel"])
    plt.ylabel(plot["yLabel"], color=plot["ColorLabel"])
    if (("yLimits" in plot) and (plot["yLimits"] != "")):
        plt.ylim(eval(plot["yLimits"], ctx1, ctx2))
    # Draw Series
    for serie in plot["Series"] :
        # if ("Condition"  in serie) :
        #     if (not eval(serie["Condition"], ctx1, ctx2)) : continue
        if (serie["Smooth"] is True) :
            xsmooth, ysmooth = plot_smooth(eval(serie["xValues"], ctx1, ctx2), eval(serie["yValues"], ctx1, ctx2))
        else:
            xsmooth, ysmooth = eval(serie["xValues"], ctx1, ctx2), eval(serie["yValues"], ctx1, ctx2)
        color = "Teal"
        if ("Color"  in serie) : color = serie["Color"]
        if ("ColorC" in serie) : color = "#"+serie["ColorC"]
        plt.plot(xsmooth, ysmooth, color=color, linestyle=serie["Style"], linewidth=serie["Width"], label=serie["Label"])
    # Title & Legend
    titre = eval(plot["Title"], ctx1, ctx2)
    plt.title(titre, color=plot["ColorTitle"], weight='bold')
    if (("Legend" in plot) and (plot["Legend"] != "")):
        if (plot["Legend"] == "below") :
            # plt.legend(bbox_to_anchor=(1.0, 1.0))
            plt.legend(loc='lower center', bbox_to_anchor=(0.8, -0.1), fancybox=True, shadow=True)
        else:
            plt.legend(fancybox=True, shadow=True)
    else:
        plt.plot(legend=None)
    ## Save Locally
    image_file_name = output_dir + ds.get_fullname() + "_" + key.replace(" ", "_") + ".png"
    plt.savefig(image_file_name)
    ds.add_metric(key="GRAPHIQUE_"+key.upper() , meta=titre , source="PLOT" , mode="EQUAL" , type="STR" , data=image_file_name , expr=image_file_name)
    return image_file_name


def plot_smooth(x_values : list[int], y_values : list[int]):
    """ Smoothing Plot Lines """
    # 300 represents number of points to make between T.min and T.max
    x_values = np.array(x_values)
    y_values = np.array(y_values)
    xnew = np.linspace(x_values.min(), x_values.max(), 300)
    spl  = make_interp_spline(x_values, y_values, k=3, bc_type="natural")  # type: BSpline
    ysmooth = spl(xnew)
    return xnew, ysmooth


def plot_logements(ds: DataStore):
    """ Graphique Logements en HTML et png. """
    data_dict = ds.get_row_as_dict(ds.store_index)
    label = "Logements"

    # Draw plot
    plt = new_figure()

    xsmooth, ysmooth = plot_smooth([2008, 2013, 2018, 2020],
                                   [data_dict["P08_RP"]   - data_dict["P08_RP"],
                                    data_dict["P13_RP"]   - data_dict["P08_RP"],
                                    data_dict["P18_RP"]   - data_dict["P08_RP"],
                                    data_dict["LOG_2020"] - data_dict["P08_RP"]])
    plt.plot(xsmooth, ysmooth, color=data_dict["THEME_COLOR"],  linestyle='-',       linewidth=3, label="Residences Principales des menages - Historique")

    xsmooth, ysmooth = plot_smooth([2020, 2030],
                                   [data_dict["LOG_2020"] - data_dict["P08_RP"],
                                    data_dict["LOG_2030"] - data_dict["P08_RP"]])
    plt.plot(xsmooth, ysmooth, color=data_dict["THEME_COLOR"],  linestyle='--',     linewidth=3, label="Residences Principales des menages - Projection des Besoins")

    xsmooth, ysmooth = plot_smooth([2008, 2013, 2018],
                                   [data_dict["P08_RP"] - data_dict["P08_RP"] + data_dict["P08_RSECOCC"] - data_dict["P08_RSECOCC"] + data_dict["P08_LOGVAC"] - data_dict["P08_LOGVAC"],
                                    data_dict["P13_RP"] - data_dict["P08_RP"] + data_dict["P13_RSECOCC"] - data_dict["P08_RSECOCC"] + data_dict["P13_LOGVAC"] - data_dict["P08_LOGVAC"],
                                    data_dict["P18_RP"] - data_dict["P08_RP"] + data_dict["P18_RSECOCC"] - data_dict["P08_RSECOCC"] + data_dict["P18_LOGVAC"] - data_dict["P08_LOGVAC"]])
    plt.plot(xsmooth, ysmooth, color='#74248f', linestyle='-',      linewidth=3, label="Residences Principales + Secondaires + Vacants")

    xsmooth, ysmooth = plot_smooth([2018, 2020],
                                   [data_dict["P18_RP"] - data_dict["P08_RP"] + data_dict["P18_RSECOCC"] - data_dict["P08_RSECOCC"] + data_dict["P18_LOGVAC"] - data_dict["P08_LOGVAC"],
                                    data_dict["NOUV_LOG_0813"] + data_dict["NB_LGT_TOT_COMMENCES_1316"] + data_dict["NB_LGT_TOT_COMMENCES_1721"]])
    plt.plot(xsmooth, ysmooth, color='#74248f', linestyle='dotted', linewidth=3, label="Indeterminees (Principales / Secondaires / Non-Vendues)")

    xsmooth, ysmooth = plot_smooth([2013, 2016, 2020],
                                   [data_dict["NOUV_LOG_0813"],
                                    data_dict["NOUV_LOG_0813"] + data_dict["NB_LGT_TOT_COMMENCES_1316"],  # Logements Construits
                                    data_dict["NOUV_LOG_0813"] + data_dict["NB_LGT_TOT_COMMENCES_1316"] + data_dict["NB_LGT_TOT_COMMENCES_1721"]])
    plt.plot(xsmooth, ysmooth, color='#663300', linestyle='-',      linewidth=3, label="Logements Construits")

    xsmooth, ysmooth = plot_smooth([2020, 2021],
                                   [data_dict["NOUV_LOG_0813"] + data_dict["NB_LGT_TOT_COMMENCES_1316"] + data_dict["NB_LGT_TOT_COMMENCES_1721"],
                                    data_dict["NOUV_LOG_0813"] + data_dict["PROJ_LOG_REALISES_2021"]])
    plt.plot(xsmooth, ysmooth, color='#663300', linestyle='dotted', linewidth=3, label="Projection Logements Construits")

    if ((data_dict["SRU_CARENCE_2020"] != 0) or (data_dict["NB_LGT_PRET_LOC_SOCIAL_1316"] + data_dict["NB_LGT_PRET_LOC_SOCIAL_1721"]) != 0):
        xsmooth, ysmooth = plot_smooth([2013, 2016, 2020],
                                       [data_dict["NOUV_LOG_0813"],
                                        data_dict["NOUV_LOG_0813"] + data_dict["NB_LGT_PRET_LOC_SOCIAL_1316"],  # Logements Sociaux Construits
                                        data_dict["NOUV_LOG_0813"] + data_dict["NB_LGT_PRET_LOC_SOCIAL_1316"] + data_dict["NB_LGT_PRET_LOC_SOCIAL_1721"]])
        plt.plot(xsmooth, ysmooth, color='#006600', linestyle='dotted', linewidth=3, label="Logements Sociaux Construits")

    # Naming the axis
    plt.xlabel('Annees')
    plt.ylabel('Logements')

    # Title & Legend
    plt.title(label + " sur " + ds.store_name, color=data_dict["THEME_COLOR"], weight='bold')
    # plt.legend(bbox_to_anchor=(1.0, 1.0))
    # plt.legend(loc='lower center', bbox_to_anchor=(0.5, -0.1), fancybox=True, shadow=True)
    plt.plot(legend=None)

    ## Save Locally
    image_file_name = output_dir + ds.get_fullname() + "_" + label.replace(" ", "_") + ".png"
    plt.savefig(image_file_name)

    ## Save for HTML Format
    encoded_fig = fig_to_base64(plt)
    browser_html = '<h2>'+label+'</h2><p><img src="data:image/png;base64, {}"></p>'.format(encoded_fig.decode('utf-8'))
    browser_html = browser_html + "<p> Taux de Residences Secondaires : " + round0str(100 * round(data_dict["TX_RES_SEC_18"], 4), 1) + "%</p>"
    browser_html = browser_html + "<p> Taux de Residences Vacantes : " + round0str(100 * round(data_dict["TX_RES_VAC_18"],    4), 1) + "%</p>"

    browser_html = browser_html + "<p> Periode 2013 - 2020 : " + str(round(ds["EXCES_BESOINS_1320"])) + "</p>"
    browser_html = browser_html + "<p> - Besoins en logements : " + str(round(ds["BESOINS_1320"])) + "</p>"
    browser_html = browser_html + "<p> - Logements Construits : " + str(round(ds["NB_LGT_TOT_COMMENCES_1321"])) + "</p>"
    browser_html = browser_html + "<p> - Exces de Logements Construits par rapport aux besoins : " + str(round(ds["EXCES_BESOINS_1320"])) + "</p>"
    browser_html = browser_html + "<p> -> Nouvelles Residences Secondaires : "  + str(round(ds["NOUV_RESSEC_1318"])) + "</p>"
    browser_html = browser_html + "<p> -> Nouvelles Residences Vacantes : "     + str(round(ds["NOUV_LOGVAC_1318"])) + "</p>"
    browser_html = browser_html + "<p> -> Indetermines / Non-Vendus : "     + str(round(ds["EXCES_BESOINS_1320"]-ds["NOUV_LOGVAC_1318"]-ds["NOUV_RESSEC_1318"])) + "</p>"
    browser_html = browser_html + "<p> Periode 2013 - 2021 : " + str(round(ds["EXCES_BESOINS_1320"])) + "</p>"
    browser_html = browser_html + "<p> - Logements Construits : " + str(ds["NB_LGT_TOT_COMMENCES_1321"])     + "</p>"
    browser_html = browser_html + "<p> - Logements Sociaux  : "   + str(ds["NB_LGT_PRET_LOC_SOCIAL_1321"]) + "</p>"
    browser_html = browser_html + "<p> - Taux LS Construits : " + str(round(100 * ds["TX_LGT_PRET_LOC_SOCIAL_1321"], 2)) + "% </p>"
    # email_html   = "<img alt=" + str(lKpi.get_data(c_name)) + " src=\"cid:" + image_file_name + "\" >"
    global_context["HTML_PLOT_LOGEMENTS"] = browser_html
    return browser_html


def plot_logements_pie(ds: DataStore):

    label = "Repartition des Logements"
    data_dict = ds.get_row_as_dict(ds.store_index)

    fig, ax = plt.subplots(figsize=(6, 3), subplot_kw=dict(aspect="equal"))

    ingredients = ["Principales",            "Secondaires",            "Vacantes"]
    data =        [data_dict["P18_RP"],      data_dict["P18_RSECOCC"], data_dict["P18_LOGVAC"]]

    def func(pct, allvals):
        absolute = int(round(pct/100.*np.sum(allvals)))
        return "{:.1f}%\n({:d})".format(pct, absolute)

    wedges, texts, autotexts = ax.pie(data, autopct=lambda pct: func(pct, data), textprops=dict(color="w"))

    ax.legend(wedges, ingredients,
              title="Residences",
              loc="center left",
              bbox_to_anchor=(1, 0, 0.5, 1))

    plt.setp(autotexts, size=8, weight="bold")

    ax.set_title(label)

    ## Save Locally
    image_file_name = output_dir + ds.get_fullname() + "_" + label.replace(" ", "_") + ".png"
    fig.savefig(image_file_name)

    ## Save for HTML Format
    encoded_fig = fig_to_base64(fig)
    browser_html = '<h2>'+label+'</h2><p><img src="data:image/png;base64, {}"></p>'.format(encoded_fig.decode('utf-8'))
    # email_html   = "<img alt=" + str(lKpi.get_data(c_name)) + " src=\"cid:" + image_file_name + "\" >"
    global_context["HTML_PLOT_TAILLE_DES_MENAGES"] = browser_html
    return browser_html


def plot_logements_donuts(ds: DataStore):

    label = "Repartition des Logements"
    data_dict = ds.get_row_as_dict(ds.store_index)

    # Draw plot
    fig, ax = plt.subplots(figsize=(6, 3), subplot_kw=dict(aspect="equal"))

    recipe = ["Residences Principales :\n" + str(data_dict["TX_RES_PR_18"]*100+"%"),
              "Residences Secondaires :\n" + str(data_dict["TX_RES_SEC_18"]*100+"%"),
              "Logements Vacants :\n"      + str(data_dict["TX_RES_VAC_18"]*100)+"%"]

    data = [data_dict["TX_RES_PR_18"], data_dict["TX_RES_SEC_18"], data_dict["TX_RES_VAC_18"]]

    wedges, texts = ax.pie(data, wedgeprops=dict(width=0.5), startangle=-40)

    bbox_props = dict(boxstyle="square,pad=0.3", fc="w", ec="k", lw=0.72)
    kw = dict(arrowprops=dict(arrowstyle="-"),
              bbox=bbox_props, zorder=0, va="center")

    for i, p in enumerate(wedges):
        ang = (p.theta2 - p.theta1)/2. + p.theta1
        y = np.sin(np.deg2rad(ang))
        x = np.cos(np.deg2rad(ang))
        horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
        connectionstyle = "angle,angleA=0,angleB={}".format(ang)
        kw["arrowprops"].update({"connectionstyle": connectionstyle})
        ax.annotate(recipe[i], xy=(x, y), xytext=(1.35*np.sign(x), 1.4*y),
                    horizontalalignment=horizontalalignment, **kw)

    ax.set_title("Repartition des Logements")

    ## Save Locally
    image_file_name = output_dir + ds.get_fullname() + "_" + label.replace(" ", "_") + ".png"
    fig.savefig(image_file_name)

    ## Save for HTML Format
    encoded_fig = fig_to_base64(fig)
    browser_html = '<h2>'+label+'</h2><p><img src="data:image/png;base64, {}"></p>'.format(encoded_fig.decode('utf-8'))
    # email_html   = "<img alt=" + str(lKpi.get_data(c_name)) + " src=\"cid:" + image_file_name + "\" >"
    global_context["HTML_PLOT_TAILLE_DES_MENAGES"] = browser_html
    return browser_html


def plot_taille_menages(ds: DataStore):
    """ Graphique taille des menages en HTML et png. """
    data_dict = ds.get_row_as_dict(ds.store_index)
    label = "Taille des Menages"

    # Draw plot
    plt = new_figure()
    xsmooth, ysmooth = plot_smooth([2008, 2013, 2018, 2020],
                                   [data_dict["TM_2008"],
                                    data_dict["TM_2013"],
                                    data_dict["TM_2018"],
                                    data_dict["TM_2020"]])
    plt.plot(xsmooth, ysmooth, color=data_dict["THEME_COLOR"],  linestyle='-',       linewidth=3, label=label + " Historique")

    xsmooth, ysmooth = plot_smooth([2020, 2030, 2040, 2050],
                                   [data_dict["TM_2020"],
                                    data_dict["TM_2030"],
                                    data_dict["TM_2040"],
                                    data_dict["TM_2050"]])
    plt.plot(xsmooth, ysmooth, color=data_dict["THEME_COLOR"],  linestyle='dotted',  linewidth=3, label=label + " Projetee")

    # Naming the axis
    plt.xlabel('Annees', color='grey')
    plt.ylabel('Taille des Menages', color='grey')
    plt.ylim((1, 3))

    # Title & Legend
    plt.title(label + " sur " + ds.store_name, color=data_dict["THEME_COLOR"], weight='bold')
    plt.legend()

    ## Save Locally
    image_file_name = output_dir + ds.get_fullname() + "_" + label.replace(" ", "_") + ".png"
    plt.savefig(image_file_name)

    ## Save for HTML Format
    encoded_fig = fig_to_base64(plt)
    browser_html = '<h2>'+label+'</h2><p><img src="data:image/png;base64, {}"></p>'.format(encoded_fig.decode('utf-8'))
    # email_html   = "<img alt=" + str(lKpi.get_data(c_name)) + " src=\"cid:" + image_file_name + "\" >"
    global_context["HTML_PLOT_TAILLE_DES_MENAGES"] = browser_html
    return browser_html


def plot_population(ds: DataStore):
    """ Graphique Population en HTML et png."""
    data_dict = ds.get_row_as_dict(ds.store_index)
    label = "Population"

    # Draw plot
    plt = new_figure()

    xsmooth, ysmooth = plot_smooth([2008, 2013, 2018, 2020],
                                   [data_dict["P08_POP"],
                                    data_dict["P13_POP"],
                                    data_dict["P18_POP"],
                                    data_dict["POP_2020"]])
    plt.plot(xsmooth, ysmooth, color=data_dict["THEME_COLOR"],   linestyle='-',      linewidth=3, label=label + " Historique")

    xsmooth, ysmooth = plot_smooth([2020, 2030, 2040, 2050],
                                   [data_dict["POP_2020"],
                                    data_dict["POP_2030"],
                                    data_dict["POP_2040"],
                                    data_dict["POP_2050"]])
    plt.plot(xsmooth, ysmooth, color=data_dict["THEME_COLOR"],   linestyle='dotted', linewidth=3, label=label + " Projetee")

    # Naming the axis
    plt.xlabel('Annees', color='grey')
    plt.ylabel('Population', color='grey')
    # plt.ylim( (data_dict["POP_2050"]-500, data_dict["POP_2050"]+500) )

    # Title & Legend
    plt.title(label + " sur " + ds.store_name + " (Base 2008)", color=data_dict["THEME_COLOR"], weight='bold')
    plt.legend()

    ## Save Locally
    image_file_name = output_dir + ds.get_fullname() + "_" + label.replace(" ", "_") + ".png"
    plt.savefig(image_file_name)

    ## Save for HTML Format
    encoded_fig = fig_to_base64(plt)
    browser_html = '<h2>'+label+'</h2><p><img src="data:image/png;base64, {}"></p>'.format(encoded_fig.decode('utf-8'))
    # email_html   = "<img alt=" + str(lKpi.get_data(c_name)) + " src=\"cid:" + image_file_name + "\" >"
    global_context["HTML_PLOT_POPULATION"] = browser_html
    return browser_html

###
### Reports
###


def gen_index(region="93"):
    """ Generates Index for a Region or France (if region = None) """
    report_region_dict(region, filename=france_file)
    report_select_dict(region, filename=selection_file)
    return render_index()


def gen_tracker(ds : DataStore):
    """ Generates Report for DataStore """
    ds.render_report(input_dir  + "tracker_template.html", suffix="_tracker")


def gen_report(ds : DataStore):
    """ Generates Report for DataStore """
    global_context["HTML_PLOT_LOGEMENTS"]          = plot_logements(ds)
    global_context["HTML_PLOT_LOGEMENTS_PIE"]      = plot_logements_pie(ds)
    global_context["HTML_PLOT_POPULATION"]         = plot_population(ds)
    global_context["HTML_PLOT_TAILLE_DES_MENAGES"] = plot_taille_menages(ds)
    global_context["HTML_SOURCES"]                 = report_source(ds)
    global_context["HTML_DIAGNOSTIC"]              = report_diagnostic(ds)
    global_context["HTML_TABLE_HISTORIQUE"]        = report_historique(ds)
    global_context["HTML_TABLE_PROJECTIONS"]       = report_projection(ds)
    global_context["HTML_TABLE_ARTIFICIALISATION"] = report_artificialisation(ds)
    global_context["HTML_TABLE_SRU"]               = report_sru(ds)
    global_context["HTML_TABLE_SUMMARY"]           = report_summary_data(ds)
    global_context["HTML_FULL_DATA"]               = report_full_data(ds)
    return ds.render_report()


def readme_to_html():
    print_green("> README.md to HTML     : " + "README.html")
    markdown.markdownFromFile(input='README.md', output='README.html')


def ftp_push_ds(ds : DataStore):
    global FAST
    file_list = list()
    file_dir     = output_dir
    file_prefix  = ds.get_fullname()
    file_exts     = ["_s.json"] if FAST else [".xlsx", ".csv", "_s.json" ]
    for file_ext in file_exts :
        file_list.append(file_dir + file_prefix + file_ext)
    if (not FAST):
        for png_file in [f for f in os.listdir(output_dir) if re.match(ds.get_fullname().upper()+'.*\.png', f)] :
            file_list.append(output_dir + png_file)
        for png_file in [f for f in os.listdir(output_dir) if re.match(ds.get_fullname()+'.*\.png', f)] :
            file_list.append(output_dir + png_file)
    ftp_push_file(file_list)


def ftp_push_file(filename, p_host="ftpupload.net", p_user="epiz_30239961", p_password="oqEwtTaACCaANF", p_directory="htdocs", prod=True):
    host = p_host
    user = p_user
    password = p_password
    directory = p_directory
    if (FTP_PROD and (prod == True)):
        host = "sobrietefonciere.fnepaca.fr"
        user = "sobrietefonciere@fnepaca.fr"
        password = "6TlHc!cjmd8u"
        directory = ""
    if (FTP_PROD and (prod == True)):
        # ftp_push_file(filename, prod=False)
        pass
    if (isinstance(filename, str)):
        filename = [filename]
    if (isinstance(filename, list)):
        ftp = None
        try:
            ftp = ftplib.FTP(host)
            ftp.login(user, password)
            ftp.cwd(directory)
            # remote_files = ftp.nlst()
            # print(remote_files)
            for file in filename:
                if (not os.path.exists(file)):
                    print_red("File not found : " + file)
                    continue
                print_blue("FTP Push [" + host + "] : " + file)
                ftp.storbinary('STOR ' + file.replace("\\" , "/"), open(file, 'rb'))
            ftp.close()
        except Exception as e:
            print_red("FTP Upload Failed for  [" + host + "] : " + str(filename))
            print_red(str(e))
            logging.error("FTP Upload Failed for  [" + host + "] : " + str(filename))
            logging.error(str(e))
            if (ftp) : ftp.close()


def ftp_push_files():
    global FAST
    filelist = ["input/plots.json",             "output/select.json",
                "output/calculations.json",     "output/datametrics.json",  "output/diagnostics.json",
                "index.html",
                "ConsommationFonciereV3.html",  "ConsommationFonciereV3.js"
               ]
    if not FAST:
        filelist.extend([
                "input/Configuration.xlsx",
                "index.html",
                "output/france.json",
                "Header.png", "Body.png",
                "README.md",  "README.html",    "README.dillinger.html",
                "ConsommationFonciere.py",
                "input/QuestionnerUnProjetDePLU-SCoT.pdf", "input/Alpes-Maritimes-NoteExcedentsLogements.pdf",
                "input/Guide_Pratique_Artif_complet.pdf",  "input/GuideSF_6_ZCZAE_web.pdf",
                "input/GuideSF_16_Observatoires_web.pdf",  "input/GuideSF_17_fiscalite_web.pdf",
                "input/Urbascope.pdf",
                "ConsommationFonciere.html",    "ConsommationFonciere.js",
                "ConsommationFonciereV2.html",  "ConsommationFonciereV2.js",
                "input/Legend_Logements.png",   "input/CommentCaMarche.png",
                "input/Densite.png",            "input/Densite_1.jpg",
                "input/Densite_2.jpg",          "input/Densite_3.jpg",
                "input/Densite_4.jpg",          "input/Densite_5.png",
                "input/Densite_6.jpg",          "input/Densite_7.png",
                "input/Gadseca-Logo.png",       "input/Gadseca_50Ans.jpg",
                "input/Logo2-Vert-FV.png",      "input/Logo2-Vert.png",
                "input/Logo3-Color-FV.png",     "input/Logo3-Color.png",
                "input/CartePaca1.jpg",         "input/CartePaca2.png",
                "input/CartePaca4.png",         "input/CartoPaca4.png",
                "input/CartoPaca3.png",         "input/CartoPaca3-Green.png",
                "input/myShophia-Logo.jpg",     "input/CartoPaca3-Light-Green.png",
                ])
    ftp_push_file(filelist)


def report_commune(code_insee : str = None, code_postal: str = None, force=True, data_only : bool = False, ftp_push : bool = False):
    """ Generates Report for a COMMUNE identified by Code INSEE or Postal """
    commune = None
    if (code_insee):
        commune = nom_commune(code_insee=code_insee, clean=True)
    elif (code_postal):
        code_insee, commune = get_code_insee_commune(code_postal)
    if ((not commune) or (str(commune).startswith("Pas"))):
        print_red("COMMUNE non trouvee pour code : INSEE ["+str(code_insee+"] POSTAL : ["+str(code_postal)+"]"))
        return None

    if (code_insee in communes_zone("SCoT_Ouest")):
        print_blue("Scot Ouest pour Commune "+str(commune)+" : INSEE ["+str(code_insee+"]"))
        scot_ouest(code_insee=code_insee, start_date="2020-01-01")

    if (commune.upper() == "ANTIBES"):
        print_blue("Scot Ouest pour Commune "+str(commune)+" : INSEE ["+str(code_insee+"]"))
        scot_ouest(code_insee=code_insee, start_date="2018-01-01", file="scot_Antibes")

    entite = entite_commune
    name   = commune
    code   = code_insee
    return DataStore(store_name=name, store_type=entite, store_code=code).report(force=force, data_only=data_only, ftp_push=ftp_push)


def report_zone(zone_name: str = "SCoT_Ouest_Littoral", force=True, with_communes=False, data_only : bool = False, ftp_push : bool = False):
    """ Generates Report for a Zone identified by Zone Name  """
    entite = entite_zone
    name   = clean_name(zone_name)
    code   = clean_name(zone_name)
    if ((not name) or (str(name).startswith("Pas")) or (zone_name not in list_zones())):
        print_red("Zone non trouvee pour code :  [" + str(zone_name) + "]")
        return None
    if (with_communes):
        comm_len = len(communes_zone(name))
        comm_idx = 0
        for commune in communes_zone(name):
            comm_idx = comm_idx + 1
            print_grey("###> Commune " + str(comm_idx) + "/" + str(comm_len))
            report_commune(code_insee=str(commune), force=force, data_only=data_only, ftp_push=ftp_push)
    return DataStore(store_name=name, store_type=entite, store_code=code).report(force=force, data_only=data_only, ftp_push=ftp_push)


def report_epci(epci_id: str = "200039915", force=True, with_communes=False, data_only : bool = False, ftp_push : bool = False):
    """ Generates Report for a EPCI identified by Code INSEE  """
    entite = entite_epci
    name   = nom_epci(epci_id, clean=True)
    code   = epci_id
    if ((not name) or (str(name).startswith("Pas"))):
        print_red("EPCI non trouvee pour code :  [" + str(epci_id) + "]")
        return None
    if (with_communes):
        comm_len = len(communes_epci(epci_id))
        comm_idx = 0
        for commune in communes_epci(epci_id):
            comm_idx = comm_idx + 1
            print_grey("###> Commune " + str(comm_idx) + "/" + str(comm_len))
            report_commune(code_insee=str(commune), force=force, data_only=data_only, ftp_push=ftp_push)
    return DataStore(store_name=name, store_type=entite, store_code=code).report(force=force, data_only=data_only, ftp_push=ftp_push)


def report_dept(dept_id: str = "06", force=True, with_communes=False, data_only : bool = False, ftp_push : bool = False):
    """ Generates Report for a Departement identified by Code INSEE  """
    entite = entite_dept
    name   = nom_dept(dept_id, clean=True)
    code   = dept_id
    if ((not name) or (str(name).startswith("Pas"))):
        print_red("DEPT non trouve pour code :  [" + str(dept_id) + "]")
        return None
    if (with_communes):
        comm_len = len(communes_dept(dept_id))
        comm_idx = 0
        epci_len = len(epci_dept(dept_id))
        epci_idx = 0
        zone_len = len(list_zones_dept(dept_id))
        zone_idx = 0
        for commune in communes_dept(dept_id):
            comm_idx = comm_idx + 1
            print_grey("###> Commune " + str(comm_idx) + "/" + str(comm_len) + " # EPCI " + str(epci_idx) + "/" + str(epci_len) + " # Zone " + str(zone_idx) + "/" + str(zone_len))
            report_commune(code_insee=str(commune), force=force, data_only=data_only, ftp_push=ftp_push)
        for epci in epci_dept(dept_id):
            epci_idx = epci_idx + 1
            print_grey("###> Commune " + str(comm_idx) + "/" + str(comm_len) + " # EPCI " + str(epci_idx) + "/" + str(epci_len) + " # Zone " + str(zone_idx) + "/" + str(zone_len))
            report_epci(epci_id=str(epci), force=force, with_communes=False, data_only=data_only, ftp_push=ftp_push)
        for zone in list_zones_dept(dept_id):
            zone_idx = zone_idx + 1
            print_grey("###> Commune " + str(comm_idx) + "/" + str(comm_len) + " # EPCI " + str(epci_idx) + "/" + str(epci_len) + " # Zone " + str(zone_idx) + "/" + str(zone_len))
            report_zone(zone_name=str(zone), force=force, with_communes=False, data_only=data_only, ftp_push=ftp_push)
    return DataStore(store_name=name, store_type=entite, store_code=code).report(force=force, data_only=data_only, ftp_push=ftp_push)


def report_region(reg_id: str = "93", force=True, with_communes=False, data_only : bool = False, ftp_push : bool = False):
    """ Generates Report for a Region identified by Code INSEE  """
    entite = "REGION"
    name   = nom_region(reg_id, clean=True)
    code   = reg_id
    if ((not name) or (str(name).startswith("Pas"))):
        print_red("REGION non trouvee for pour code :  [" + str(reg_id) + "]")
        return None
    if (with_communes):
        for dept in list_dept(reg_id):
            report_dept(str(dept), force, with_communes=True, data_only=data_only, ftp_push=ftp_push)
        for commune in communes_region(reg_id):
            report_commune(code_insee=str(commune), force=force, data_only=data_only, ftp_push=ftp_push)
    return DataStore(store_name=name, store_type=entite, store_code=code).report(force=force, data_only=data_only, ftp_push=ftp_push)


def report_paca(force=True, data_only : bool = False, ftp_push : bool = False):
    """ Generates Report for PACA Region  """
    print_yellow("DEPT 06 - Alpes-Maritimes : ")
    for epci in epci_dept("06"):
        report_epci(str(epci), force=force, data_only=data_only, ftp_push=ftp_push)
    report_dept("06", force=force, data_only=data_only, ftp_push=ftp_push)

    print_yellow("DEPT 04 - Alpes-de-Haute-Provence : ")
    for epci in epci_dept("04"):
        report_epci(str(epci), force=force, data_only=data_only, ftp_push=ftp_push)
    report_dept("04", force=force, data_only=data_only, ftp_push=ftp_push)

    print_yellow("DEPT 05 - Hautes-Alpes : ")
    for epci in epci_dept("05"):
        report_epci(str(epci), force=force, data_only=data_only, ftp_push=ftp_push)
    report_dept("05", force=force, data_only=data_only, ftp_push=ftp_push)

    print_yellow("DEPT 13 - Bouches-du-Rhone : ")
    for epci in epci_dept("13"):
        report_epci(str(epci), force=force, data_only=data_only, ftp_push=ftp_push)
    report_dept("13", force=force, data_only=data_only, ftp_push=ftp_push)

    print_yellow("DEPT 83 - Var : ")
    for epci in epci_dept("83"):
        report_epci(str(epci), force=force, data_only=data_only, ftp_push=ftp_push)
    report_dept("83", force=force, data_only=data_only, ftp_push=ftp_push)

    print_yellow("DEPT 84 - Vaucluse : ")
    for epci in epci_dept("84"):
        report_epci(str(epci), force=force, data_only=data_only, ftp_push=ftp_push)
    report_dept("84", force=force, data_only=data_only, ftp_push=ftp_push)

    for zone in list_zones():
        report_zone(str(zone), force=force, data_only=data_only, ftp_push=ftp_push)

    print_yellow("REGION 93 - Region PACA : ")
    report_region("93", force=force, data_only=data_only, ftp_push=ftp_push)


report_select = {}

"""
Region
    List of Communes
    List of ECPI
        [Communes]
    List of Dept
        [Communes]
        [ECPIS]
"""


def report_select_dict(region=None, filename=None, force=False) -> dict:
    global report_select
    if str(region) in report_select: return report_select[str(region)]

    if ((force == False) and (os.path.isfile(filename))):
        print_yellow("+ Reading  Select Index : " + filename)
        with open(filename, "r") as read_file:
            print_grey("Converting JSON encoded data into Python dictionary : "+filename)
            select = jsonc.load(read_file)
            report_select[str(region)] = select
            return select
    print_yellow("+ Creating Select Index : " + filename)
    select = {}
    select["REGIONS"] = []
    if (not region):
        lr = list_region()
    else:
        lr = [str(region)]
    for r in lr:
        rd = {}
        rd["TYPE"]   = "REGION"
        rd["INSEE"]  = str(r)
        rd["Nom"]    = nom_region(r, clean=False)
        rd["Clean"]  = "REGION_" + nom_region(r, clean=True) + "_" + rd["INSEE"]
        rd["Region"] = str(r)
        rd["DEPARTEMENTS"] = []
        rd["DEPARTEMENTS_CODES"] = list_dept(r)
        rd["EPCIS"] = []
        rd["EPCIS_CODES"] = epci_region(r)
        rd["ZONES_CODES"] = list_zones()
        rd["COMMUNES"] = []
        rd["COMMUNES_CODES"] = communes_region(r)
        for c in rd["COMMUNES_CODES"]  :
            cd = {}
            cd["TYPE"]     = "COMMUNE"
            cd["INSEE"]    = str(c)
            pos, lib       = get_code_postal_commune(c)
            cd["Postal"]   = pos
            cd["Libelle"]  = lib
            cd["Nom"]      = nom_commune(code_insee=c, clean=False)
            cd["Clean"]    = "COMMUNE_" + nom_commune(code_insee=c, clean=True) + "_" + cd["INSEE"]
            rd["COMMUNES"].append(cd)
        for d in rd["DEPARTEMENTS_CODES"]:
            dd = {}
            dd["TYPE"]     = "DEPT"
            dd["INSEE"]    = str(d)
            dd["Nom"]      = nom_dept(d, clean=False)
            dd["Clean"]    = "DEPT_" + nom_dept(d, clean=True) + "_" + dd["INSEE"]
            dd["COMMUNES_CODES"] = communes_dept(d)
            dd["EPCIS_CODES"]    = epci_dept(d)
            dd["ZONES_CODES"]    = list_zones_dept(str(d))
            rd["DEPARTEMENTS"].append(dd)
        for e in rd["EPCIS_CODES"]:
            de = {}
            de["TYPE"]     = "EPCI"
            de["INSEE"]    = str(e)
            de["Nom"]      = nom_epci(e, clean=False)
            de["Clean"]    = "EPCI_" + nom_epci(e, clean=True) + "_" + de["INSEE"]
            de["COMMUNES_CODES"] = communes_epci(e)
            rd["EPCIS"].append(de)
        scot_data = scot_consolidation()
        rd["ZONES"] = []
        rd["ZONES_CODES"] = []
        for zone_name in scot_data["GROUPES_COMMUNES"]:
            zone = {}
            zone["TYPE"]           = scot_data["GROUPES_COMMUNES"][zone_name]["TYPE"]
            zone["Nom"]            = scot_data["GROUPES_COMMUNES"][zone_name]["NAME"]
            zone["INSEE"]          = clean_name(zone["Nom"])
            zone["Clean"]          = "ZONE_" + clean_name(zone["Nom"])
            zone["COMMUNES_CODES"] = scot_data["GROUPES_COMMUNES"][zone_name]["COMMUNES"]
            rd["ZONES"].append(zone)
            rd["ZONES_CODES"].append(zone["Clean"])
    select["REGIONS"].append(rd)
    if (filename):
        save_file(to_json(select, indent=4), filename)
        print_yellow("+ Saved    Select Index : " + filename)
    report_select[str(region)] = select
    return select


report_france = {}


def report_region_dict(region=None, filename=None, force=False) -> dict:
    global report_france
    if str(region) in report_france : return report_france[str(region)]

    if ((force == False) and (os.path.isfile(filename))):
        print_yellow("+ Reading  Region Index : " + filename)
        with open(filename, "r") as read_file:
            print_grey("Converting JSON encoded data into Python dictionary : " + filename)
            france = jsonc.load(read_file)
            report_france[str(region)] = france
            return france
    print_yellow("+ Creating Region Index : " + filename)
    france = {}
    france["REGIONS"] = []
    if (not region):
        lr = list_region()
    else:
        lr = [str(region)]
    for r in lr:
        rd = {}
        rd["TYPE"]   = "REGION"
        rd["INSEE"]  = str(r)
        rd["Nom"]    = nom_region(r, clean=False)
        rd["Clean"]  = "REGION_" + nom_region(r, clean=True)
        rd["Region"] = str(r)
        rd["DEPARTEMENTS"] = []
        france["REGIONS"].append(rd)
        for d in list_dept(r):
            dd = {}
            dd["TYPE"] = "DEPT"
            dd["INSEE"] = str(d)
            dd["Nom"]   = nom_dept(d, clean=False)
            dd["Clean"] = "DEPT_" + nom_dept(d, clean=True)
            dd["Departement"] = str(d)
            dd["Region"] = str(r)
            dd["Nom_Region"] = nom_region(r, clean=True)
            dd["EPCI"] = []
            dd["COMMUNES"] = []
            rd["DEPARTEMENTS"].append(dd)
            for e in epci_dept(d):
                de = {}
                de["TYPE"]   = "EPCI"
                de["INSEE"]  = str(e)
                de["Nom"]    = nom_epci(e,   clean=False)
                de["Clean"]  = "EPCI_" + nom_epci(e, clean=True)
                de["Departement"]     = str(d)
                de["Nom_Departement"] = nom_dept(d, clean=False)
                de["Region"]          = str(r)
                de["Nom_Region"]      = nom_region(r, clean=False)
                de["COMMUNES"]        = []
                dd["EPCI"].append(de)
                for c in communes_epci(e):
                    cd = {}
                    cd["TYPE"]    = "COMMUNE"
                    cd["INSEE"]   = str(c)
                    pos, lib      = get_code_postal_commune(c)
                    cd["Postal"]  = pos
                    cd["Libelle"] = lib
                    cd["Nom"]     = nom_commune(code_insee=c, clean=False)
                    cd["Clean"]   = "COMMUNE_" + nom_commune(code_insee=c, clean=True)
                    cd["Departement"]     = str(d)
                    cd["Nom_Departement"] = nom_dept(d, clean=False)
                    cd["Region"]          = str(r)
                    cd["Nom_Region"]      = nom_region(r, clean=False)
                    cd["EPCI"]            = epci_commune(c)
                    cd["Nom_EPCI"]        = nom_epci(epci_commune(c))
                    de["COMMUNES"].append(cd)
            for c in communes_dept(d):
                cd = {}
                cd["TYPE"]    = "COMMUNE"
                cd["INSEE"]   = str(c)
                pos, lib      = get_code_postal_commune(c)
                cd["Postal"]  = pos
                cd["Libelle"] = lib 
                cd["Nom"]     = nom_commune(code_insee=c,   clean=False)
                cd["Clean"]   = "COMMUNE_" + nom_commune(code_insee=c, clean=True)
                cd["Departement"]     = str(d)
                cd["Nom_Departement"] = nom_dept(d, clean=False)
                cd["Region"]          = str(r)
                cd["Nom_Region"]      = nom_region(r, clean=False)
                cd["EPCI"]            = epci_commune(c)
                cd["Nom_EPCI"]        = nom_epci(epci_commune(c))
                dd["COMMUNES"].append(cd)
    if (filename):
        save_file(to_json(france, indent=4),  filename)
        print_yellow("+ Saved    Region Index : " + filename)
    report_france[str(region)] = france
    return france


def load_min_data():
    load_departements()
    load_codes()
    load_interco()
    load_scot_data()
    load_collectData()


def load_all_data():
    load_min_data()
    load_sru(sruFile)
    load_communes()
    load_projections()
    load_projections_paca()
    load_sitadel()
    load_sitadel_locaux()
    load_evolution()
    load_artificialisation()


def fast():
    load_collectData()


def report_meta():
    load_collectData()
    ftp_push_files()


class TestConsommation(unittest.TestCase):

    def setUp(self) -> None:
        global DISPLAY_HTML
        DISPLAY_HTML = False
        print_red("> Setup")
        load_min_data()
        print_red("< Setup")

    def testScot_Ouest(self):
        print_yellow("> Scot Ouest")
        scot_ouest(code_insee="06108", start_date="2020-01-01") # 06108 / 06085
        scot_ouest(code_insee="06085", start_date="2020-01-01") # 06108 / 06085
        print_yellow("< Scot Ouest")

    def testSaintTropez(self):
        print_yellow("> Saint-Tropez")
        ds = report_commune(code_postal="83990", force=False)
        self.assertEqual(ds.get("NOM_COMMUNE"), "ST TROPEZ")
        print_yellow("< Saint-Tropez")

    def testMougins(self):
        print_yellow("> Mougins")
        scot_consolidation()
        ds = report_commune(code_postal="06250", force=True)
        self.assertEqual(ds.get("NOM_COMMUNE"), "MOUGINS")
        print_yellow("< Mougins")

    def testAntibes(self):
        print_yellow("> Antibes")
        scot_consolidation()
        ds = report_commune(code_postal="06600", force=True)
        print_yellow("< Antibes")

    def testMouginsReport(self):
        global DISPLAY_HTML
        DISPLAY_HTML = True
        self.testMougins()

    def testCannesReport(self):
        global DISPLAY_HTML
        DISPLAY_HTML = True
        print_yellow("> Cannes")
        ds = report_commune(code_postal="06400", force=False)
        self.assertEqual(ds.get("NOM_COMMUNE"), "CANNES")
        print_yellow("< Cannes")

    def testCAPL(self):
        global FAST
        FAST = True
        sel = report_select_dict("93", filename=selection_file, force=True)
        print_yellow("> CA Cannes Pays de Lerins")
        ds = report_epci(epci_id="200039915", force=True, with_communes=True)
        self.assertEqual(str(ds.get("EPCI")), "200039915")
        self.assertEqual(str(ds.get("NOM_COMMUNE")), "CANNES, LE CANNET, MANDELIEU LA NAPOULE, MOUGINS, THEOULE SUR MER")
        print_yellow("< CA Cannes Pays de Lerins")

    def testSCoT_Ouest(self):
        global FAST
        FAST = True
        sel = report_select_dict("93", filename=selection_file, force=True)
        print_yellow("> SCoT Ouest")
        ds = report_zone(zone_name="SCoT_Ouest", force=True, with_communes=True)
        self.assertEqual(str(ds.get("EPCI")), "SCoT_Ouest")
        print_yellow("< SCoT Ouest")

    def testZone(self):
        global FAST
        FAST = True
        report_select_dict("93", filename=selection_file, force=True)
        lz = list_zones()
        self.assertIn("SCoT_Ouest_Littoral", lz)
        self.assertIn("Zone_Menton", lz)
        self.assertIn("Zone_04", lz)
        lz06 = list_zones_dept("06")
        self.assertIn("SCoT_Ouest_Littoral", lz06)
        self.assertIn("Zone_Menton", lz06)
        self.assertNotIn("Zone_04", lz06)
        self.assertEqual(nom_zone("Zone_Menton"), "Zone Menton")
        self.assertEqual(dept_zone("Zone_Menton"), "06")
        lz04 = list_zones_dept("04")
        self.assertNotIn("Zone_Menton", lz04)
        self.assertIn("Zone_04", lz04)

        lz06 = list_nom_zone(list_zones_dept("06"))
        self.assertIn("SCoT Ouest Littoral", lz06)
        self.assertIn("Zone Menton", lz06)
        self.assertNotIn("Zone 04", lz06)

        cz06 = communes_zone("SCoT_Ouest_Littoral")
        self.assertIn("06004", cz06)
        self.assertIn("06085", cz06)
        self.assertNotIn("06000", cz06)

        strl = str_list(communes_zone("SCoT_Ouest_Littoral"))
        self.assertEqual(strl, "06004, 06085")

        print_yellow("> SCoT Ouest Littoral")
        ds  = report_zone(zone_name="SCoT_Ouest_Littoral", force=True, with_communes=True)
        self.assertEqual(str(ds.get("EPCI")), "SCoT_Ouest_Littoral")
        self.assertEqual(str(ds.get("NOM_COMMUNE")), "ANTIBES, MOUGINS")
        dpt = dept_zone("SCoT_Ouest_Littoral")
        self.assertEqual(str(dpt), "06")
        print_yellow("< SCoT Ouest Littoral")

    def testDataListZone(self):
        global FAST
        FAST = True
        sel = report_select_dict("93", filename=selection_file, force=True)
        lz = list_zones_dept("06")
        dpt = dept_zone("SCoT_Ouest_Littoral")
        self.assertEqual(str(dpt), "06")
        print_yellow("< SCoT Ouest Littoral")

    def testCAPLReport(self):
        global DISPLAY_HTML
        DISPLAY_HTML = True
        self.testCAPL()

    def testAlpesMaritimes(self):
        global FAST
        FAST = True
        sel = report_select_dict("93", filename=selection_file, force=True)
        print_yellow("> Departement Alpes Maritimes")
        ds = report_dept(dept_id="06", force=False, with_communes=True)
        self.assertEqual(str(ds.get("NOM_COMMUNE")), "Alpes-Maritimes")
        print_yellow("< Departement Alpes Maritimes")

    def testAlpesMaritimesReport(self):
        global DISPLAY_HTML
        DISPLAY_HTML = True
        self.testAlpesMaritimes()

    def testPaca(self):
        print_yellow("> Region Provence Alpes Cote d'Azur")
        ds = report_region(reg_id="93", force=False)
        self.assertEqual(str(ds.get("NOM_COMMUNE")), "946")
        print_yellow("< Region Provence Alpes Cote d'Azur")

    def testSelectReportRegion(self):
        selection_data = report_select_dict("93", filename=selection_file, force=True)
        print_yellow(to_json(selection_data, indent=4))
        region_data = report_region_dict("93", filename=france_file, force=True)
        print_blue(to_json(region_data, indent=4))

    def test_render_index(self):
        global DISPLAY_HTML
        DISPLAY_HTML = True
        html_index = gen_index()
        display_in_browser(html_index)

    def testPacaReport(self):
        global DISPLAY_HTML
        DISPLAY_HTML = True
        self.testPaca()

    def testReportPaca(self):
        load_all_data()
        report_paca()

    def testFTP_Push_Files(self):
        global FAST
        FAST = False
        ftp_push_files()

    def testFTP_Push_Prod(self):
        global FAST, FTP_PROD
        FAST = False
        FTP_PROD = True
        ftp_push_files()

    def testData(self):
        load_min_data()
        print_yellow("> Liste des Departements")
        the_list = list_dept()
        print(the_list)
        self.assertIn("06", the_list)
        self.assertNotIn("200004802", the_list)

        print_yellow("> Liste des Regions")
        the_list = list_region()
        print(the_list)
        self.assertIn("93", the_list)

        print_yellow("> Liste EPCI Region 93")
        the_list = epci_region("93")
        print(the_list)
        self.assertIn("200004802", the_list)

        print_yellow("> Liste EPCI Departement 06")
        the_list = epci_dept("06")
        print(the_list)
        self.assertIn("200039915", the_list)
        self.assertNotIn("200035723", the_list)

        print_yellow("> Liste Communes EPCI  200039915")
        the_list = communes_epci("200039915")
        print(the_list)
        self.assertIn("06030",    the_list)
        self.assertNotIn("04022", the_list)
        dept = dept_epci("200039915")
        self.assertEqual("06",    dept)

        print_yellow("> Liste Communes EPCI  200039915")
        the_list = list_nom_communes(communes_epci("200039915"))
        print(the_list)
        self.assertIn("Mougins",   the_list)
        self.assertNotIn("Grasse", the_list)

        print_yellow("> Liste Communes Departement  06")
        print(communes_dept("06"))
        self.assertIn("06030", communes_dept("06"))
        self.assertNotIn("04022", communes_epci("200039915"))

        print_yellow("> Liste Communes Region  93")
        print(communes_region("93"))
        self.assertIn("06030", communes_region("93"))
        self.assertIn("04022", communes_region("93"))

        print_yellow("> Nom Region 93")
        nom = nom_region("93", clean=True)
        print(nom)
        self.assertEqual("Provence-Alpes-Cote_d_Azur", nom)
        nom = nom_region("93", clean=False)
        print(nom)
        self.assertEqual("Provence-Alpes-Côte d'Azur", nom)

        print_yellow("> Nom EPCI 200039915")
        nom = nom_epci("200039915", clean=True)
        print(nom)
        self.assertEqual("CA_Cannes_Pays_de_Lerins", nom)
        nom = nom_epci("200039915", clean=False)
        print(nom)
        self.assertEqual("CA Cannes Pays de Lérins", nom)

        print_yellow("> Nom Departement 06")
        nom = nom_dept("06", clean=True)
        print(nom)
        self.assertEqual("Alpes-Maritimes", nom)
        nom = nom_dept("06", clean=False)
        print(nom)
        self.assertEqual("Alpes-Maritimes", nom)

        print_yellow("> Nom Commune 06250")
        nom = nom_commune(code_postal="06250", clean=False)
        print(nom)
        self.assertEqual("Mougins", nom)
        nom = nom_commune(code_postal="06250", clean=True)
        print(nom)
        self.assertEqual("Mougins", nom)
        nom = nom_commune(code_insee="06085", clean=False)
        print(nom)
        self.assertEqual("Mougins", nom)
        nom = nom_commune(code_insee="06085", clean=True)
        print(nom)
        self.assertEqual("Mougins", nom)
        ecpi = epci_commune("06085")
        print(str(ecpi))
        self.assertEqual("200039915", ecpi)
        lat, long = get_gps_insee("06085")
        print(str(lat))
        print(str(long))
        self.assertEqual("43.5961410556", lat)
        self.assertEqual("7.00129444919", long)
        lat = get_gps_lat_insee("06085")
        print(str(lat))
        self.assertEqual("43.5961410556", lat)
        long = get_gps_long_insee("06085")
        print(str(long))
        self.assertEqual("7.00129444919", long)
        long = get_gps_long_insee("06222085")
        print(str(long))
        self.assertEqual("", long)

    def testPlots(self):
        ds = report_commune(code_postal="06250", force=False)
        plots(ds)

    def testExcelFlux(self):
        excel_flux()


    def testExcel(self):
        file_123 = "123"+".xlsx"
        sheet_123 = "1234"
        sheet_sum = "Summary"


        if (os.path.exists(file_123)) :
            work_book = load_workbook(file_123)
        else:
            work_book = Workbook()
        if (sheet_123 in work_book.sheetnames):
            work_book.remove(work_book[sheet_123])
        work_sheet = work_book.create_sheet(0)
        work_sheet.title = sheet_123

        # Data
        df_123 = pd.DataFrame([[11, 21, 31, 41], [12, 22, 32, 42], [31, 32, 33, 34]], index=['one', 'two', 'three'], columns=['a', 'b', 'c', 'd'])
        for row in dataframe_to_rows(df_123, index=True, header=True):
            work_sheet.append(row)

        red_color = 'ffc7ce'
        red_color_font = '9c0103'
        red_font = styles.Font(size=14, bold=True, color=red_color_font)
        red_fill = styles.PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')
        redFill = styles.PatternFill(start_color='EE1111', end_color='EE1111', fill_type='solid')

        work_sheet.conditional_formatting.add('C2:C10', formatting.rule.CellIsRule(operator='lessThan', formula=['20'], stopIfTrue=True, fill=redFill))
        work_sheet.conditional_formatting.add('B1:B10', formatting.rule.CellIsRule(operator='lessThan', formula=['20'], fill=red_fill, font=red_font))
        work_sheet.conditional_formatting.add('D1:D10', formatting.rule.CellIsRule(operator='greaterThan', formula=['20'], fill=red_fill))

        # Summary
        try :
            df_sum = pd.read_excel(file_123, sheet_name=sheet_sum)
        except :
            df_sum = pd.DataFrame()
        df_sum.loc["MOUGINS","DATA"] = "123"
        df_sum.loc["MOUGINS","DATA2"] = "12"

        if (sheet_sum in work_book.sheetnames):
            work_book.remove(work_book[sheet_sum])
        work_sheet = work_book.create_sheet(0)
        work_sheet.title = sheet_sum

        for row in dataframe_to_rows(df_sum, index=True, header=True):
            work_sheet.append(row)


        work_book.save(filename=file_123)


    def testMeta(self):
        load_collectData()

    def testCalc(self):
        self.assertEqual("0",    round0str(0,   rounding=0))
        self.assertEqual("1",    round0str(1,   rounding=0))
        self.assertEqual("1",    round0str(1.1, rounding=0))
        self.assertEqual("0",    round0str(0.2, rounding=0))
        self.assertEqual("0.0",  round0str(0,   rounding=1))
        self.assertEqual("1.0",  round0str(1,   rounding=1))
        self.assertEqual("1.1",  round0str(1.1, rounding=1))
        self.assertEqual("0.2",  round0str(0.2, rounding=1))
        self.assertEqual("1.11", round0str(1.1123, rounding=2))
        self.assertEqual("0.30", round0str(0.2993, rounding=2))
        self.assertEqual("20%",   perCentStr(0.2,    rounding=0))
        self.assertEqual("1.6%",  perCentStr(0.0156, rounding=1))
        self.assertEqual("1.56%", perCentStr(0.0156, rounding=2))
        self.assertEqual(0,      taux(24, 100, rounding=0))
        self.assertEqual(0.2,    taux(24, 100, rounding=1))
        self.assertEqual(0.24,   taux(24, 100, rounding=2))
        self.assertEqual(0.12,   taux(24, 200, rounding=2))

        self.assertEqual(100.1,   calc_after(2000, 100, 2001, 0.1,  rounding=3))
        self.assertEqual(100.2,   calc_after(2000, 100, 2002, 0.1,  rounding=3))
        self.assertEqual(100.3,   calc_after(2000, 100, 2003, 0.1,  rounding=3))
        self.assertEqual(100.2,   calc_after(2000, 100, 2020, 0.01, rounding=3))

        self.assertEqual(0.01,    calc_taux(2000, 100, 2020, 100.2, rounding=3))

    def testSRUData(self):
        self.assertEqual(25,      get_sru2017("Taux de LLS à atteindre", "06085", rounding=2))
        self.assertEqual(8909,    get_sru2020("NBR RP au 01/01/2020",    "06085", rounding=0))

    def testSRUArtificialisation(self):
        self.assertEqual(1589,    get_art("pop1217", "06085", rounding=0))


###
### Command Line Arguments
###

# Arg Options
DISPLAY_HTML       = False
FORCE              = False
CONFIGURATION_FILE = None
TEMPLATE_FILE      = None
CODE_COMMUNE       = None
CODE_EPCI          = None
CODE_DEPT          = None
CODE_REGION        = None
DEBUG              = False
WITH_COMMUNES      = False
LIST_COMMUNE       = False
DATA_ONLY          = False
FTP_PUSH           = False
FTP_PROD           = False
FAST               = False
VERBOSE            = False


def read_command_line_args(argv):
    global DISPLAY_HTML, FORCE, DEBUG, WITH_COMMUNES, LIST_COMMUNE
    global CONFIGURATION_FILE, TEMPLATE_FILE, DATA_ONLY, FTP_PUSH, FTP_PROD, FAST
    global CODE_COMMUNE, CODE_EPCI, CODE_DEPT, CODE_REGION, VERBOSE
    # print_yellow("Command Line Arguments : " + str(argv))

    usage = """
    Usage: -v -f -a -p -n -b -t -m -l -c -t <commune_code> -e <epci_code> -d <dept_code> -r <region_code> -z <zone_name>    
           -l --list        : List for all communes/epci/dept/zone in Territory       
           -c --commune <c> : Report for Commune Code INSEE 'c'                 
           -e --ecpi    <e> : Report for ECPI Code INSEE 'e'                      
           -d --dept    <d> : Report for Departement Code INSEE 'd'                 
           -r --region  <r> : Report for Region Code INSEE 'r'          
           -z --zone    <z> : Report for Zone or Scot 'z'          
           -a --all         : Report for all communes in Territory           
           -n --data        : No report - Generate only data & Graphics        
           -p --push        : FTP Push Data to Infinity Free Host WebSite         
           -s --prod        : FTP Push Data to Production Sobriete Fonciere         
           -f --force       : Report reading source data (cache ignored)     
           -t --fast        : FasT (Only Data Refresh during Dev)     
           -m --meta        : Meta (Only Meta Refresh during Dev)     
           -v --verbose     : Verbose     
           --browse         : Start Browser on generated report (debug)  
           --cxlsx            <ConfigurationFile.xlsx> : Use Configuration File  
           --rhtml            <ReportTemplate.html>    : Use ReportTemplate      
           --clean          : Delete Report files      
    """

    try:
        opts, args = getopt.getopt(argv, "hanlbmpstfc:e:d:r:z:", ["help", "list", "data" , "fast" , "push" , "prod" , "clean" , "meta" , "commune=", "epci=", "dep=", "reg=", "zone=", "no_debug"])
    except getopt.GetoptError:
        print(usage)
        sys.exit(2)
    for opt, arg in opts:
        if opt in ("-h", "--help"):
            print(usage)
            quit()
        elif (opt == "--debug"):
            DEBUG = True
            continue
        elif opt in ("-b", "-B"):
            DISPLAY_HTML = True
            continue
        elif opt in ("-p", "-P", "-push", "-PUSH", "-ftp", "-FTP"):
            FTP_PUSH = True
            continue
        elif opt in ("-s", "-S", "-prod", "-PROD", "-sobriete", "-SOBRIETE"):
            FTP_PROD = True
            continue
        elif opt in ("-v", "-V", "-verbose", "-VERBOSE"):
            VERBOSE = True
            continue
        elif opt in ("-t", "-T", "-fast", "-FAST"):
            FAST = True
            continue
        elif opt in ("-f", "-F"):
            FORCE = True
            continue
        elif opt in ("-a", "-A", "-all", "-All", "-ALL"):
            WITH_COMMUNES = True
            continue
        elif opt in ("-n", "-N", "-data", "-Data", "-DATA"):
            DATA_ONLY = True
            continue
        elif opt in ("-b", "--browse"):
            DISPLAY_HTML = True
            continue
        elif opt in ("-m", "--meta"):
            print_yellow("> Meta ")
            report_meta()
            print_yellow("< Meta ")
            quit()
        elif (opt == "--cxlsx"):
            if (not os.path.isfile(arg)):
                print_red("Configuration File not found : "+arg)
                quit()
            CONFIGURATION_FILE = arg
            continue
        elif (opt == "--clean"):
            delete_pattern(output_dir, "*.png")
            delete_pattern(output_dir, "*.csv")
            delete_pattern(output_dir, "*.log")
            delete_pattern(output_dir, "context.yaml")
            delete_pattern(output_dir, "france.json")
            delete_pattern(output_dir, "*_*.json")
            delete_pattern(output_dir, "*.xlsx")
            delete_pattern(output_dir, "*.html")
            quit()
        elif (opt == "--rhtml"):
            if (not os.path.isfile(arg)):
                print_red("Template File not found : "+arg)
                quit()
            TEMPLATE_FILE = arg
            continue
        elif opt in ("-l", "--list"):
            LIST_COMMUNE = True
            continue
        elif opt in ("-c", "--commune"):
            CODE_COMMUNE = arg
            print_yellow("> Commune "+str(CODE_COMMUNE))
            report_commune(code_insee=CODE_COMMUNE, force=FORCE, data_only=DATA_ONLY, ftp_push=FTP_PUSH)
            print_yellow("< Commune "+str(CODE_COMMUNE))
            quit()
        elif opt in ("-e", "--epci"):
            CODE_EPCI = arg
            if (LIST_COMMUNE):
                print_commune(communes_epci(CODE_EPCI))
            else:
                print_yellow("> EPCI " + str(CODE_EPCI))
                report_select_dict("93", filename=selection_file, force=True)
                if (not FAST) : report_region_dict("93", filename=france_file,    force=True)
                report_epci(CODE_EPCI, force=FORCE, with_communes=WITH_COMMUNES, data_only=DATA_ONLY, ftp_push=FTP_PUSH)
                print_yellow("< EPCI " + str(CODE_EPCI))
                quit()
        elif opt in ("-z", "--zone"):
                ZONE_NAME = arg
                if (LIST_COMMUNE):
                    print_commune(communes_zone(ZONE_NAME))
                else:
                    print_yellow("> Zone " + str(ZONE_NAME))
                    report_select_dict("93", filename=selection_file, force=True)
                    if (not FAST): report_region_dict("93", filename=france_file, force=True)
                    report_zone(ZONE_NAME, force=FORCE, with_communes=WITH_COMMUNES, data_only=DATA_ONLY,ftp_push=FTP_PUSH)
                    print_yellow("< Zone " + str(ZONE_NAME))
                    quit()
        elif opt in ("-d", "--dept"):
            CODE_DEPT = arg
            if (LIST_COMMUNE):
                print_commune(communes_dept(CODE_DEPT))
                print_epci(epci_dept(CODE_DEPT))
                quit()
            else:
                print_yellow("> Departement " + str(CODE_DEPT))
                report_select_dict("93", filename=selection_file, force=True)
                if (not FAST) : report_region_dict("93", filename=france_file,    force=True)
                report_dept(CODE_DEPT, force=FORCE, with_communes=WITH_COMMUNES, data_only=DATA_ONLY, ftp_push=FTP_PUSH)
                print_yellow("< Departement " + str(CODE_DEPT))
                quit()
        elif opt in ("-r", "--reg"):
            CODE_REGION = arg
            if (LIST_COMMUNE):
                print_commune(communes_region(CODE_REGION))
                print_epci(epci_region(CODE_REGION))
                print_dept(list_dept(CODE_REGION))
                quit()
            else:
                print_yellow("> Region " + str(CODE_REGION))
                report_select_dict("93",              filename=selection_file, force=True)
                if (not FAST) : report_region_dict(str(CODE_REGION),  filename=france_file,    force=True)
                report_region(CODE_REGION, force=FORCE, with_communes=WITH_COMMUNES, data_only=DATA_ONLY, ftp_push=FTP_PUSH)
                print_yellow("< Region " + str(CODE_REGION))
                quit()

###
### Main
###

read_command_line_args(argv=sys.argv[1:])

if __name__ == '__main__':
    print_yellow("Consommation Fonciere - Test Suite > " + __name__)
    load_min_data()
    unittest.main()

###
# https://app.infinityfree.net/accounts/epiz_30239961#
#
# FTP Username	epiz_30239961
# FTP Password  oqEwtTaACCaANF
# FTP Hostname	ftpupload.net
# FTP Port (optional)	21
###


###
# https://dillinger.io/
# http://consommationfonciere.infinityfreeapp.com/README.md
# Export As Styled HTML into README.dillinger.html
###
